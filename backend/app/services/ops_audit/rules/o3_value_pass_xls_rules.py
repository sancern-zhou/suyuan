"""O3 value-pass XLS attachment checks."""

from __future__ import annotations

import json
import os
import re
import subprocess
import tempfile
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import requests

from app.services.ops_audit.models import Issue
from app.services.ops_audit.rules.base import add_issue

RULE_ID = "ATTACHMENT_O3_VALUE_PASS_XLS_VALUE_MISMATCH"
MISSING_XLS_REVIEW_RULE_ID = "ATTACHMENT_O3_VALUE_PASS_XLS_MISSING_REVIEW"
HISTORY_CONFLICT_REVIEW_RULE_ID = "RF_O3_UPPER_STANDARD_HISTORY_CONFLICT_REVIEW"
FLOW_MISSING_RULE_ID = "RF_O3_VALUE_PASS_FLOW_VALUE_MISSING"
FIELD_POSITION_RULE_ID = "RF_O3_VALUE_PASS_FIELD_POSITION_SUSPECT"
RF_TABLE = "RF_HY_O3VALUEPASS"
XLSX_SCAN_MAX_ROWS = 200
XLSX_SCAN_MAX_COLUMNS = 8

COMPARISONS = [
    {"field": "DEVICEDELIVERMODEL", "label": "斜率", "cell": "F"},
    {"field": "DELIVERFC", "label": "截距(ppb)", "cell": "F"},
    {"field": "DENSITY1VALUE", "label": "相对于前一次传递的改变(%)", "cell": "F"},
]

UPPER_STANDARD_COMPARISONS = (
    {"field": "DELIVER6VALUE", "label": "上级标准型号", "comparison_type": "text"},
    {"field": "DELIVERFROM6VALUE", "label": "上级标准设备号", "comparison_type": "text"},
    {"field": "AVALUE", "label": "上级标准序列号", "comparison_type": "text"},
    {"field": "WORKDENSITY6VALUE", "label": "上级标准传递日期", "comparison_type": "date"},
    {"field": "DELIVERTO6VALUE", "label": "上级标准传递公式", "comparison_type": "formula"},
    {"field": "BVALUE", "label": "上级标准有效期", "comparison_type": "date"},
)

UPPER_STANDARD_SECTION_LABELS = ("上级臭氧传递标准", "参考光电仪")
UPPER_STANDARD_LABEL_FIELDS = {
    "型号": "DELIVER6VALUE",
    "设备号": "DELIVERFROM6VALUE",
    "序列号": "AVALUE",
    "传递日期": "WORKDENSITY6VALUE",
    "认证日期": "WORKDENSITY6VALUE",
    "传递公式": "DELIVERTO6VALUE",
    "认证公式": "DELIVERTO6VALUE",
    "传递有效期限": "BVALUE",
    "认证有效期限": "BVALUE",
}


def check_o3_value_pass_xls_values(
    order: dict[str, Any],
    forms: list[tuple[str, dict[str, Any]]],
    attachments: list[dict[str, Any]],
    wo_commonfiles: list[dict[str, Any]],
    issues: list[Issue],
    *,
    attachment_read_cache: dict[tuple[Any, ...], dict[str, Any]] | None = None,
) -> None:
    """Compare RF_HY_O3VALUEPASS values with the uploaded XLS first sheet."""

    relevant_forms = [form for table, form in forms if table == RF_TABLE and not form.get("_query_error")]
    if not relevant_forms:
        return

    records = attachments + wo_commonfiles
    xls_items = _xls_items(records)
    pdf_items = _pdf_items(records)
    read_cache = attachment_read_cache if attachment_read_cache is not None else {}
    for form in relevant_forms:
        _check_o3_value_pass_form_fields(order, form, issues)
        if not xls_items:
            _add_missing_xls_review_issue(order, form, records, issues)
        selected_xls = _select_matching_xls_item(form, xls_items, read_cache=read_cache)
        selected_items = [item for item in (selected_xls, _select_item(pdf_items)) if item]
        if not selected_items:
            continue
        for item in selected_items:
            cell_values = _read_cells_cached(item, read_cache)
            if cell_values.get("status") != "success":
                if item.get("attachment_kind") == "pdf" or _is_unavailable_attachment_source_error(cell_values.get("error")):
                    continue
                _add_issue(order, form, item, [{"status": "xls_read_error", "error": cell_values.get("error")}], issues)
                continue

            comparisons = _compare_values(form, cell_values["cells"])
            failed = [comparison for comparison in comparisons if comparison["status"] != "match"]
            if failed:
                _add_issue(order, form, item, failed, issues)


def build_o3_upper_standard_history_conflicts(
    forms_by_code: dict[str, list[tuple[str, dict[str, Any]]]],
    current_codes: set[str],
) -> dict[str, list[Issue]]:
    grouped: dict[
        tuple[Any, ...],
        dict[tuple[str, str], list[dict[str, Any]]],
    ] = defaultdict(lambda: defaultdict(list))

    for code, forms in forms_by_code.items():
        for table, form in forms:
            if table != RF_TABLE or form.get("_query_error"):
                continue
            fingerprint = _upper_standard_fingerprint(form)
            identity = _upper_standard_identity_pair(form)
            if fingerprint is None or identity is None:
                continue
            grouped[fingerprint][identity].append(
                {
                    "working_order_code": str(code),
                    "model": form.get("DELIVER6VALUE"),
                    "device_number": form.get("DELIVERFROM6VALUE"),
                    "serial_number": form.get("AVALUE"),
                    "transfer_date": form.get("WORKDENSITY6VALUE"),
                    "transfer_formula": form.get("DELIVERTO6VALUE"),
                    "expiry_date": form.get("BVALUE"),
                }
            )

    issues_by_code: dict[str, list[Issue]] = defaultdict(list)
    for fingerprint, alternatives_by_identity in grouped.items():
        if len(alternatives_by_identity) < 2:
            continue
        alternatives = []
        affected_codes = set()
        for records in alternatives_by_identity.values():
            order_codes = sorted({record["working_order_code"] for record in records})
            affected_codes.update(order_codes)
            alternatives.append(
                {
                    "model": records[0]["model"],
                    "device_number": records[0]["device_number"],
                    "order_codes": order_codes,
                }
            )
        alternatives.sort(key=lambda item: (str(item.get("model") or ""), str(item.get("device_number") or "")))
        sample_records = next(iter(alternatives_by_identity.values()))
        sample = sample_records[0]
        for code in sorted(affected_codes & current_codes):
            evidence = {
                "current_working_order_code": code,
                "rf_table": RF_TABLE,
                "needs_manual_review": True,
                "review_reason": "same_upper_standard_batch_has_conflicting_identity",
                "fingerprint": {
                    "serial_number": sample["serial_number"],
                    "transfer_date": sample["transfer_date"],
                    "transfer_formula": sample["transfer_formula"],
                    "expiry_date": sample["expiry_date"],
                    "normalized": list(fingerprint),
                },
                "alternatives": alternatives,
            }
            issue_list: list[Issue] = []
            add_issue(
                issue_list,
                HISTORY_CONFLICT_REVIEW_RULE_ID,
                "跨工单证据复核",
                "中",
                "rf.RF_HY_O3VALUEPASS.upper_standard_identity",
                "同一O3上级标准批次存在不同型号/设备号填法，需结合证书人工确认",
                json.dumps(evidence, ensure_ascii=False, default=str),
            )
            issues_by_code[code].extend(issue_list)
    return dict(issues_by_code)


def _upper_standard_fingerprint(form: dict[str, Any]) -> tuple[Any, ...] | None:
    serial_number = _normalize_identity_text(form.get("AVALUE"))
    if serial_number in {"", "NA", "NONE", "NULL"}:
        return None
    transfer_formula = _parse_linear_formula(form.get("DELIVERTO6VALUE"))
    transfer_date = _normalize_date_value(form.get("WORKDENSITY6VALUE"))
    expiry_date = _normalize_date_value(form.get("BVALUE"))
    if transfer_formula is None or len(transfer_date) < 3 or len(expiry_date) < 3:
        return None
    return (
        serial_number,
        round(transfer_formula[0], 8),
        round(transfer_formula[1], 8),
        transfer_date,
        expiry_date,
    )


def _upper_standard_identity_pair(form: dict[str, Any]) -> tuple[str, str] | None:
    model = _normalize_identity_text(form.get("DELIVER6VALUE"))
    if not model:
        return None
    raw_device = str(form.get("DELIVERFROM6VALUE") or "").strip()
    device = _normalize_identity_text(raw_device)
    if raw_device in {"", "/", "-", "\\"} or device in {"", "NA", "NONE", "NULL"}:
        device = "NO_DEVICE_NUMBER"
    return model, device


def _add_missing_xls_review_issue(
    order: dict[str, Any],
    form: dict[str, Any],
    records: list[dict[str, Any]],
    issues: list[Issue],
) -> None:
    upper_standard = {
        "model": form.get("DELIVER6VALUE"),
        "device_number": form.get("DELIVERFROM6VALUE"),
        "serial_number": form.get("AVALUE"),
        "transfer_date": form.get("WORKDENSITY6VALUE"),
        "transfer_formula": form.get("DELIVERTO6VALUE"),
        "expiry_date": form.get("BVALUE"),
    }
    available_attachments = []
    for record in records:
        sources = _source_candidates(record)
        available_attachments.append(
            {
                "filename": _first_present(record, ["FILENAME", "filename", "FileName", "NAME", "name"]),
                "source_path": sources[0] if sources else None,
                "source_paths": sources,
            }
        )
    evidence = {
        "working_order_code": order.get("WORKINGORDERCODE") or form.get("WORKINGORDERCODE"),
        "rf_table": RF_TABLE,
        "needs_manual_review": True,
        "review_reason": "missing_o3_transfer_workbook",
        "upper_standard": upper_standard,
        "available_attachments": available_attachments,
    }
    add_issue(
        issues,
        MISSING_XLS_REVIEW_RULE_ID,
        "附件证据复核",
        "中",
        f"attachment.{RF_TABLE}.xls",
        "O3量值传递工单缺少XLS/XLSX计算附件，当前传递斜率及上级标准身份需人工复核",
        json.dumps(evidence, ensure_ascii=False, default=str),
    )


def _check_o3_value_pass_form_fields(
    order: dict[str, Any],
    form: dict[str, Any],
    issues: list[Issue],
) -> None:
    missing = []
    position_suspects = []
    # In the current RF_HY_O3VALUEPASS template, groups 4-6 store metadata such
    # as dates, model, serial number, and formula fragments. Only groups 1-3
    # are stable enough for deterministic transfer-flow completeness checks.
    for point in range(1, 4):
        flow_field = f"DELIVER{point}VALUE"
        related_fields = [
            f"WORKDENSITY{point}VALUE",
            f"DELIVERFROM{point}VALUE",
            f"DELIVERTO{point}VALUE",
        ]
        raw_flow = form.get(flow_field)
        related_values = {field: form.get(field) for field in related_fields if field in form}
        if _is_low_value(raw_flow) and any(not _is_low_value(value) for value in related_values.values()):
            missing.append(
                {
                    "point": point,
                    "field": flow_field,
                    "value": raw_flow,
                    "related_values": related_values,
                }
            )
        if _looks_like_non_flow_value(raw_flow):
            position_suspects.append(
                {
                    "point": point,
                    "field": flow_field,
                    "value": raw_flow,
                    "suggested_check": "传递流量字段疑似填入日期或设备信息",
                }
            )
    if missing:
        _add_form_field_issue(
            order,
            form,
            FLOW_MISSING_RULE_ID,
            "表单完整性",
            "高",
            f"rf.{RF_TABLE}.{missing[0]['field']}",
            f"O3量值传递第{missing[0]['point']}组传递流量未填写",
            missing,
            issues,
        )
    if position_suspects:
        _add_form_field_issue(
            order,
            form,
            FIELD_POSITION_RULE_ID,
            "表单数值逻辑",
            "高",
            f"rf.{RF_TABLE}.{position_suspects[0]['field']}",
            "O3量值传递传递流量字段疑似填入日期或设备信息",
            position_suspects,
            issues,
        )


def _add_form_field_issue(
    order: dict[str, Any],
    form: dict[str, Any],
    rule_id: str,
    category: str,
    severity: str,
    field: str,
    message: str,
    violations: list[dict[str, Any]],
    issues: list[Issue],
) -> None:
    evidence = {
        "working_order_code": order.get("WORKINGORDERCODE") or form.get("WORKINGORDERCODE"),
        "rf_table": RF_TABLE,
        "violations": violations[:20],
    }
    add_issue(
        issues,
        rule_id,
        category,
        severity,
        field,
        message,
        json.dumps(evidence, ensure_ascii=False, default=str),
    )


def _is_low_value(value: Any) -> bool:
    return str(value or "").strip() in {"", "/", "-", "无", "NA", "N/A", "nan", "NaN", "null", "None"}


def _looks_like_non_flow_value(value: Any) -> bool:
    text = str(value or "").strip()
    if not text or _is_low_value(text):
        return False
    if re.search(r"\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}", text):
        return True
    upper = text.upper()
    return any(token in upper for token in ("TE-", "49I", "146I", "CM", "LGH", "MODEL", "????"))


def _xls_items(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items = []
    for record in records:
        name = _first_present(record, ["FILENAME", "filename", "FileName", "NAME", "name", "TITLE", "title"])
        sources = _source_candidates(record)
        source = sources[0] if sources else None
        if not _is_xls_attachment(name, sources):
            continue
        items.append({"filename": name, "source_path": source, "source_paths": sources, "raw": record})
    return items


def _pdf_items(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items = []
    for record in records:
        name = _first_present(record, ["FILENAME", "filename", "FileName", "NAME", "name", "TITLE", "title"])
        sources = _source_candidates(record)
        source = sources[0] if sources else None
        if not _is_pdf_attachment(name, sources):
            continue
        descriptor = " ".join(str(value or "") for value in [name, *sources]).lower()
        if "o3" not in descriptor and "臭氧" not in descriptor:
            continue
        items.append({"filename": name, "source_path": source, "source_paths": sources, "raw": record, "attachment_kind": "pdf"})
    return items


def _is_xls_attachment(name: Any, sources: list[str]) -> bool:
    for value in [name, *sources]:
        suffix = Path(urlparse(str(value or "")).path).suffix.lower()
        if suffix in {".xls", ".xlsx"}:
            return True
    return False


def _is_pdf_attachment(name: Any, sources: list[str]) -> bool:
    for value in [name, *sources]:
        suffix = Path(urlparse(str(value or "")).path).suffix.lower()
        if suffix == ".pdf":
            return True
    return False


def _select_item(items: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not items:
        return None
    return sorted(items, key=lambda item: str(item.get("filename") or item.get("source_path") or ""))[0]


def _select_matching_xls_item(
    form: dict[str, Any],
    items: list[dict[str, Any]],
    *,
    read_cache: dict[tuple[Any, ...], dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    """Select the XLS whose structure and values best represent the RF form.

    An O3 transfer order can contain both the operator's transfer worksheet and
    a supplier/authority report.  Filename ordering is not a reliable way to
    distinguish them (and previously selected the GDEEMC report in some
    orders).  Parse every candidate and rank by comparison quality first;
    filename hints are only a deterministic tie-breaker.
    """
    if not items:
        return None

    read_cache = read_cache if read_cache is not None else {}
    ranked: list[tuple[tuple[int, int, int, int, str], dict[str, Any]]] = []
    for item in items:
        read_result = _read_cells_cached(item, read_cache)
        if read_result.get("status") != "success":
            ranked.append(((0, 0, -1, _xls_filename_priority(item), _item_name(item)), item))
            continue
        comparisons = _compare_values(form, read_result.get("cells", {}))
        comparable = [comparison for comparison in comparisons if comparison.get("field") in {
            "DEVICEDELIVERMODEL", "DELIVERFC", "DENSITY1VALUE"
        }]
        matches = sum(comparison.get("status") == "match" for comparison in comparable)
        populated = sum(comparison.get("status") != "missing_xls_value" for comparison in comparable)
        problems = sum(comparison.get("status") not in {"match"} for comparison in comparable)
        # More complete/matching templates win.  A name hint is only used when
        # the parsed evidence is otherwise tied.
        key = (matches, populated, -problems, _xls_filename_priority(item), _item_name(item))
        ranked.append((key, item))
    return max(ranked, key=lambda pair: pair[0])[1]


def _item_name(item: dict[str, Any]) -> str:
    return str(item.get("filename") or item.get("source_path") or "")


def _xls_filename_priority(item: dict[str, Any]) -> int:
    """Use naming conventions only as a tie-breaker, never as the selector."""
    name = _item_name(item).lower()
    score = 0
    if "报告" in name or "gdeemc" in name:
        score -= 2
    if "标准传递" in name or "valuepass" in name:
        score += 1
    return score


def _read_cells(item: dict[str, Any]) -> dict[str, Any]:
    if item.get("attachment_kind") == "pdf":
        return _read_cells_from_pdf_item(item)
    sources = item.get("source_paths")
    if not isinstance(sources, list) or not sources:
        sources = [item.get("source_path")]
    errors = []
    for source_value in sources:
        source = str(source_value or "").strip()
        if not source:
            continue
        resolved = _resolve_source(source)
        if resolved.get("status") == "success":
            return _read_cells_from_path(Path(str(resolved["path"])))
        errors.append(f"{source}: {resolved.get('error')}")
    return {"status": "error", "error": "；".join(errors) or "附件路径为空"}


def _read_cells_cached(
    item: dict[str, Any],
    cache: dict[tuple[Any, ...], dict[str, Any]],
) -> dict[str, Any]:
    cache_key = _attachment_read_cache_key(item)
    if cache_key not in cache:
        cache[cache_key] = _read_cells(item)
    return cache[cache_key]


def _attachment_read_cache_key(item: dict[str, Any]) -> tuple[Any, ...]:
    sources = item.get("source_paths")
    if not isinstance(sources, list) or not sources:
        sources = [item.get("source_path")]
    source_keys = sorted(
        {
            _source_cache_identity(str(source or "").strip())
            for source in sources
            if str(source or "").strip()
        },
        key=repr,
    )
    if not source_keys:
        source_keys = [("filename", _item_name(item))]
    return (str(item.get("attachment_kind") or "xls"), *source_keys)


def _source_cache_identity(source: str) -> tuple[Any, ...]:
    parsed = urlparse(source)
    if parsed.scheme in {"http", "https"}:
        return "url", source
    path = Path(source).expanduser()
    if path.is_file():
        stat = path.stat()
        return "file", str(path.resolve()), stat.st_size, stat.st_mtime_ns
    return "source", source


def _read_cells_from_pdf_item(item: dict[str, Any]) -> dict[str, Any]:
    result = _read_pdf_text(item)
    if result.get("status") != "success":
        return result
    return {"status": "success", "cells": _cells_from_pdf_text(str(result.get("text") or ""))}


def _read_pdf_text(item: dict[str, Any]) -> dict[str, Any]:
    sources = item.get("source_paths")
    if not isinstance(sources, list) or not sources:
        sources = [item.get("source_path")]
    errors = []
    for source_value in sources:
        source = str(source_value or "").strip()
        if not source:
            continue
        resolved = _resolve_source(source)
        if resolved.get("status") != "success":
            errors.append(f"{source}: {resolved.get('error')}")
            continue
        path = Path(str(resolved["path"]))
        try:
            completed = subprocess.run(
                ["pdftotext", "-layout", str(path), "-"],
                check=True,
                capture_output=True,
                text=True,
                timeout=20,
            )
        except Exception as exc:
            errors.append(f"{source}: {exc}")
            continue
        return {"status": "success", "text": completed.stdout}
    return {"status": "error", "error": "；".join(errors) or "PDF附件路径为空"}


def _cells_from_pdf_text(text: str) -> dict[str, list[dict[str, Any]]]:
    cells = {
        "__source_type": "pdf_text",
        "DEVICEDELIVERMODEL": [],
        "DELIVERFC": [],
        "DENSITY1VALUE": [],
        "TRANSFER_FORMULA": [],
    }
    for index, line in enumerate(text.splitlines(), start=1):
        normalized = _normalize_label(line)
        if not normalized:
            continue
        if "斜率" in normalized and not cells["DEVICEDELIVERMODEL"]:
            cells["DEVICEDELIVERMODEL"].append({"cell": f"pdf:{index}", "value": _first_number_text(line)})
        if "截距" in normalized and not cells["DELIVERFC"]:
            cells["DELIVERFC"].append({"cell": f"pdf:{index}", "value": _first_number_text(line)})
        if "改变" in normalized and "前一次" in normalized:
            cells["DENSITY1VALUE"].append({"cell": f"pdf:{index}", "value": _first_number_text(line)})
        if "公式" in normalized and ("传递" in normalized or "拟合" in normalized or "线性" in normalized):
            cells["TRANSFER_FORMULA"].append({"cell": f"pdf:{index}", "value": line.strip()})
    return cells


def _first_number_text(text: str) -> str:
    match = re.search(r"[-+]?\d+(?:\.\d+)?%?", str(text or ""))
    return match.group(0) if match else ""


def _read_cells_from_path(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            wb = load_workbook(path, data_only=True, read_only=True)
            try:
                cells = _worksheet_dynamic_cells(wb.worksheets[0])
            finally:
                wb.close()
            return {"status": "success", "cells": cells}
        if suffix == ".xls":
            import xlrd

            book = xlrd.open_workbook(str(path))
            sheet = book.sheet_by_index(0)
            cells = _xls_sheet_dynamic_cells(sheet)
            return {"status": "success", "cells": cells}
    except Exception as exc:
        return {"status": "error", "error": str(exc)}
    return {"status": "error", "error": f"不支持的附件格式：{suffix}"}


def _source_candidates(record: dict[str, Any]) -> list[str]:
    candidates = []
    for field in (
        "file_url",
        "fileUrl",
        "FILEURL",
        "FILE_URL",
        "URL",
        "url",
        "FILEPATH",
        "filepath",
        "PATH",
        "path",
    ):
        value = record.get(field)
        if value is None:
            continue
        source = str(value).strip()
        if source and source not in candidates:
            candidates.append(source)
    return candidates


def _worksheet_dynamic_cells(ws: Any) -> dict[str, list[dict[str, Any]]]:
    row_count = min(int(ws.max_row or 0), XLSX_SCAN_MAX_ROWS)
    if row_count <= 0:
        return _tabular_dynamic_cells([])
    rows = list(
        ws.iter_rows(
            min_row=1,
            max_row=row_count,
            min_col=1,
            max_col=XLSX_SCAN_MAX_COLUMNS,
            values_only=True,
        )
    )
    return _tabular_dynamic_cells(rows)


def _tabular_dynamic_cells(rows: list[tuple[Any, ...]]) -> dict[str, list[dict[str, Any]]]:
    def value_at(row_index: int, column_index: int) -> Any:
        if row_index < 1 or row_index > len(rows):
            return None
        row = rows[row_index - 1]
        if column_index < 1 or column_index > len(row):
            return None
        return row[column_index - 1]

    slope_row, intercept_row = _find_o3_value_pass_rows(
        (row_index, value_at(row_index, 4)) for row_index in range(1, len(rows) + 1)
    )
    cells = _dynamic_value_cells(
        slope_row,
        intercept_row,
        lambda row_index: value_at(row_index, 6),
    )
    formula_row = _find_formula_row(
        (row_index, value_at(row_index, 4)) for row_index in range(1, len(rows) + 1)
    )
    if formula_row is not None:
        cells["TRANSFER_FORMULA"] = [{"cell": f"F{formula_row}", "value": value_at(formula_row, 6)}]
    cells.update(
        _upper_standard_cells(
            len(rows),
            XLSX_SCAN_MAX_COLUMNS,
            value_at,
        )
    )
    return cells


def _xls_sheet_dynamic_cells(sheet: Any) -> dict[str, list[dict[str, Any]]]:
    slope_row, intercept_row = _find_o3_value_pass_rows(
        (row_index + 1, sheet.cell_value(row_index, 3)) for row_index in range(sheet.nrows)
    )
    cells = _dynamic_value_cells(
        slope_row,
        intercept_row,
        lambda row_index: sheet.cell_value(row_index - 1, 5) if row_index <= sheet.nrows else None,
    )
    formula_row = _find_formula_row((row_index + 1, sheet.cell_value(row_index, 3)) for row_index in range(sheet.nrows))
    if formula_row is not None:
        cells["TRANSFER_FORMULA"] = [
            {"cell": f"F{formula_row}", "value": sheet.cell_value(formula_row - 1, 5) if formula_row <= sheet.nrows else None}
        ]
    cells.update(
        _upper_standard_cells(
            sheet.nrows,
            min(sheet.ncols, 8),
            lambda row_index, column_index: sheet.cell_value(row_index - 1, column_index - 1),
        )
    )
    return cells


def _upper_standard_cells(row_count: int, column_count: int, value_at: Any) -> dict[str, list[dict[str, Any]]]:
    cells = {comparison["field"]: [] for comparison in UPPER_STANDARD_COMPARISONS}
    anchor_row = None
    for row_index in range(1, row_count + 1):
        row_text = "".join(_normalize_label(value_at(row_index, column_index)) for column_index in range(1, column_count + 1))
        if any(section_label in row_text for section_label in UPPER_STANDARD_SECTION_LABELS):
            anchor_row = row_index
    if anchor_row is None:
        return cells

    last_row = min(row_count, anchor_row + 10)
    for row_index in range(anchor_row + 1, last_row + 1):
        for column_index in range(1, column_count + 1):
            label = _normalize_label(value_at(row_index, column_index)).rstrip(":：")
            field = UPPER_STANDARD_LABEL_FIELDS.get(label)
            if field is None:
                continue
            value_column, value = _next_nonempty_cell(value_at, row_index, column_index + 1, column_count)
            if value_column is None:
                continue
            cells[field].append(
                {
                    "cell": f"{_column_letter(value_column)}{row_index}",
                    "value": value,
                }
            )
    return cells


def _next_nonempty_cell(
    value_at: Any,
    row_index: int,
    first_column: int,
    column_count: int,
) -> tuple[int | None, Any]:
    for column_index in range(first_column, column_count + 1):
        value = value_at(row_index, column_index)
        if not _is_blank(value):
            if _normalize_label(value).rstrip(":：") in UPPER_STANDARD_LABEL_FIELDS:
                return None, None
            return column_index, value
    return None, None


def _column_letter(column_index: int) -> str:
    result = ""
    value = column_index
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _find_o3_value_pass_rows(labels: Any) -> tuple[int | None, int | None]:
    slope_row = None
    intercept_row = None
    for row_index, value in labels:
        label = _normalize_label(value)
        if slope_row is None and "斜率" in label:
            slope_row = row_index
        if intercept_row is None and "截距" in label:
            intercept_row = row_index
        if slope_row is not None and intercept_row is not None:
            break
    return slope_row, intercept_row


def _find_formula_row(labels: Any) -> int | None:
    for row_index, value in labels:
        label = _normalize_label(value)
        if not label:
            continue
        if "公式" in label and ("传递" in label or "拟合" in label or "线性" in label):
            return row_index
    return None


def _dynamic_value_cells(
    slope_row: int | None,
    intercept_row: int | None,
    value_at: Any,
) -> dict[str, list[dict[str, Any]]]:
    cells = {
        "DEVICEDELIVERMODEL": [],
        "DELIVERFC": [],
        "DENSITY1VALUE": [],
    }
    if slope_row is not None:
        cells["DEVICEDELIVERMODEL"].append({"cell": f"F{slope_row}", "value": value_at(slope_row)})
    if intercept_row is not None:
        cells["DELIVERFC"].append({"cell": f"F{intercept_row}", "value": value_at(intercept_row)})
        cells["DENSITY1VALUE"].extend(
            {"cell": f"F{row_index}", "value": value_at(row_index)}
            for row_index in range(intercept_row + 1, intercept_row + 4)
        )
    return cells


def _normalize_label(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("（", "(").replace("）", ")")


def _resolve_source(source: str) -> dict[str, Any]:
    if not source:
        return {"status": "error", "error": "附件路径为空"}

    parsed = urlparse(source)
    if parsed.scheme in {"http", "https"}:
        return _download_to_temp(source)

    path = Path(source).expanduser()
    if path.exists():
        return {"status": "success", "path": str(path)}

    attachment_root = os.getenv("OPS_ATTACHMENT_ROOT") or os.getenv("ATTACHMENT_ROOT")
    if attachment_root:
        rooted = Path(attachment_root).expanduser() / source.lstrip("/")
        if rooted.exists():
            return {"status": "success", "path": str(rooted)}

    attachment_base_url = os.getenv("OPS_ATTACHMENT_BASE_URL") or os.getenv("ATTACHMENT_BASE_URL")
    if attachment_base_url and source.startswith("/"):
        return _download_to_temp(urljoin(attachment_base_url.rstrip("/") + "/", source.lstrip("/")))

    return {"status": "error", "error": f"文件不存在且未配置附件根路径/基础URL：{source}"}


def _download_to_temp(url: str) -> dict[str, Any]:
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
    except requests.RequestException as exc:
        return {"status": "error", "error": f"下载附件失败：{exc}"}
    suffix = Path(urlparse(url).path).suffix or ".xls"
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        handle.write(response.content)
        return {"status": "success", "path": handle.name}
    finally:
        handle.close()


def _is_unavailable_attachment_source_error(error: Any) -> bool:
    text = str(error or "")
    return "附件路径为空" in text or "文件不存在且未配置附件根路径/基础URL" in text


def _compare_values(form: dict[str, Any], cells: dict[str, Any]) -> list[dict[str, Any]]:
    comparisons = []
    for comparison in COMPARISONS:
        field = comparison["field"]
        if cells.get("__source_type") == "pdf_text" and not cells.get(field):
            continue
        cell = comparison["cell"]
        form_raw = form.get(field)
        form_value = _number(form_raw)
        form_precision = _decimal_places(form_raw)
        cell_candidates = _cell_candidates(cells.get(field), comparison)
        matched_cell = _matched_cell(form_value, form_precision, cell_candidates, field)
        first_numeric_cell = next((item for item in cell_candidates if item["number"] is not None), None)
        used_cell = first_numeric_cell or (
            cell_candidates[0] if cell_candidates else {"cell": cell, "value": None, "number": None}
        )
        if form_value is None:
            status = "missing_form_value"
        elif matched_cell is not None:
            status = "match"
            used_cell = matched_cell
        elif first_numeric_cell is None:
            status = "missing_xls_value"
        else:
            status = "mismatch"
        comparisons.append(
            {
                **comparison,
                "configured_cell": cell,
                "candidate_cells": [item["cell"] for item in cell_candidates],
                "cell_candidates": [
                    {"cell": item["cell"], "value": item["value"], "number": item["number"]}
                    for item in cell_candidates
                ],
                "cell": used_cell["cell"],
                "form_value": form_raw,
                "xls_value": used_cell["value"],
                "form_number": form_value,
                "xls_number": used_cell["number"],
                "comparison_number": used_cell.get("comparison_number"),
                "comparison_transform": used_cell.get("comparison_transform", ""),
                "form_precision": form_precision,
                "status": status,
            }
        )
    comparisons.extend(_compare_transfer_formula(form, cells.get("TRANSFER_FORMULA")))
    comparisons.extend(_compare_upper_standard_values(form, cells))
    return comparisons


def _compare_upper_standard_values(form: dict[str, Any], cells: dict[str, Any]) -> list[dict[str, Any]]:
    results = []
    for comparison in UPPER_STANDARD_COMPARISONS:
        field = comparison["field"]
        candidates = cells.get(field) or []
        if not candidates:
            continue
        candidate = next((item for item in candidates if not _is_blank(item.get("value"))), candidates[0])
        form_value = form.get(field)
        xls_value = candidate.get("value")
        if _is_blank(form_value):
            status = "missing_form_value"
        elif _upper_standard_values_match(
            form_value,
            xls_value,
            comparison["comparison_type"],
            field=field,
        ):
            status = "match"
        else:
            status = "mismatch"
        results.append(
            {
                **comparison,
                "configured_cell": candidate.get("cell"),
                "candidate_cells": [item.get("cell") for item in candidates],
                "cell_candidates": candidates,
                "cell": candidate.get("cell"),
                "form_value": form_value,
                "xls_value": xls_value,
                "status": status,
            }
        )
    return results


def _upper_standard_values_match(
    form_value: Any,
    xls_value: Any,
    comparison_type: str,
    *,
    field: str = "",
) -> bool:
    if comparison_type == "formula":
        form_formula = _parse_linear_formula(form_value)
        xls_formula = _parse_linear_formula(xls_value)
        return form_formula is not None and xls_formula is not None and all(
            abs(left - right) <= 1e-9 for left, right in zip(form_formula, xls_formula, strict=True)
        )
    if comparison_type == "date":
        form_date = _normalize_date_value(form_value)
        xls_date = _normalize_date_value(xls_value)
        return bool(form_date and xls_date and form_date == xls_date)
    if field == "DELIVER6VALUE":
        return _normalize_upper_standard_model(form_value) == _normalize_upper_standard_model(xls_value)
    return _normalize_identity_text(form_value) == _normalize_identity_text(xls_value)


def _normalize_identity_text(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _normalize_upper_standard_model(value: Any) -> str:
    normalized = _normalize_identity_text(value)
    match = re.fullmatch(r"(?:TE|THERMO|THERMOSCIENTIFIC)(49IPS)", normalized)
    return match.group(1) if match else normalized


def _normalize_date_value(value: Any) -> tuple[int, ...]:
    if isinstance(value, datetime):
        return value.year, value.month, value.day
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 80000:
        date_value = datetime(1899, 12, 30) + timedelta(days=float(value))
        return date_value.year, date_value.month, date_value.day
    text = str(value or "").strip()
    numbers = [int(item) for item in re.findall(r"\d+", text)]
    if len(numbers) >= 6 and numbers[0] >= 1900 and numbers[3] >= 1900:
        return tuple(numbers[:6])
    if len(numbers) == 4 and numbers[0] >= 1900:
        year, month, start_day, end_day = numbers
        return year, month, start_day, year, month, end_day
    if len(numbers) == 5 and numbers[0] >= 1900 and _has_abbreviated_date_range(text):
        year, month, start_day, end_month, end_day = numbers
        return year, month, start_day, year, end_month, end_day
    if len(numbers) >= 3 and numbers[0] >= 1900:
        return tuple(numbers[:3])
    if len(numbers) >= 3 and numbers[2] >= 1900:
        first, second, year = numbers[:3]
        if first > 12:
            return year, second, first
        return year, first, second
    return tuple(numbers)


def _has_abbreviated_date_range(value: str) -> bool:
    return bool(
        re.search(r"(?:~|～|至|\bto\b)", value, flags=re.IGNORECASE)
        or re.search(r"\d\s*-\s*\d{1,2}[./]\d{1,2}\s*$", value)
    )


def _compare_transfer_formula(form: dict[str, Any], cell_data: Any) -> list[dict[str, Any]]:
    cell_candidates = _cell_candidates(cell_data, {"cell": "F", "field": "TRANSFER_FORMULA"})
    formula_cell = next((item for item in cell_candidates if str(item.get("value") or "").strip()), None)
    if formula_cell is None:
        return []

    form_slope = _number(form.get("DEVICEDELIVERMODEL"))
    form_intercept = _number(form.get("DELIVERFC"))
    parsed = _parse_linear_formula(formula_cell.get("value"))
    base = {
        "field": "TRANSFER_FORMULA",
        "label": "最佳拟合线性的传递公式",
        "configured_cell": "F",
        "candidate_cells": [item["cell"] for item in cell_candidates],
        "cell_candidates": [
            {"cell": item["cell"], "value": item["value"], "number": item["number"]}
            for item in cell_candidates
        ],
        "cell": formula_cell["cell"],
        "form_value": f"斜率={_display_value(form.get('DEVICEDELIVERMODEL'))}, 截距={_display_value(form.get('DELIVERFC'))}",
        "xls_value": formula_cell.get("value"),
        "form_number": None,
        "xls_number": None,
        "comparison_number": None,
        "comparison_transform": "linear_formula",
        "form_precision": max(_decimal_places(form.get("DEVICEDELIVERMODEL")), _decimal_places(form.get("DELIVERFC"))),
    }
    if form_slope is None or form_intercept is None:
        return [{**base, "status": "missing_form_value"}]
    if parsed is None:
        return [{**base, "status": "formula_unparseable"}]
    formula_slope, formula_intercept = parsed
    slope_precision = _decimal_places(form.get("DEVICEDELIVERMODEL"))
    intercept_precision = _decimal_places(form.get("DELIVERFC"))
    if (
        round(formula_slope, slope_precision) == round(form_slope, slope_precision)
        and round(formula_intercept, intercept_precision) == round(form_intercept, intercept_precision)
    ):
        return [{**base, "status": "match", "formula_slope": formula_slope, "formula_intercept": formula_intercept}]
    return [{**base, "status": "mismatch", "formula_slope": formula_slope, "formula_intercept": formula_intercept}]


def _parse_linear_formula(value: Any) -> tuple[float, float] | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"([-+]?\d+(?:\.\d+)?)\s*\*?\s*\(?\s*[xXｘＸ]\s*\)?", text)
    if not match:
        return None
    slope = float(match.group(1))
    tail = text[match.end():]
    intercept_match = re.search(r"([-+]\s*\d+(?:\.\d+)?)", tail)
    if not intercept_match:
        return None
    intercept = float(intercept_match.group(1).replace(" ", ""))
    return slope, intercept


def _cell_candidates(cell_data: Any, comparison: dict[str, Any]) -> list[dict[str, Any]]:
    if isinstance(cell_data, list):
        raw_items = cell_data
    elif isinstance(cell_data, dict):
        raw_items = [cell_data]
    else:
        raw_items = [{"cell": comparison["cell"], "value": cell_data}]

    candidates = []
    for item in raw_items:
        if isinstance(item, dict):
            cell = str(item.get("cell") or comparison["cell"])
            value = item.get("value")
        else:
            cell = str(comparison["cell"])
            value = item
        candidates.append({"cell": cell, "value": value, "number": _number(value)})
    return candidates


def _matched_cell(
    form_value: float | None,
    form_precision: int,
    candidates: list[dict[str, Any]],
    field: str,
) -> dict[str, Any] | None:
    if form_value is None:
        return None
    for candidate in candidates:
        for comparison_number, transform in _comparison_numbers(candidate.get("number"), field):
            if round(comparison_number, form_precision) == round(form_value, form_precision):
                matched = dict(candidate)
                matched["comparison_number"] = comparison_number
                matched["comparison_transform"] = transform
                return matched
    return None


def _comparison_numbers(value: float | None, field: str) -> list[tuple[float, str]]:
    if value is None:
        return []
    values = [(value, "")]
    if field == "DENSITY1VALUE":
        values.append((value * 100, "percent_scale"))
    return values


def _decimal_places(value: Any) -> int:
    text = str(value or "").strip()
    match = re.search(r"[-+]?\d+(?:\.(\d+))?", text)
    if not match or match.group(1) is None:
        return 0
    return len(match.group(1))


def _add_issue(
    order: dict[str, Any],
    form: dict[str, Any],
    item: dict[str, Any] | None,
    comparisons: list[dict[str, Any]],
    issues: list[Issue],
) -> None:
    for comparison in comparisons:
        evidence = {
            "working_order_code": order.get("WORKINGORDERCODE") or form.get("WORKINGORDERCODE"),
            "rf_table": RF_TABLE,
            "attachment": item,
            "comparison": comparison,
            "comparisons": [comparison],
        }
        comparison_field = str(comparison.get("field") or comparison.get("cell") or "xls")
        add_issue(
            issues,
            RULE_ID,
            "附件读数一致性",
            "高",
            f"attachment.{RF_TABLE}.{comparison_field}",
            _issue_message([comparison]),
            json.dumps(evidence, ensure_ascii=False, default=str),
        )


def _issue_message(comparisons: list[dict[str, Any]]) -> str:
    details = [_comparison_detail(comparison) for comparison in comparisons[:3]]
    details = [detail for detail in details if detail]
    if details:
        return f"O3量值传递表单与XLS附件不一致：{'；'.join(details)}"
    return "O3量值传递表单与XLS附件不一致"


def _comparison_detail(comparison: dict[str, Any]) -> str:
    status = comparison.get("status")
    if status == "xls_read_error":
        return f"XLS附件读取失败，原因：{_display_value(comparison.get('error'))}"
    if status == "missing_xls_attachment":
        return "未找到O3量值传递XLS附件"

    label = _display_value(comparison.get("label"))
    field = _display_value(comparison.get("field"))
    cell = _display_value(comparison.get("cell") or comparison.get("configured_cell"))
    form_value = _display_value(comparison.get("form_value"))
    xls_value = _display_value(comparison.get("xls_value"))

    if status == "mismatch":
        return f"{label}不一致：表单字段{field}表单值{form_value}，XLS {cell}值{xls_value}"
    if status == "missing_form_value":
        return f"{label}表单字段{field}为空，XLS {cell}值{xls_value}"
    if status == "missing_xls_value":
        return f"{label}XLS {cell}为空，表单字段{field}值{form_value}"
    if status == "formula_unparseable":
        return f"{label}XLS {cell}无法解析，公式文本{xls_value}"
    return f"{label}异常：{_display_value(status)}"


def _display_value(value: Any) -> str:
    text = str(value).strip() if value is not None else ""
    return text or "空"


def _first_present(record: dict[str, Any], fields: list[str]) -> Any:
    for field in fields:
        value = record.get(field)
        if value is not None and str(value).strip():
            return value
    return None


def _number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"/", "-", "无", "NA", "N/A"}:
        return None
    text = text.replace("％", "%")
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def _is_blank(value: Any) -> bool:
    return value is None or not str(value).strip()
