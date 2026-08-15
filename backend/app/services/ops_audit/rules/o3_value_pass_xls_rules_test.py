import json
from io import BytesIO

from openpyxl import Workbook

from app.services.ops_audit.rules import o3_value_pass_xls_rules


def _xlsx(path, *, slope="0.999", intercept="-0.079", change="-0.07"):
    wb = Workbook()
    ws = wb.active
    ws["D25"] = "斜率"
    ws["F25"] = slope
    ws["D26"] = "截距(ppb)"
    ws["F26"] = intercept
    ws["F28"] = change
    wb.save(path)


def _xlsx_with_formula(path, *, slope="0.999", intercept="-0.079", change="-0.07", formula="y=1.001x-0.079"):
    wb = Workbook()
    ws = wb.active
    ws["D25"] = "斜率"
    ws["F25"] = slope
    ws["D26"] = "截距(ppb)"
    ws["F26"] = intercept
    ws["F28"] = change
    ws["D30"] = "最佳拟合线性的传递公式"
    ws["F30"] = formula
    wb.save(path)


def _xlsx_bytes(*, slope="0.999", intercept="-0.079", change="-0.07"):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws["D25"] = "斜率"
    ws["F25"] = slope
    ws["D26"] = "截距(ppb)"
    ws["F26"] = intercept
    ws["F28"] = change
    wb.save(buffer)
    return buffer.getvalue()


def _form(**overrides):
    form = {
        "WORKINGORDERCODE": "WO-001",
        "DEVICEDELIVERMODEL": "0.999",
        "DELIVERFC": "-0.079",
        "DENSITY1VALUE": "-0.07",
    }
    form.update(overrides)
    return form


def _attachment(path):
    return {
        "REFID": "WO-001",
        "TYPECODE": "RF_HY_O3ValuePass",
        "FILENAME": path.name,
        "FILEPATH": str(path),
    }


def test_o3_value_pass_xls_matching_values_adds_no_issue(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    _xlsx(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert issues == []


def test_o3_value_pass_prefers_matching_template_over_report_xls(tmp_path):
    report_path = tmp_path / "GDEEMC-2025-42 臭氧标准传递报告.xlsx"
    report_wb = Workbook()
    report_ws = report_wb.active
    report_ws["D34"] = "相关系数"
    report_ws["F34"] = "Coefficient (Ri)"
    report_wb.save(report_path)

    template_path = tmp_path / "综合观测点臭氧标准传递20260724.xlsx"
    _xlsx(template_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [
            {**_attachment(report_path), "FILENAME": "GDEEMC-2025-42 臭氧标准传递报告.xls"},
            {**_attachment(template_path), "FILENAME": "综合观测点臭氧标准传递20260724.xls"},
        ],
        issues,
    )

    assert issues == []


def test_o3_value_pass_xls_accepts_any_xls_attachment_by_format(tmp_path):
    xls_path = tmp_path / "daily-record.xlsx"
    _xlsx(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [
            {
                "REFID": "WO-001",
                "TYPECODE": "UNRELATED_ATTACHMENT",
                "FILENAME": "daily-record.xlsx",
                "FILEPATH": str(xls_path),
            }
        ],
        [],
        issues,
    )

    assert issues == []


def test_o3_value_pass_xls_missing_attachment_adds_manual_review_issue():
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [
            {
                "REFID": "WO-001",
                "TYPECODE": "RF_HY_O3ValuePass",
                "FILENAME": "传递点100.jpg",
                "FILEPATH": "/WebFiles/NewFiles/o3/point100.jpg",
            },
            {
                "REFID": "WO-001",
                "TYPECODE": "RF_HY_O3ValuePass",
                "FILENAME": "臭氧标准传递报告.pdf",
                "FILEPATH": "/WebFiles/NewFiles/o3/report.pdf",
            },
        ],
        issues,
    )

    assert [issue.rule_id for issue in issues] == ["ATTACHMENT_O3_VALUE_PASS_XLS_MISSING_REVIEW"]
    evidence = json.loads(issues[0].evidence)
    assert evidence["needs_manual_review"] is True


def test_o3_value_pass_xls_unavailable_local_path_is_not_reported():
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [
            {
                "REFID": "WO-001",
                "TYPECODE": "RF_HY_O3ValuePass",
                "FILENAME": "臭氧标准传递.xls",
                "FILEPATH": "/WebFiles/NewFiles/o3/missing.xls",
            }
        ],
        issues,
    )

    assert issues == []


def test_o3_value_pass_xls_mismatch_adds_issue(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    _xlsx(xls_path, slope="1.001")
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert len(issues) == 1
    assert issues[0].rule_id == "ATTACHMENT_O3_VALUE_PASS_XLS_VALUE_MISMATCH"
    assert "斜率" in issues[0].message
    assert "DEVICEDELIVERMODEL" in issues[0].message
    assert "表单值0.999" in issues[0].message
    assert "XLS F25值1.001" in issues[0].message
    evidence = json.loads(issues[0].evidence)
    assert evidence["comparisons"][0]["field"] == "DEVICEDELIVERMODEL"
    assert evidence["comparisons"][0]["cell"] == "F25"
    assert evidence["comparisons"][0]["status"] == "mismatch"


def test_o3_value_pass_xls_emits_each_field_mismatch_separately(tmp_path):
    xls_path = tmp_path / "o3-transfer-multiple-mismatches.xlsx"
    _xlsx(xls_path, slope="1.001", intercept="0.25")
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert len(issues) == 2
    assert {json.loads(issue.evidence)["comparison"]["field"] for issue in issues} == {
        "DEVICEDELIVERMODEL",
        "DELIVERFC",
    }
    assert all(len(json.loads(issue.evidence)["comparisons"]) == 1 for issue in issues)
    assert all("；" not in issue.message for issue in issues)


def test_o3_value_pass_xls_formula_text_mismatch_adds_issue(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    _xlsx_with_formula(xls_path, formula="y=1.001x-0.079")
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert len(issues) == 1
    assert "传递公式" in issues[0].message
    evidence = json.loads(issues[0].evidence)
    comparison = evidence["comparisons"][0]
    assert comparison["field"] == "TRANSFER_FORMULA"
    assert comparison["status"] == "mismatch"


def test_o3_value_pass_xls_formula_text_match_adds_no_issue(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    _xlsx_with_formula(xls_path, formula="y=0.999x-0.079")
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert issues == []


def test_o3_value_pass_text_pdf_formula_mismatch_adds_issue(monkeypatch, tmp_path):
    pdf_path = tmp_path / "o3-transfer-report.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    def fake_read_pdf_text(item):
        return {
            "status": "success",
            "text": "斜率 0.999\n截距(ppb) -0.079\n最佳拟合线性的传递公式 y=1.001x-0.079",
        }

    monkeypatch.setattr(o3_value_pass_xls_rules, "_read_pdf_text", fake_read_pdf_text)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [
            {
                "REFID": "WO-001",
                "TYPECODE": "RF_HY_O3ValuePass",
                "FILENAME": "臭氧传递报告.pdf",
                "FILEPATH": str(pdf_path),
            }
        ],
        issues,
    )

    mismatches = [issue for issue in issues if issue.rule_id == "ATTACHMENT_O3_VALUE_PASS_XLS_VALUE_MISMATCH"]
    assert len(mismatches) == 1
    assert "传递公式" in mismatches[0].message
    evidence = json.loads(mismatches[0].evidence)
    assert evidence["comparisons"][0]["field"] == "TRANSFER_FORMULA"
    assert evidence["attachment"]["filename"] == "臭氧传递报告.pdf"
    assert [issue.rule_id for issue in issues if issue.rule_id != "ATTACHMENT_O3_VALUE_PASS_XLS_VALUE_MISMATCH"] == [
        "ATTACHMENT_O3_VALUE_PASS_XLS_MISSING_REVIEW"
    ]


def test_o3_value_pass_xls_missing_form_field_adds_issue(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    _xlsx(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form(DENSITY1VALUE=""))],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert len(issues) == 1
    evidence = json.loads(issues[0].evidence)
    assert evidence["comparisons"][0]["field"] == "DENSITY1VALUE"
    assert evidence["comparisons"][0]["cell"] == "F28"
    assert evidence["comparisons"][0]["status"] == "missing_form_value"


def test_o3_value_pass_xls_uses_file_url_when_filepath_is_unavailable(monkeypatch):
    class Response:
        content = _xlsx_bytes()

        def raise_for_status(self):
            return None

    requested_urls = []

    def fake_get(url, timeout):
        requested_urls.append((url, timeout))
        return Response()

    monkeypatch.setattr(o3_value_pass_xls_rules.requests, "get", fake_get)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [
            {
                "REFID": "WO-001",
                "TYPECODE": "RF_HY_O3ValuePass",
                "FILENAME": "o3-transfer.xlsx",
                "FILEPATH": "/WebFiles/NewFiles/o3-transfer.xlsx",
                "file_url": "http://files.example.test/WebFiles/NewFiles/o3-transfer.xlsx",
            }
        ],
        issues,
    )

    assert issues == []
    assert requested_urls == [("http://files.example.test/WebFiles/NewFiles/o3-transfer.xlsx", 30)]


def test_o3_value_pass_xls_accepts_values_in_actual_template_cells(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["D25"] = "斜率"
    ws["F25"] = "0.999"
    ws["D26"] = "截距(ppb)"
    ws["F26"] = "-0.079"
    ws["F28"] = "-0.07"
    wb.save(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert issues == []


def test_o3_value_pass_xls_uses_d_column_labels_for_value_cells(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["D24"] = "斜率"
    ws["F24"] = "1.002"
    ws["D25"] = "截距（ppb）"
    ws["F25"] = "0.110"
    ws["F26"] = "not-the-change"
    ws["F28"] = "-0.5"
    wb.save(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [
            (
                "RF_HY_O3VALUEPASS",
                _form(
                    DEVICEDELIVERMODEL="1.002",
                    DELIVERFC="0.110",
                    DENSITY1VALUE="-0.5",
                ),
            )
        ],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert issues == []


def test_o3_value_pass_field_position_ignores_actual_template_metadata_fields():
    issues = []
    form = _form(
        DELIVER1VALUE="389",
        WORKDENSITY1VALUE="42.6",
        DELIVERFROM1VALUE="8.5",
        DELIVERTO1VALUE="291",
        DELIVER2VALUE="8.5",
        WORKDENSITY2VALUE="195",
        DELIVERFROM2VALUE="25.5",
        DELIVERTO2VALUE="8.5",
        DELIVER3VALUE="17.4",
        WORKDENSITY3VALUE="8.5",
        DELIVERFROM3VALUE="48.5",
        DELIVERTO3VALUE="13.1",
        DELIVER4VALUE="2026-6-22",
        WORKDENSITY4VALUE="2026-9-21",
        DELIVERFROM4VALUE="TE-146i",
        DELIVER5VALUE="2026-6-17",
        WORKDENSITY5VALUE="??????",
        DELIVER6VALUE="49ips",
        WORKDENSITY6VALUE="2026-3-4to6",
    )

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-ACTUAL-TEMPLATE"},
        [("RF_HY_O3VALUEPASS", form)],
        [],
        [],
        issues,
    )

    assert [issue for issue in issues if issue.rule_id == "RF_O3_VALUE_PASS_FIELD_POSITION_SUSPECT"] == []
    assert [issue.rule_id for issue in issues] == ["ATTACHMENT_O3_VALUE_PASS_XLS_MISSING_REVIEW"]


def test_o3_value_pass_xls_accepts_change_value_within_three_rows_after_intercept(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["D25"] = "斜率"
    ws["F25"] = "0.999"
    ws["D26"] = "截距(ppb)"
    ws["F26"] = "-0.079"
    ws["F27"] = "not-the-change"
    ws["F28"] = "-0.5"
    wb.save(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form(DENSITY1VALUE="-0.5"))],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert issues == []


def test_o3_value_pass_xls_does_not_use_fixed_cells_without_d_column_labels(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["F25"] = "0.999"
    ws["F26"] = "-0.079"
    ws["F28"] = "-0.07"
    wb.save(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [("RF_HY_O3VALUEPASS", _form())],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert len(issues) == 3
    statuses = {
        json.loads(issue.evidence)["comparison"]["field"]:
        json.loads(issue.evidence)["comparison"]["status"]
        for issue in issues
    }
    assert statuses == {
        "DEVICEDELIVERMODEL": "missing_xls_value",
        "DELIVERFC": "missing_xls_value",
        "DENSITY1VALUE": "missing_xls_value",
    }


def test_o3_value_pass_xls_matches_using_form_precision_and_percent_scale(tmp_path):
    xls_path = tmp_path / "o3-transfer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["D24"] = "斜率"
    ws["F24"] = 0.9994584484828917
    ws["D25"] = "截距(ppb)"
    ws["F25"] = -0.11881566910229442
    ws["F27"] = -0.010978043912175658
    wb.save(xls_path)
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "WO-001"},
        [
            (
                "RF_HY_O3VALUEPASS",
                _form(
                    DEVICEDELIVERMODEL="0.999",
                    DELIVERFC="-0.119",
                    DENSITY1VALUE="-1.10",
                ),
            )
        ],
        [],
        [_attachment(xls_path)],
        issues,
    )

    assert issues == []


def test_xlsx_scan_uses_one_bounded_sequential_iteration():
    rows = [[None] * 8 for _ in range(30)]
    rows[24][3], rows[24][5] = "斜率", "0.999"
    rows[25][3], rows[25][5] = "截距(ppb)", "-0.079"
    rows[27][5] = "-0.07"

    class Worksheet:
        max_row = 421

        def __init__(self):
            self.calls = []

        def iter_rows(self, **kwargs):
            self.calls.append(kwargs)
            return iter(tuple(row) for row in rows)

        def __getitem__(self, key):
            raise AssertionError(f"unexpected random cell access: {key}")

        def cell(self, **kwargs):
            raise AssertionError(f"unexpected random cell access: {kwargs}")

    worksheet = Worksheet()
    cells = o3_value_pass_xls_rules._worksheet_dynamic_cells(worksheet)

    assert worksheet.calls == [
        {
            "min_row": 1,
            "max_row": o3_value_pass_xls_rules.XLSX_SCAN_MAX_ROWS,
            "min_col": 1,
            "max_col": o3_value_pass_xls_rules.XLSX_SCAN_MAX_COLUMNS,
            "values_only": True,
        }
    ]
    assert cells["DEVICEDELIVERMODEL"] == [{"cell": "F25", "value": "0.999"}]
    assert cells["DELIVERFC"] == [{"cell": "F26", "value": "-0.079"}]
    assert {item["value"] for item in cells["DENSITY1VALUE"]} == {None, "-0.07"}


def test_same_attachment_is_parsed_once_across_forms_and_orders(monkeypatch, tmp_path):
    xls_path = tmp_path / "shared-o3-transfer.xlsx"
    _xlsx(xls_path)
    read_count = 0
    original_read_cells = o3_value_pass_xls_rules._read_cells

    def counted_read_cells(item):
        nonlocal read_count
        read_count += 1
        return original_read_cells(item)

    monkeypatch.setattr(o3_value_pass_xls_rules, "_read_cells", counted_read_cells)
    shared_cache = {}
    for code in ("WO-001", "WO-002"):
        issues = []
        form = _form(WORKINGORDERCODE=code)
        attachment = {**_attachment(xls_path), "REFID": code}
        o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
            {"WORKINGORDERCODE": code},
            [("RF_HY_O3VALUEPASS", form), ("RF_HY_O3VALUEPASS", dict(form))],
            [],
            [attachment],
            issues,
            attachment_read_cache=shared_cache,
        )
        assert issues == []

    assert read_count == 1


def test_o3_value_pass_missing_transfer_flow_value_is_reported():
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "CH2606151781505099937"},
        [
            (
                "RF_HY_O3VALUEPASS",
                _form(
                    DELIVER1VALUE="398",
                    DELIVER2VALUE="",
                    DELIVERFROM2VALUE="28.8%",
                    WORKDENSITY2VALUE="199",
                ),
            )
        ],
        [],
        [],
        issues,
    )

    matched = [issue for issue in issues if issue.rule_id == "RF_O3_VALUE_PASS_FLOW_VALUE_MISSING"]
    assert len(matched) == 1
    assert "第2组传递流量未填写" in matched[0].message
    evidence = json.loads(matched[0].evidence)
    assert evidence["violations"][0]["field"] == "DELIVER2VALUE"


def test_o3_value_pass_date_in_metadata_fields_is_not_reported_as_position_suspect():
    issues = []

    o3_value_pass_xls_rules.check_o3_value_pass_xls_values(
        {"WORKINGORDERCODE": "CH2606151781505099937"},
        [
            (
                "RF_HY_O3VALUEPASS",
                _form(
                    DELIVER4VALUE="2026-06-15",
                    WORKDENSITY4VALUE="2026-09-15",
                    DELIVERFROM4VALUE="TE-146i",
                    DELIVERTO4VALUE="?????",
                ),
            )
        ],
        [],
        [],
        issues,
    )

    matched = [issue for issue in issues if issue.rule_id == "RF_O3_VALUE_PASS_FIELD_POSITION_SUSPECT"]
    assert matched == []
