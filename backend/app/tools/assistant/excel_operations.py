# -*- coding: utf-8 -*-
"""
会商Excel文件操作模块

功能：
- 读取和更新会商Excel文件
- 计算统计指标（平均、同比、排序）
- 数据验证
- 合并多个Excel文件

author: Claude
date: 2026-05-08
"""

import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple, Optional
import structlog
from pathlib import Path

logger = structlog.get_logger()


class ConsultationExcelOperator:
    """会商Excel文件操作器"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None

    def load_file(self, sheet_index: int = 0):
        """加载Excel文件"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook.worksheets[sheet_index]
            logger.info(f"Loaded Excel file: {self.file_path}, sheet: {self.sheet.title}")
        except Exception as e:
            logger.error(f"Failed to load Excel file {self.file_path}: {str(e)}")
            raise

    def save_file(self):
        """保存Excel文件"""
        try:
            self.workbook.save(self.file_path)
            logger.info(f"Saved Excel file: {self.file_path}")
        except Exception as e:
            logger.error(f"Failed to save Excel file {self.file_path}: {str(e)}")
            raise

    def update_column_data(
        self,
        column: str,
        data: List[float],
        start_row: int = 2,
        precision: int = 2,
        end_row: Optional[int] = None
    ):
        """
        更新列数据

        Args:
            column: 列字母（如"B"）
            data: 数据列表
            start_row: 起始行号
            precision: 小数位数
        """
        for i, value in enumerate(data):
            cell = self.sheet[f"{column}{start_row + i}"]
            cell.value = round(value, precision) if value is not None else None

        if end_row is not None:
            for row in range(start_row + len(data), end_row + 1):
                self.sheet[f"{column}{row}"].value = None

        logger.info(f"Updated column {column} with {len(data)} rows")

    def align_records_to_names(
        self,
        records: List[Dict[str, Any]],
        name_column: str = "A",
        start_row: int = 2,
        end_row: int = 32
    ) -> List[float]:
        """按模板名称列对齐API记录，避免依赖API返回顺序。"""
        value_by_name = {}
        duplicates = []

        for record in records:
            name = self._normalize_area_name(record.get("name", ""))
            if not name:
                continue
            if name in value_by_name:
                duplicates.append(name)
                continue
            value_by_name[name] = record.get("value")

        if duplicates:
            raise ValueError(f"API数据存在重复地区: {sorted(set(duplicates))}")

        aligned = []
        missing = []
        template_names = []
        for row in range(start_row, end_row + 1):
            raw_name = self.sheet[f"{name_column}{row}"].value
            name = self._normalize_area_name(raw_name)
            if not name:
                continue
            if name in template_names:
                raise ValueError(f"模板{name_column}列存在重复地区: {name}")
            template_names.append(name)

            if name not in value_by_name:
                missing.append(name)
                aligned.append(None)
            else:
                aligned.append(value_by_name[name])

        if missing:
            raise ValueError(f"API数据缺少模板地区: {missing}")

        return aligned

    def update_column_name(self, column: str, name: str, row: int = 1):
        """
        更新列名

        Args:
            column: 列字母（如"B"）
            name: 列名
            row: 行号（默认第1行）
        """
        cell = self.sheet[f"{column}{row}"]
        cell.value = name
        logger.info(f"Updated column {column} name to: {name}")

    def read_column_data(
        self,
        column: str,
        start_row: int = 2,
        end_row: int = None
    ) -> List[float]:
        """
        读取列数据

        Args:
            column: 列字母（如"B"）
            start_row: 起始行号
            end_row: 结束行号（None表示到数据末尾）

        Returns:
            数据列表
        """
        if end_row is None:
            end_row = self.sheet.max_row

        data = []
        for row in range(start_row, end_row + 1):
            cell = self.sheet[f"{column}{row}"]
            value = cell.value
            if value is not None:
                try:
                    data.append(float(value))
                except (ValueError, TypeError):
                    data.append(0.0)
            else:
                data.append(0.0)

        return data

    def calculate_column_average(
        self,
        column: str,
        start_row: int = 2,
        end_row: int = None,
        precision: int = 2
    ) -> float:
        """
        计算列平均值

        Args:
            column: 列字母
            start_row: 起始行号
            end_row: 结束行号
            precision: 小数位数

        Returns:
            平均值
        """
        data = self.read_column_data(column, start_row, end_row)
        if data:
            average = sum(data) / len(data)
            return round(average, precision)
        return 0.0

    def calculate_column_difference(
        self,
        col1: str,
        col2: str,
        output_col: str,
        start_row: int = 2,
        precision: int = 2,
        end_row: int = None
    ):
        """
        计算两列差值（col1 - col2）

        Args:
            col1: 列1
            col2: 列2
            output_col: 输出列
            start_row: 起始行号
            precision: 小数位数
        """
        data1 = self.read_column_data(col1, start_row, end_row)
        data2 = self.read_column_data(col2, start_row, end_row)

        diff_data = []
        for v1, v2 in zip(data1, data2):
            diff = v1 - v2
            diff_data.append(round(diff, precision))

        self.update_column_data(output_col, diff_data, start_row, end_row=end_row)
        logger.info(f"Calculated {col1} - {col2} -> {output_col}")

    def sort_and_update(
        self,
        data_column: str,
        name_column: str,
        output_name_col: str,
        output_data_col: str,
        output_rank_col: str,
        start_row: int = 2,
        end_row: int = None,
        ascending: bool = True
    ):
        """
        对数据排序并更新到指定列

        Args:
            data_column: 数据列（要排序的列）
            name_column: 名称列（省份/城市名称）
            output_name_col: 输出名称列
            output_data_col: 输出数据列
            output_rank_col: 输出排名列
            start_row: 起始行号
            ascending: 是否升序
        """
        # 读取数据
        if end_row is None:
            end_row = self.sheet.max_row

        data = self.read_column_data(data_column, start_row, end_row)
        names = []
        for row in range(start_row, end_row + 1):
            cell = self.sheet[f"{name_column}{row}"]
            names.append(cell.value or "")

        paired = [
            (name, value)
            for name, value in zip(names, data)
            if self._normalize_area_name(name)
        ]
        self._validate_unique_names([name for name, _ in paired], "排序源数据")
        paired.sort(key=lambda x: x[1], reverse=not ascending)

        for row in range(start_row, end_row + 1):
            self.sheet[f"{output_name_col}{row}"].value = None
            self.sheet[f"{output_data_col}{row}"].value = None
            self.sheet[f"{output_rank_col}{row}"].value = None

        # 更新排序后的数据
        for i, (name, value) in enumerate(paired):
            row = start_row + i
            self.sheet[f"{output_name_col}{row}"].value = name
            self.sheet[f"{output_data_col}{row}"].value = value
            self.sheet[f"{output_rank_col}{row}"].value = i + 1

        self._validate_unique_names(
            [name for name, _ in paired],
            f"{output_name_col}列排序结果"
        )

        logger.info(f"Sorted {data_column} and updated to {output_data_col}")

    def update_consultation_file(
        self,
        current_data: List[float],
        last_year_data: List[float],
        current_period: str,
        last_year_period: str,
        data_start_row: int = 2,
        data_end_row: int = 32  # 全国31个省份 + 1行 header
    ):
        """
        更新会商文件的完整流程

        Args:
            current_data: 当年数据
            last_year_data: 去年数据
            current_period: 当年时间段描述（如"2026年1-3月份"）
            last_year_period: 去年时间段描述（如"2025年1-3月份"）
            data_start_row: 数据起始行
            data_end_row: 数据结束行
        """
        # 步骤1：更新B列（当年数据）和列名
        self.update_column_name("B", current_period)
        self.update_column_data("B", current_data, data_start_row, end_row=data_end_row)

        # 步骤2：更新D列（去年数据）和列名
        self.update_column_name("D", last_year_period)
        self.update_column_data("D", last_year_data, data_start_row, end_row=data_end_row)

        # 步骤3：计算C列（B列平均值）
        avg_current = self.calculate_column_average("B", data_start_row, data_end_row)
        self.sheet[f"C{data_start_row}"].value = avg_current

        # 步骤4：计算E列（同比变化 = B列 - D列）
        self.calculate_column_difference("B", "D", "E", data_start_row, end_row=data_end_row)

        # 步骤5：更新G-I列（去年数据排序）
        self.sort_and_update(
            data_column="D",
            name_column="A",
            output_name_col="G",
            output_data_col="H",
            output_rank_col="I",
            start_row=data_start_row,
            end_row=data_end_row,
            ascending=True
        )

        # 步骤6：更新K-M列（同比数据排序）
        self.sort_and_update(
            data_column="E",
            name_column="A",
            output_name_col="K",
            output_data_col="L",
            output_rank_col="M",
            start_row=data_start_row,
            end_row=data_end_row,
            ascending=True
        )

        # 步骤7：计算同比平均值（M列下方）
        avg_diff = self.calculate_column_average("E", data_start_row, data_end_row)
        self.sheet[f"M{data_start_row}"].value = avg_diff

        logger.info(f"Updated consultation file: {self.file_path}")

    def update_consultation_file_by_name(
        self,
        current_records: List[Dict[str, Any]],
        last_year_records: List[Dict[str, Any]],
        current_period: str,
        last_year_period: str,
        data_start_row: int = 2,
        data_end_row: int = 32
    ):
        """按A列地区名称对齐后更新会商文件。"""
        current_data = self.align_records_to_names(
            current_records,
            start_row=data_start_row,
            end_row=data_end_row
        )
        last_year_data = self.align_records_to_names(
            last_year_records,
            start_row=data_start_row,
            end_row=data_end_row
        )

        self.update_consultation_file(
            current_data=current_data,
            last_year_data=last_year_data,
            current_period=current_period,
            last_year_period=last_year_period,
            data_start_row=data_start_row,
            data_end_row=data_end_row
        )

    @staticmethod
    def _normalize_area_name(name: Any) -> str:
        normalized = str(name or "").strip()
        for suffix in ("壮族自治区", "回族自治区", "维吾尔自治区", "自治区", "省", "市"):
            if normalized.endswith(suffix):
                normalized = normalized[: -len(suffix)]
                break
        return normalized

    def _validate_unique_names(self, names: List[Any], context: str):
        normalized = [self._normalize_area_name(name) for name in names if self._normalize_area_name(name)]
        duplicates = sorted({name for name in normalized if normalized.count(name) > 1})
        if duplicates:
            raise ValueError(f"{context}存在重复地区: {duplicates}")


def merge_excel_files(
    source_files: List[str],
    output_file: str,
    sheet_names: List[str] = None
) -> bool:
    """
    合并多个Excel文件为一个汇总文件

    Args:
        source_files: 源文件列表
        output_file: 输出文件路径
        sheet_names: sheet名称列表（默认使用源文件名）

    Returns:
        是否成功
    """
    try:
        # 创建新工作簿
        output_wb = openpyxl.Workbook()

        # 删除默认sheet
        if "Sheet" in output_wb.sheetnames:
            output_wb.remove(output_wb["Sheet"])

        # 复制每个源文件的第一个sheet
        for i, source_file in enumerate(source_files):
            if not Path(source_file).exists():
                logger.warning(f"Source file not found: {source_file}")
                continue

            source_wb = openpyxl.load_workbook(source_file)
            source_sheet = source_wb.worksheets[0]

            # 确定sheet名称
            sheet_name = sheet_names[i] if sheet_names else source_sheet.title
            if sheet_name in output_wb.sheetnames:
                sheet_name = f"{sheet_name}_{i}"

            # 创建新sheet并复制数据
            target_sheet = output_wb.create_sheet(title=sheet_name)

            for row in source_sheet.iter_rows():
                for cell in row:
                    target_sheet[cell.coordinate].value = cell.value
                    # 复制样式
                    if cell.has_style:
                        target_sheet[cell.coordinate].font = cell.font.copy()
                        target_sheet[cell.coordinate].border = cell.border.copy()
                        target_sheet[cell.coordinate].fill = cell.fill.copy()
                        target_sheet[cell.coordinate].number_format = cell.number_format
                        target_sheet[cell.coordinate].protection = cell.protection.copy()
                        target_sheet[cell.coordinate].alignment = cell.alignment.copy()

            source_wb.close()

        # 保存合并文件
        output_wb.save(output_file)
        output_wb.close()

        logger.info(f"Merged {len(source_files)} files into {output_file}")
        return True

    except Exception as e:
        logger.error(f"Failed to merge files: {str(e)}")
        return False


def validate_excel_data(
    file_path: str,
    pollutant: str,
    current_period: str,
    last_year_period: str
) -> Dict[str, Any]:
    """
    验证Excel数据的合理性

    Args:
        file_path: 文件路径
        pollutant: 污染物名称
        current_period: 当年时间段
        last_year_period: 去年时间段

    Returns:
        验证结果
    """
    result = {
        "valid": True,
        "errors": [],
        "warnings": []
    }

    try:
        operator = ConsultationExcelOperator(file_path)
        operator.load_file()

        # 检查1：列名格式
        b_name = operator.sheet["B1"].value
        d_name = operator.sheet["D1"].value

        if b_name != current_period:
            result["errors"].append(f"B列列名不正确：期望'{current_period}'，实际'{b_name}'")
            result["valid"] = False

        if d_name != last_year_period:
            result["errors"].append(f"D列列名不正确：期望'{last_year_period}'，实际'{d_name}'")
            result["valid"] = False

        # 检查2：数据范围
        normal_ranges = {
            "PM2.5": (5, 150),
            "PM10": (10, 300),
            "NO2": (5, 100),
            "O3": (10, 160),
            "AQI": (80, 100)
        }

        if pollutant in normal_ranges:
            min_val, max_val = normal_ranges[pollutant]
            b_data = operator.read_column_data("B", 2, 32)
            d_data = operator.read_column_data("D", 2, 32)

            for i, (val_b, val_d) in enumerate(zip(b_data, d_data)):
                if val_b < min_val or val_b > max_val:
                    result["warnings"].append(
                        f"第{i+2}行当年数据{val_b}超出正常范围[{min_val}, {max_val}]"
                    )
                if val_d < min_val or val_d > max_val:
                    result["warnings"].append(
                        f"第{i+2}行去年数据{val_d}超出正常范围[{min_val}, {max_val}]"
                    )

        # 检查3：数据一致性
        if b_data == d_data:
            result["errors"].append("当年数据和去年数据完全相同，可能查询了同一年数据")
            result["valid"] = False

        # 检查4：数据全为0
        if all(v == 0 for v in b_data):
            result["errors"].append("当年数据全为0，数据查询可能失败")
            result["valid"] = False

    except Exception as e:
        result["errors"].append(f"验证过程出错：{str(e)}")
        result["valid"] = False

    return result
