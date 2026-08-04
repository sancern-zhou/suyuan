# -*- coding: utf-8 -*-
"""
Execute Python Code Tool

执行 Python 代码工具（用于数据处理、可视化、中间资源生成）

特性：
- 在 Bubblewrap Linux namespace 沙箱和隔离的临时目录中执行代码
- 30 秒超时保护（可配置）
- 始终在独立子进程中执行，避免带崩 Web worker
- 返回生成的文件列表
- 自动清理临时文件
- 支持所有 Python 库（python-docx, matplotlib, pandas 等）

安全措施：
1. 文件系统、PID、IPC、UTS、cgroup 和网络 namespace 隔离
2. 仅挂载 Python 运行时、本次会话数据和临时输出目录
3. 默认无网络、丢弃全部 capabilities，并清除主进程敏感环境变量
4. CPU、虚拟内存、文件大小、进程数和文件描述符限制
5. 超时、输出截断（1MB）和子进程组级别终止
"""

import asyncio
import tempfile
import os
import shutil
import subprocess
import sys
import time
import base64
import re
import ast
import json
import signal
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
import structlog

from app.tools.base.tool_interface import LLMTool, ToolCategory
from app.tools.resource_refs import build_data_file_ref, build_file_ref, build_visual_ref, merge_refs
from app.utils.path_config import (
    PROJECT_ROOT,
    format_agent_path,
    get_charts_dir,
    get_data_registry,
    get_images_dir,
    get_python_output_dir,
    get_reports_dir,
    resolve_agent_path,
)

logger = structlog.get_logger()


class ExecutePythonTool(LLMTool):
    """
    Python 代码执行工具

    参考：
    - KIMI ipython 工具
    - openwork-dev: Python 执行环境
    """

    def __init__(self):
        # 永久文件存储目录（统一使用 backend/backend_data_registry）
        self.PERMANENT_DIR = str(get_data_registry() / "python_generated_files")

        # ✅ 图表文件存储目录
        self.CHARTS_DIR = str(get_charts_dir())
        self.REPORTS_DIR = str(get_reports_dir())

        # 确保永久目录和图表目录存在
        os.makedirs(self.PERMANENT_DIR, exist_ok=True)
        os.makedirs(self.CHARTS_DIR, exist_ok=True)
        os.makedirs(self.REPORTS_DIR, exist_ok=True)

        super().__init__(
            name="execute_python",
            description=(
                "执行 Python 代码，用于数据处理、数值计算、Excel/文件处理、"
                "⭐ **自定义图表生成**：matplotlib/seaborn/plotly/bokeh 绘制复杂/3D/科研图表。"
                "每次调用是独立环境；用 load_data(file_path) 读取会话数据文件，"
                "跨调用复用结果请用 save_data(...) 保存并获取 file_path。"
                "⚠️ **图表选择策略**："
                "① 标准报告图表（bar/line/scatter/pile/histogram等）→ 优先使用 create_report_chart；"
                "② 复杂/自定义图表（3D图/多子图/任意极坐标/科研图表）→ 使用 execute_python + matplotlib/seaborn/plotly；"
                "③ 流程图/架构图/步骤图 → 使用 call_sub_agent(target_mode='board') 调用画板Agent生成draw.io图片文件。"
                f"生成文件保存到统一数据目录：{get_data_registry()}。"
            ),
            category=ToolCategory.QUERY,
            version="1.0.3",
            requires_context=True  # ✅ 需要上下文以支持数据访问功能
        )

        # Agent 生成的代码绝不能在 Web worker 进程内执行。
        # 生产环境默认 fail-closed：沙箱不可用时返回工具失败，不回退到宿主机直接执行。
        self.execution_engine = os.getenv("PYTHON_EXECUTION_ENGINE", "bubblewrap").strip().lower()
        self.default_timeout = 30
        self.max_output_size = 1024 * 1024  # 1MB
        self.enable_echarts_visuals = False

        logger.info(
            "execute_python_tool_initialized",
            execution_engine=self.execution_engine,
            permanent_dir=self.PERMANENT_DIR,
            charts_dir=self.CHARTS_DIR,
            timeout=self.default_timeout
        )

    def _find_data_file_accesses(self, code: str) -> List[str]:
        file_paths: List[str] = []
        assigned_strings: Dict[str, str] = {}

        try:
            tree = ast.parse(code)
        except SyntaxError:
            return []

        for node in ast.walk(tree):
            if isinstance(node, ast.Assign) and isinstance(node.value, ast.Constant) and isinstance(node.value.value, str):
                for target in node.targets:
                    if isinstance(target, ast.Name):
                        assigned_strings[target.id] = node.value.value

            if isinstance(node, ast.Call):
                func_name = self._call_name(node.func)
                # Canonical session paths are documented as directly usable by
                # Python. Discover ordinary open(...) calls as well as the
                # injected helpers so Bubblewrap can stage the authorized file.
                if func_name in {"open", "io.open", "get_raw_data", "load_data"}:
                    if node.args:
                        file_path = self._literal_or_assigned_string(node.args[0], assigned_strings)
                        if file_path:
                            file_paths.append(file_path)

        return list(dict.fromkeys(file_paths))

    def _call_name(self, node: ast.AST) -> str:
        if isinstance(node, ast.Name):
            return node.id
        if isinstance(node, ast.Attribute):
            parent = self._call_name(node.value)
            return f"{parent}.{node.attr}" if parent else node.attr
        return ""

    def _literal_or_assigned_string(self, node: ast.AST, assigned_strings: Dict[str, str]) -> Optional[str]:
        if isinstance(node, ast.Constant) and isinstance(node.value, str):
            return node.value
        if isinstance(node, ast.Name):
            return assigned_strings.get(node.id)
        return None

    def _build_allowed_data_files_payload(self, code: str, context) -> str:
        """Build a small allowlist; the child loads data instead of embedding it in code."""
        session_data_dir = Path(context.data_manager.memory.session.data_dir).resolve()
        available_paths = {
            str(Path(path).resolve())
            for path in getattr(context, "available_file_paths", []) or []
            if path
        }
        requested_paths = []
        for file_path in self._find_data_file_accesses(code):
            resolved = resolve_agent_path(file_path)
            resolved_path = str(resolved)
            is_session_file = resolved == session_data_dir or session_data_dir in resolved.parents
            if resolved_path not in available_paths and not is_session_file:
                logger.warning(
                    "execute_python_data_file_not_in_context",
                    file_path=resolved_path,
                )
                continue
            requested_paths.append(resolved_path)
        return json.dumps(list(dict.fromkeys(requested_paths)), ensure_ascii=False)

    async def execute(
        self,
        context=None,
        code: str = None,
        timeout: Optional[int] = None,
        **kwargs
    ) -> Dict[str, Any]:
        """
        执行 Python 代码

        Args:
            code: Python 代码
            timeout: 超时时间（秒），默认 30 秒

        Returns:
            {
                "success": True/False,
                "data": {
                    "output": "代码输出",
                    "files": ["/path/to/generated/file.docx"],
                    "engine": "subprocess"
                },
                "summary": "执行成功"
            }
        """
        if not code:
            return {
                "status": "failed",  # ✅ 添加 status 字段
                "success": False,
                "error": "Missing required parameter: code",
                "data": None,
                "metadata": {
                    "tool_name": "execute_python",
                    "error_type": "MISSING_PARAMETER"
                },
                "summary": "缺少代码参数"
            }

        # ✅ 确保图表目录存在（每次执行时都检查）
        os.makedirs(self.CHARTS_DIR, exist_ok=True)

        # 创建临时目录
        temp_dir = tempfile.mkdtemp(prefix="python_exec_")

        # 获取 backend 目录（用于相对路径访问数据文件）
        # ✅ 修复：从 execute_python_tool.py 往上 3 级到达 backend/ 目录
        # 文件位置：backend/app/tools/utility/execute_python_tool.py
        # 需要到达：backend/
        backend_dir = os.path.join(os.path.dirname(__file__), "..", "..", "..")
        backend_dir = os.path.abspath(backend_dir)

        try:
            # 只在独立子进程中执行，且不修改 Web worker 的全局 cwd。
            original_code = code
            code = self._inject_data_context(code, context)
            code = self._inject_matplotlib_save_support(code)
            code = self._inject_excel_helpers(code)

            logger.info(
                "code_injection_completed",
                original_code_length=len(original_code),
                injected_code_length=len(code),
                code_modified=(code != original_code),
                mode=self.execution_engine,
                has_context=context is not None,
                available_data_count=len(context.available_file_paths) if context and hasattr(context, 'available_file_paths') else 0,
            )

            result = await self._execute_with_subprocess(
                code,
                timeout or self.default_timeout,
                working_dir=temp_dir,
                context=context,
            )

            # ✅ 从 output 中提取用户保存的文件路径（绝对路径保存的文件）
            output = result["data"].get("output", "")
            user_saved_files = self._extract_file_paths_from_output(output)

            # ✅ 查找临时目录生成的文件（相对路径保存的文件）
            temp_files = self._find_generated_files(temp_dir)

            logger.info(
                "execute_python_file_detection",
                temp_files_count=len(temp_files),
                user_saved_files_count=len(user_saved_files),
                user_saved_files=user_saved_files
            )

            # 移动临时文件到永久目录
            moved_temp_files = self._move_to_permanent_dir(temp_files)

            # 合并文件列表（移动的临时文件 + 用户保存的文件）
            final_files = moved_temp_files + user_saved_files

            result["data"]["files"] = final_files
            result["data"]["engine"] = self.execution_engine
            result["data"]["artifact_schema"] = self._artifact_schema()

            # ✅ DOCX 后处理：优先用同目录 report.html 作为源重建正式 Word；
            # 若没有 HTML 源，再兜底替换图片占位符。
            docx_postprocess = []
            for generated_file in final_files:
                if Path(generated_file).suffix.lower() != ".docx":
                    continue
                try:
                    from app.services.report.government_docx_style import (
                        convert_html_report_to_government_docx,
                        infer_report_html_path_for_docx,
                        replace_image_placeholders_in_docx,
                    )

                    html_path = infer_report_html_path_for_docx(generated_file)
                    if html_path:
                        cleanup = convert_html_report_to_government_docx(html_path, generated_file)
                        logger.info(
                            "execute_python_docx_rebuilt_from_html_report",
                            file=generated_file,
                            html_path=str(html_path),
                            embedded_images=cleanup.get("embedded_images"),
                            missing_images=cleanup.get("missing_images", [])
                        )
                    else:
                        cleanup = replace_image_placeholders_in_docx(generated_file)
                        if cleanup.get("replaced", 0):
                            logger.info(
                                "execute_python_docx_images_embedded",
                                file=generated_file,
                                replaced=cleanup.get("replaced"),
                                missing=cleanup.get("missing", [])
                            )
                    docx_postprocess.append(cleanup)
                except Exception as cleanup_error:
                    logger.warning(
                        "execute_python_docx_postprocess_failed",
                        file=generated_file,
                        error=str(cleanup_error)
                    )
            if docx_postprocess:
                result["data"]["docx_postprocess"] = docx_postprocess

            # ✅ 处理办公文件：生成 PDF 预览（与 Office 工具统一格式）
            office_extensions = {'.docx', '.xlsx', '.pptx', '.pdf', '.doc', '.xls', '.ppt'}
            office_files = [f for f in final_files if Path(f).suffix.lower() in office_extensions]

            if office_files:
                # 只处理第一个 office 文件（与 Office 工具行为一致）
                office_file = office_files[0]
                office_suffix = Path(office_file).suffix.lower()
                if office_suffix in {'.xlsx', '.xls'}:
                    result["data"]["spreadsheet_preview"] = {
                        "file_type": office_suffix.lstrip("."),
                        "editable": True,
                        "size": Path(office_file).stat().st_size,
                    }
                    result["data"]["file_path"] = office_file
                    result["data"]["file_name"] = Path(office_file).name
                    if result.get("success", False):
                        result["summary"] = f"✅ 工具已执行完成，生成文档：{Path(office_file).name}"
                    logger.info(
                        "execute_python_spreadsheet_preview_created",
                        office_file=office_file,
                        execution_success=result.get("success", False)
                    )
                else:
                    try:
                        from app.services.pdf_converter import pdf_converter
                        pdf_preview = await pdf_converter.convert_to_pdf(office_file)
                        result["data"]["pdf_preview"] = pdf_preview
                        result["data"]["file_path"] = office_file
                        result["data"]["file_name"] = Path(office_file).name
                        # ✅ 只在执行成功时覆盖 summary
                        if result.get("success", False):
                            result["summary"] = f"✅ 工具已执行完成，生成文档：{Path(office_file).name}"
                        logger.info(
                            "execute_python_pdf_generated",
                            pdf_id=pdf_preview["pdf_id"],
                            office_file=office_file,
                            execution_success=result.get("success", False)
                        )
                    except Exception as pdf_error:
                        logger.warning("execute_python_pdf_conversion_failed", error=str(pdf_error))
                        # PDF 转换失败时，仍然返回文件信息
                    result["data"]["file_path"] = office_file
                    # ✅ 只在执行成功时覆盖 summary
                    if result.get("success", False):
                        result["summary"] = f"✅ 工具已执行完成，生成文件：{Path(office_file).name}"

            if final_files and "file_path" not in result["data"]:
                # ✅ 只在执行成功时覆盖 summary
                if result.get("success", False):
                    file_names = [Path(f).name for f in final_files]
                    result["summary"] = f"✅ 工具已执行完成，生成文件: {', '.join(file_names)}"
            else:
                # ✅ 只在执行成功时覆盖 summary
                if result.get("success", False):
                    result["summary"] = "✅ 工具已执行完成，计算任务已完成"

            # ✅ 检测图表输出（CHART_SAVED:xxx.png 或 CHART_SAVED:data:image/png;base64,...）
            chart_data = self._extract_chart_paths(result["data"].get("output", ""))

            # 检测 Python 中通过 save_data() 保存的会话数据文件。
            python_data_paths = self._extract_python_data_file_paths(
                result["data"].get("output", "")
            )
            if python_data_paths:
                for file_path in python_data_paths:
                    if file_path not in context.available_file_paths:
                        context.available_file_paths.append(file_path)
                result["data"]["data_file_paths"] = python_data_paths
                result.setdefault("metadata", {})
                result["metadata"]["data_file_paths"] = python_data_paths
                if result.get("success", False):
                    result["summary"] = (
                        f"{result.get('summary', '✅ 工具已执行完成')} | "
                        f"已保存中间结果文件: {', '.join(python_data_paths)}"
                    )

            # ECharts 标准 JSON 只由 execute_echarts_python 专用工具处理。
            echarts_options = (
                self._extract_echarts_formats(result["data"].get("output", ""))
                if self.enable_echarts_visuals
                else []
            )

            logger.info(
                "chart_paths_extracted",
                chart_paths=chart_data.get("paths", []),
                base64_count=len(chart_data.get("base64_data", [])),
                echarts_found=bool(echarts_options),
                echarts_count=len(echarts_options),
                output_preview=result["data"].get("output", "")[:200]
            )

            # 如果检测到图表，自动缓存到 ImageCache
            if chart_data.get("paths") or chart_data.get("base64_data"):
                from app.services.image_cache import ImageCache
                image_cache = ImageCache()

                # 处理文件路径格式的图表
                for chart_path in chart_data.get("paths", []):
                    # ✅ 修复：将相对路径转换为绝对路径（因为工作目录已切换到temp_dir）
                    # Python代码在backend_dir执行，所以相对路径需要基于backend_dir解析
                    if not os.path.isabs(chart_path):
                        abs_chart_path = os.path.abspath(os.path.join(backend_dir, chart_path))
                    else:
                        abs_chart_path = chart_path

                    logger.info(
                        "checking_chart_file",
                        relative_path=chart_path,
                        absolute_path=abs_chart_path,
                        exists=os.path.exists(abs_chart_path)
                    )

                    # 检查文件是否存在（使用绝对路径）
                    if os.path.exists(abs_chart_path):
                        try:
                            logger.info(
                                "chart_file_found",
                                chart_path=chart_path,
                                abs_chart_path=abs_chart_path,
                                file_size=os.path.getsize(abs_chart_path)
                            )

                            with open(abs_chart_path, 'rb') as f:
                                base64_data = base64.b64encode(f.read()).decode('utf-8')

                            # 生成唯一的 chart_id（使用纳秒时间戳避免冲突）
                            chart_id = f"matplotlib_{time.time_ns()}"

                            logger.info(
                                "saving_to_image_cache",
                                chart_id=chart_id,
                                cache_dir=image_cache.cache_dir
                            )

                            image_info = image_cache.save(
                                base64_data=base64_data,
                                chart_id=chart_id
                            )

                            logger.info(
                                "chart_cached",
                                chart_path=chart_path,
                                image_url=image_info["url"],
                                image_id=image_info["image_id"],
                                local_path=image_info["local_path"]
                            )

                            # ✅ 添加到 visuals 字段（顶层，前端渲染使用）
                            result.setdefault("visuals", []).append({
                                "id": chart_id,
                                "type": "image",
                                "title": f"图表 {Path(chart_path).stem}",
                                "data": {
                                    "image_id": image_info["image_id"],
                                    "local_path": image_info["local_path"],  # 图片缓存真实本地路径
                                    "source_file_path": abs_chart_path,  # execute_python 生成的真实图片路径
                                },
                                "meta": {
                                    "generator": "execute_python",
                                    "schema_version": "3.1",
                                    "file_path": abs_chart_path,
                                    "report_asset_hint": {
                                        "path": abs_chart_path,
                                        "type": "image",
                                        "name": Path(chart_path).name,
                                    },
                                }
                            })

                            # 更新摘要
                            result["summary"] = "✅ 工具已执行完成，图表已登记到统一资源目录。"

                        except Exception as e:
                            logger.error(
                                "chart_cache_failed",
                                chart_path=chart_path,
                                abs_chart_path=abs_chart_path,
                                error=str(e),
                                error_type=type(e).__name__,
                                exc_info=True
                            )
                            # ⚠️ 缓存失败时，仍返回文件路径供调试
                            result.setdefault("visuals", []).append({
                                "id": chart_id,
                                "type": "image",
                                "title": f"图表 {Path(chart_path).stem}",
                                "data": {
                                    "file_path": abs_chart_path,
                                    "error": str(e)
                                },
                                "meta": {
                                    "generator": "execute_python",
                                    "schema_version": "3.1",
                                    "cache_failed": True
                                }
                            })

                # 处理 base64 格式的图表
                for base64_data_url in chart_data.get("base64_data", []):
                    try:
                        # 解析 data:image/png;base64,... 格式
                        if "," in base64_data_url:
                            mime_type, base64_data = base64_data_url.split(",", 1)

                            # 生成唯一的 chart_id
                            chart_id = f"matplotlib_{time.time_ns()}"

                            logger.info(
                                "saving_base64_to_image_cache",
                                chart_id=chart_id,
                                mime_type=mime_type,
                                cache_dir=image_cache.cache_dir
                            )

                            image_info = image_cache.save(
                                base64_data=base64_data,
                                chart_id=chart_id
                            )

                            logger.info(
                                "base64_chart_cached",
                                chart_id=chart_id,
                                image_url=image_info["url"],
                                image_id=image_info["image_id"]
                            )

                            # ✅ 添加到 visuals 字段
                            result.setdefault("visuals", []).append({
                                "id": chart_id,
                                "type": "image",
                                "title": f"图表 {chart_id}",
                                "data": {
                                    "image_id": image_info["image_id"],
                                    "local_path": image_info["local_path"],  # 图片缓存真实本地路径
                                    "source_file_path": image_info["local_path"],
                                },
                                "meta": {
                                    "generator": "execute_python",
                                    "schema_version": "3.1",
                                    "source": "base64_output",
                                    "report_asset_hint": {
                                        "path": image_info["local_path"],
                                        "type": "image",
                                        "name": f"{chart_id}.png",
                                    },
                                }
                            })

                            # 更新摘要
                            result["summary"] = "✅ 工具已执行完成，图表已登记到统一资源目录。"

                            # ✅ 从输出中移除 base64 字符串（太长，LLM不需要）
                            output = result["data"].get("output", "")
                            output = output.replace(base64_data_url, "[图表已生成，详见visuals字段]")
                            result["data"]["output"] = output

                    except Exception as e:
                        logger.error(
                            "base64_chart_cache_failed",
                            error=str(e),
                            error_type=type(e).__name__,
                            exc_info=True
                        )
                        result["summary"] = "✅ 工具已执行完成，图表生成成功（缓存失败）"

            # ✅ 新增：处理 ECharts 标准格式 JSON 数据
            if echarts_options:
                visuals = self._build_echarts_visuals(echarts_options, generator="execute_python")
                result.setdefault("visuals", []).extend(visuals)
                if not chart_data.get("paths") and not chart_data.get("base64_data"):
                    result["summary"] = f"✅ 工具已执行完成，ECharts图表生成成功：{len(visuals)} 个"

            from app.tools.resource_declarations import (
                data_file_resource,
                generated_file_products,
                resources_for_visuals,
            )

            visual_source_paths: set[str] = set()
            for visual in result.get("visuals") or []:
                if not isinstance(visual, dict):
                    continue
                visual_data = visual.get("data") if isinstance(visual.get("data"), dict) else {}
                visual_meta = visual.get("meta") if isinstance(visual.get("meta"), dict) else {}
                for path in (
                    visual.get("local_path"),
                    visual.get("file_path"),
                    visual_data.get("local_path"),
                    visual_data.get("file_path"),
                    visual_data.get("source_file_path"),
                    visual_meta.get("local_path"),
                    visual_meta.get("file_path"),
                ):
                    if isinstance(path, str) and path:
                        visual_source_paths.add(str(Path(path).expanduser().resolve()))

            deliverable_files = [
                path for path in final_files
                if str(Path(path).expanduser().resolve()) not in visual_source_paths
            ]
            pdf_preview = result["data"].get("pdf_preview")
            preview_paths: dict[str, str] = {}
            primary_file_path = result["data"].get("file_path")
            if (
                isinstance(pdf_preview, dict)
                and isinstance(pdf_preview.get("pdf_path"), str)
                and isinstance(primary_file_path, str)
            ):
                preview_paths[primary_file_path] = pdf_preview["pdf_path"]

            result["resources"] = generated_file_products(
                deliverable_files,
                tool_name=self.name,
                preview_paths=preview_paths,
            )
            result["resources"].extend(
                data_file_resource(file_path, tool_name=self.name)
                for file_path in python_data_paths
            )
            if not self.enable_echarts_visuals:
                result["resources"].extend(
                    resources_for_visuals(result.get("visuals", []), tool_name=self.name)
                )
            # Resource declarations above deliberately retain absolute host paths.
            # Only after they are complete do we expose the canonical Agent-facing
            # project-relative paths and the automatic-publication contract.
            self._attach_resume_context(result)
            return result

        except Exception as e:
            logger.error("execute_python_failed", error=str(e), exc_info=True)
            return {
                "status": "failed",  # ✅ 添加 status 字段
                "success": False,
                "error": str(e),
                "data": {"error": str(e)},
                "metadata": {
                    "tool_name": "execute_python",
                    "error_type": type(e).__name__
                },
                "summary": f"执行失败: {str(e)}"
            }
        finally:
            # 清理临时目录
            shutil.rmtree(temp_dir, ignore_errors=True)

    async def _execute_with_subprocess(
        self,
        code: str,
        timeout: int,
        *,
        working_dir: str,
        context=None,
    ) -> Dict[str, Any]:
        """在独立沙箱进程组中执行代码，同时保持 Web worker 事件循环可运行。"""
        # 写入脚本文件
        script_file = os.path.join(working_dir, "script.py")
        with open(script_file, 'w', encoding='utf-8') as f:
            f.write(code)

        if self.execution_engine != "bubblewrap":
            return {
                "status": "failed",
                "success": False,
                "data": {"error": f"不安全或未支持的 Python 执行引擎: {self.execution_engine}"},
                "summary": "Python 沙箱配置错误",
            }

        sandbox_spec = self._build_bubblewrap_command(
            code=code,
            script_file=script_file,
            working_dir=working_dir,
            timeout=timeout,
            context=context,
        )
        if sandbox_spec is None:
            return {
                "status": "failed",
                "success": False,
                "data": {"error": "Bubblewrap 沙箱不可用，已拒绝在宿主机直接执行代码"},
                "summary": "Python 沙箱不可用",
            }
        command, sandbox_sync_dirs, relative_input_mounts = sandbox_spec

        # 执行代码
        process = None
        try:
            process = subprocess.Popen(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                cwd=working_dir,
                start_new_session=True,
            )
            logger.info(
                "execute_python_subprocess_started",
                pid=process.pid,
                script_size=len(code),
                timeout_seconds=timeout,
                sandbox=self.execution_engine,
            )
            stdout, stderr = await asyncio.to_thread(process.communicate, timeout=timeout)
            self._sync_sandbox_session_outputs(sandbox_sync_dirs)
            logger.info(
                "execute_python_subprocess_completed",
                pid=process.pid,
                returncode=process.returncode,
                stdout_size=len(stdout or ""),
                stderr_size=len(stderr or ""),
            )

            output = stdout or ""
            if stderr:
                output += f"\n错误输出:\n{stderr}"

            # 截断输出
            if len(output) > self.max_output_size:
                output = output[:self.max_output_size] + "\n... (输出被截断)"

            # 如果执行失败，尝试解析错误信息
            if process.returncode != 0:
                error_info = self._parse_subprocess_error(stderr, code)
                return {
                    "status": "failed",
                    "success": False,
                    "data": {"error": error_info["error_message"], "error_details": error_info, "output": output},
                    "summary": error_info["summary"]
                }

            return {
                "status": "success",
                "success": True,
                "data": {"output": output},
                "summary": "✅ 工具已执行完成，计算任务已完成"
            }

        except subprocess.TimeoutExpired:
            if process is not None:
                logger.warning("execute_python_subprocess_timeout", pid=process.pid)
                self._terminate_process_group(process)
                await asyncio.to_thread(process.communicate)
            return {
                "status": "failed",
                "success": False,
                "data": {"error": "执行超时"},
                "summary": "执行超时"
            }
        except asyncio.CancelledError:
            if process is not None and process.poll() is None:
                self._terminate_process_group(process)
                await asyncio.to_thread(process.communicate)
            raise
        finally:
            for mountpoint in sorted(relative_input_mounts, key=lambda item: len(item.parts), reverse=True):
                try:
                    if mountpoint.is_dir():
                        mountpoint.rmdir()
                    else:
                        mountpoint.unlink(missing_ok=True)
                except OSError:
                    logger.warning("execute_python_relative_input_mount_cleanup_failed", path=str(mountpoint))

    def _build_bubblewrap_command(
        self,
        *,
        code: str,
        script_file: str,
        working_dir: str,
        timeout: int,
        context=None,
    ) -> Optional[Tuple[List[str], List[Tuple[Path, Path, set[str]]], List[Path]]]:
        """构建 fail-closed Bubblewrap 沙箱命令。

        沙箱不可见项目根目录、宿主机 home、Docker socket 和网络；
        仅会话数据目录可写，其他显式输入文件只读。
        """
        bwrap = shutil.which("bwrap")
        prlimit = shutil.which("prlimit")
        if not bwrap or not prlimit:
            logger.error(
                "execute_python_sandbox_dependency_missing",
                bubblewrap=bool(bwrap),
                prlimit=bool(prlimit),
            )
            return None

        python_env = Path(sys.prefix).resolve()
        required_mounts = [Path("/usr"), Path("/lib"), Path("/lib64"), python_env]
        if any(not path.exists() for path in required_mounts):
            logger.error(
                "execute_python_sandbox_runtime_mount_missing",
                mounts=[str(path) for path in required_mounts if not path.exists()],
            )
            return None

        command = [
            prlimit,
            f"--cpu={max(1, timeout + 2)}",
            "--as=2147483648",
            "--fsize=104857600",
            "--nofile=256",
            "--nproc=128",
            "--core=0",
            "--",
            bwrap,
            "--die-with-parent",
            "--new-session",
            "--unshare-all",
            "--cap-drop",
            "ALL",
            "--ro-bind",
            "/usr",
            "/usr",
            "--ro-bind",
            "/lib",
            "/lib",
            "--ro-bind",
            "/lib64",
            "/lib64",
            "--dir",
            "/root",
            "--dir",
            "/root/miniconda3",
            "--dir",
            "/root/miniconda3/envs",
            "--ro-bind",
            str(python_env),
            str(python_env),
            "--proc",
            "/proc",
            "--dev",
            "/dev",
            "--tmpfs",
            "/tmp",
            "--bind",
            working_dir,
            "/sandbox",
            "--chdir",
            "/sandbox",
        ]

        # 不把 API key、代理、数据库密码等 Web worker 环境变量传给 Agent 代码。
        for name in os.environ:
            command.extend(["--unsetenv", name])
        safe_env = {
            "PATH": f"{python_env}/bin:/usr/bin:/bin",
            "HOME": "/tmp",
            "TMPDIR": "/tmp",
            "MPLCONFIGDIR": "/tmp/matplotlib",
            "XDG_CACHE_HOME": "/tmp/cache",
            "PYTHONDONTWRITEBYTECODE": "1",
            "LANG": "C.UTF-8",
            "LC_ALL": "C.UTF-8",
        }
        for name, value in safe_env.items():
            command.extend(["--setenv", name, value])

        session_data_dir = None
        if context is not None:
            try:
                session_data_dir = Path(context.data_manager.memory.session.data_dir).resolve()
            except (AttributeError, TypeError):
                session_data_dir = None

        mounted_destinations = set()
        sync_dirs: List[Tuple[Path, Path, set[str]]] = []
        relative_input_mounts: List[Path] = []
        context_paths = list(getattr(context, "available_file_paths", []) or [])
        requested_paths = self._find_data_file_accesses(code)
        candidate_paths = list(dict.fromkeys([*context_paths, *requested_paths]))
        try:
            if session_data_dir and session_data_dir.is_dir():
                staged_session_dir = Path(working_dir) / "session_data"
                staged_session_dir.mkdir(parents=True, exist_ok=True)
                original_relative_paths: set[str] = set()
                for value in candidate_paths:
                    source_path = resolve_agent_path(value)
                    if not source_path.is_file() or not (
                        source_path == session_data_dir or session_data_dir in source_path.parents
                    ):
                        continue
                    relative_path = source_path.relative_to(session_data_dir)
                    staged_path = staged_session_dir / relative_path
                    staged_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(source_path, staged_path)
                    original_relative_paths.add(str(relative_path))

                self._append_bubblewrap_parent_dirs(command, session_data_dir)
                command.extend(["--bind", str(staged_session_dir), str(session_data_dir)])
                mounted_destinations.add(str(session_data_dir))
                sync_dirs.append((staged_session_dir, session_data_dir, original_relative_paths))

            # Matplotlib helpers write to the canonical images path inside the
            # sandbox. Bind a staging directory there, then publish only the
            # generated files after the subprocess exits successfully.
            images_dir = Path(get_images_dir()).resolve()
            staged_images_dir = Path(working_dir) / "images_output"
            staged_images_dir.mkdir(parents=True, exist_ok=True)
            self._append_bubblewrap_parent_dirs(command, images_dir)
            command.extend(["--bind", str(staged_images_dir), str(images_dir)])
            mounted_destinations.add(str(images_dir))
            sync_dirs.append((staged_images_dir, images_dir, set()))

            for index, value in enumerate(candidate_paths):
                raw_path = Path(value).expanduser()
                path = resolve_agent_path(value)
                if not path.exists():
                    continue
                if (
                    raw_path.is_absolute()
                    and session_data_dir
                    and (path == session_data_dir or session_data_dir in path.parents)
                ):
                    continue
                if raw_path.is_absolute():
                    destination_path = path
                else:
                    try:
                        project_relative = path.relative_to(PROJECT_ROOT)
                    except ValueError:
                        continue
                    destination_path = Path("/sandbox") / project_relative
                    host_mountpoint = Path(working_dir) / project_relative
                    host_mountpoint.parent.mkdir(parents=True, exist_ok=True)
                    if path.is_dir():
                        host_mountpoint.mkdir(exist_ok=True)
                    else:
                        host_mountpoint.touch(exist_ok=True)
                    relative_input_mounts.append(host_mountpoint)
                destination = str(destination_path)
                if destination in mounted_destinations:
                    continue
                staged_input_dir = Path(working_dir) / "inputs"
                staged_input_dir.mkdir(parents=True, exist_ok=True)
                staged_input = staged_input_dir / f"{index}_{path.name}"
                if path.is_dir():
                    shutil.copytree(path, staged_input)
                else:
                    shutil.copy2(path, staged_input)
                if raw_path.is_absolute():
                    self._append_bubblewrap_parent_dirs(command, path.parent)
                command.extend(["--ro-bind", str(staged_input), destination])
                mounted_destinations.add(destination)
        except OSError as error:
            logger.error(
                "execute_python_sandbox_mount_open_failed",
                error=str(error),
                error_type=type(error).__name__,
            )
            return None

        command.extend([str(python_env / "bin/python"), "/sandbox/script.py"])
        return command, sync_dirs, relative_input_mounts

    @staticmethod
    def _sync_sandbox_session_outputs(
        sync_dirs: List[Tuple[Path, Path, set[str]]],
    ) -> None:
        """只发布沙箱新生成的会话文件，不允许代码覆盖原始输入。"""
        for staged_dir, destination_dir, original_relative_paths in sync_dirs:
            for staged_path in staged_dir.rglob("*"):
                if not staged_path.is_file():
                    continue
                relative_path = staged_path.relative_to(staged_dir)
                if str(relative_path) in original_relative_paths:
                    continue
                destination = destination_dir / relative_path
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(staged_path, destination)

    @staticmethod
    def _append_bubblewrap_parent_dirs(command: List[str], destination: Path) -> None:
        """在空 mount namespace 中创建绑定挂载的父目录。"""
        parents = list(destination.parents)
        for parent in reversed(parents):
            parent_text = str(parent)
            if parent_text in {"/", "/usr", "/lib", "/lib64", "/root"}:
                continue
            command.extend(["--dir", parent_text])
        command.extend(["--dir", str(destination)])

    @staticmethod
    def _terminate_process_group(process: subprocess.Popen) -> None:
        """终止工具进程及其派生的所有子进程。"""
        try:
            os.killpg(process.pid, signal.SIGKILL)
        except ProcessLookupError:
            pass

    def _find_generated_files(self, temp_dir: str) -> list:
        """查找生成的文件（排除临时文件）"""
        generated_files = []
        for root, dirs, files in os.walk(temp_dir):
            if Path(root).resolve() == Path(temp_dir).resolve():
                # Bubblewrap 的输入副本和会话输出镜像由沙箱发布逻辑管理，
                # 不得被误认为用户在 cwd 生成的普通文件。
                dirs[:] = [
                    name
                    for name in dirs
                    if name not in {"inputs", "session_data", "images_output"}
                ]
            # 跳过 __pycache__ 目录
            if '__pycache__' in dirs:
                dirs.remove('__pycache__')

            for file in files:
                # 排除脚本文件和缓存文件
                if file not in ['script.py'] and not file.endswith('.pyc'):
                    file_path = os.path.join(root, file)
                    # 只包含文件，不包含目录
                    if os.path.isfile(file_path):
                        generated_files.append(file_path)
        return generated_files

    def _attach_resume_context(self, result: Dict[str, Any]) -> None:
        if not isinstance(result, dict) or not result.get("success"):
            return

        data = result.get("data") if isinstance(result.get("data"), dict) else {}
        absolute_file_paths = []
        for path in data.get("files") or []:
            if path:
                absolute_file_paths.append(str(resolve_agent_path(path)))
        if data.get("file_path"):
            absolute_file_paths.insert(0, str(resolve_agent_path(data["file_path"])))
        absolute_file_paths = list(dict.fromkeys(absolute_file_paths))
        file_paths = [format_agent_path(path) for path in absolute_file_paths]

        file_refs = []
        for absolute_path, agent_path in zip(absolute_file_paths, file_paths):
            path_obj = Path(absolute_path)
            file_refs.append(
                build_file_ref(
                    agent_path,
                    type=self._resource_type_for_path(path_obj),
                    format=path_obj.suffix.lstrip(".").lower() or None,
                    size=path_obj.stat().st_size if path_obj.exists() else None,
                    usage="generated_file",
                    preferred_for=["read_file", "list_session_resources"],
                )
            )

        data_file_paths = [
            format_agent_path(file_path)
            for file_path in (data.get("data_file_paths") or [])
            if file_path
        ]
        data_refs = [
            build_data_file_ref(file_path, usage="generated")
            for file_path in data_file_paths
        ]

        visual_refs = []
        for visual in result.get("visuals") or []:
            if not isinstance(visual, dict):
                continue
            visual_data = visual.get("data") if isinstance(visual.get("data"), dict) else {}
            visual_meta = visual.get("meta") if isinstance(visual.get("meta"), dict) else {}
            image_url = visual.get("image_url") or visual_data.get("url") or visual_data.get("image_url") or visual_meta.get("image_url")
            local_path = visual.get("local_path") or visual_data.get("local_path") or visual_meta.get("local_path")
            file_path = visual.get("file_path") or visual_data.get("source_file_path") or visual_data.get("file_path") or visual_meta.get("file_path")
            visual_refs.append(
                build_visual_ref(
                    id=visual.get("id"),
                    type=visual.get("type"),
                    title=visual.get("title"),
                    image_url=image_url,
                    local_path=local_path,
                    file_path=file_path,
                )
            )

        refs = {}
        if file_refs:
            refs["files"] = file_refs
        if data_refs:
            refs["data"] = data_refs
        visual_refs = [ref for ref in visual_refs if ref]
        if visual_refs:
            refs["visuals"] = visual_refs
        if refs:
            result["refs"] = merge_refs(result.get("refs"), refs)

        llm_resume: Dict[str, Any] = {}
        if file_paths:
            llm_resume["generated_files"] = file_paths
            llm_resume["auto_published"] = True
            llm_resume["publish_session_file_required"] = False
            llm_resume["tool_hint"] = (
                "Generated files have already been published to the unified session resource catalog. "
                "Do not call publish_session_file and do not construct a sessions/... path. "
                "Reuse the returned file_path exactly, or use list_session_resources to inspect the resource IDs."
            )
        if data_refs:
            llm_resume["file_paths"] = [ref["file_path"] for ref in data_refs]
        if visual_refs and "tool_hint" not in llm_resume:
            first_visual_path = visual_refs[0].get("tool_path")
            if first_visual_path:
                llm_resume["tool_hint"] = (
                    f"Use read_file(path='{first_visual_path}', as_multimodal_attachment=true) "
                    "to inspect this image."
                )
        if llm_resume:
            result["llm_resume"] = llm_resume

        if file_paths:
            result["file_path"] = file_paths[0]
            result["file_paths"] = file_paths
            data["files"] = file_paths
            if data.get("file_path"):
                data["file_path"] = format_agent_path(data["file_path"])
            data["generated_artifacts"] = [
                {
                    "file_path": path,
                    "file_name": Path(path).name,
                    "auto_published": True,
                    "publish_session_file_required": False,
                }
                for path in file_paths
            ]
        if data_file_paths:
            data["data_file_paths"] = data_file_paths
        pdf_preview = data.get("pdf_preview")
        if isinstance(pdf_preview, dict) and pdf_preview.get("pdf_path"):
            pdf_preview["pdf_path"] = format_agent_path(pdf_preview["pdf_path"])

    def _resource_type_for_path(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg"}:
            return "image"
        if suffix in {".pdf", ".doc", ".docx", ".ppt", ".pptx", ".xls", ".xlsx", ".qmd", ".md", ".html", ".htm"}:
            return "document"
        if suffix in {".json", ".csv", ".parquet", ".xlsx", ".xls"}:
            return "data"
        return "file"

    def _move_to_permanent_dir(self, files: list) -> list:
        """移动文件到永久目录"""
        final_files = []
        for file_path in files:
            file_name = os.path.basename(file_path)
            permanent_path = os.path.join(self.PERMANENT_DIR, file_name)

            # 如果文件已存在，添加时间戳
            if os.path.exists(permanent_path):
                name, ext = os.path.splitext(file_name)
                permanent_path = os.path.join(
                    self.PERMANENT_DIR,
                    f"{name}_{int(time.time())}{ext}"
                )

            # 使用 shutil.move 代替 os.rename（跨文件系统支持）
            shutil.move(file_path, permanent_path)
            final_files.append(permanent_path)

        return final_files

    def _artifact_schema(self) -> Dict[str, Any]:
        """Return a compact schema note for generated artifacts."""
        return {
            "version": "execute_python.artifacts.v2",
            "files": "All generated local files as project-relative paths; backend paths include the backend/ prefix.",
            "file_path": "Primary generated file for preview/download. Reuse exactly; never construct a sessions/... path.",
            "publication": "Generated files are published automatically; publish_session_file is not required.",
            "pdf_preview": "Office/PDF preview metadata for docx/xlsx/pptx/pdf artifacts.",
            "visuals": "Image/ECharts blocks used to declare durable session resources.",
        }

    def _extract_chart_paths(self, output: str) -> dict:
        """
        从 Python 代码输出中提取图表路径和base64数据

        检测格式：
        1. CHART_SAVED:/path/to/chart.png (标准格式)
        2. 图表已保存: /path/to/chart.png (中文格式)
        3. 图表已保存到: /path/to/chart.png (中文格式2)
        4. 保存成功: /path/to/chart.png (中文格式3)
        5. CHART_SAVED:data:image/png;base64,... (base64数据)

        Args:
            output: Python 代码输出

        Returns:
            {"paths": [文件路径列表], "base64_data": [base64数据列表]}
        """
        import re
        result = {"paths": [], "base64_data": []}
        output_lines = output.split('\n')

        # 图表文件扩展名
        image_extensions = r'\.(?:png|jpg|jpeg|svg|pdf|gif|bmp|tiff)'

        for line in output_lines:
            line = line.strip()

            # 检测标准格式
            if line.startswith("CHART_SAVED:"):
                chart_data = line.split("CHART_SAVED:")[1].strip()

                if chart_data.startswith("data:image/"):
                    result["base64_data"].append(chart_data)
                else:
                    result["paths"].append(chart_data)

            # 检测中文输出格式（智能匹配）
            else:
                # 常见的中文图表保存模式
                chinese_patterns = [
                    # 图表已保存: /path/to/chart.png
                    rf'(?:图表已保存|图表已保存到|保存成功|图片已保存|图片已保存到|已生成图表)[:：]\s*(.+?{image_extensions})',
                    # 保存图表到 /path/to/chart.png
                    rf'(?:保存图表到|生成图表到|图表保存到)[:：]\s*(.+?{image_extensions})',
                    # 直接以文件路径结尾（带图表扩展名）
                    rf'^(.+?{image_extensions})\s*(?:$|已保存|生成完成)',
                ]

                for pattern in chinese_patterns:
                    matches = re.findall(pattern, line)
                    for match in matches:
                        if match and match not in result["paths"]:
                            result["paths"].append(match)

        return result

    def _extract_python_data_file_paths(self, output: str) -> List[str]:
        """Extract paths printed by the injected save_data() helper."""
        if not output:
            return []
        refs: List[str] = []
        for match in re.findall(r"PYTHON_DATA_FILE_SAVED:([^\s]+\.json)", output):
            if match and match not in refs:
                refs.append(match)
        return refs

    def _extract_echarts_format(self, output: str) -> dict:
        """Backward-compatible single-option extractor."""
        options = self._extract_echarts_formats(output)
        return options[0] if options else None

    def _extract_echarts_formats(self, output: str) -> List[Dict[str, Any]]:
        """
        从 Python 代码输出中提取一个或多个 ECharts 标准格式 JSON 数据

        检测格式：
        1. 标准格式：JSON字符串包含 series 字段（ECharts标准格式）
        2. 嵌套格式：包含 echarts_option 字段（兼容某些自定义格式）
        可能包含：xAxis, yAxis, series, tooltip, legend 等字段

        Args:
            output: Python 代码输出

        Returns:
            ECharts 配置字典列表；每个有效 JSON 行对应一个图表
        """
        import json

        options: List[Dict[str, Any]] = []
        output_lines = output.split('\n')

        for line in output_lines:
            line = line.strip()
            if not line:
                continue

            try:
                # 尝试解析 JSON
                chart_data = json.loads(line)

                # ✅ 检测方式1：标准 ECharts 格式（顶层有 series 字段）
                if isinstance(chart_data, dict) and "series" in chart_data:
                    # series 必须是数组类型
                    if isinstance(chart_data["series"], list):
                        logger.info(
                            "echarts_format_detected",
                            format="standard",
                            chart_type=chart_data.get("series", [{}])[0].get("type", "unknown") if len(chart_data.get("series", [])) > 0 else "unknown",
                            has_xAxis="xAxis" in chart_data,
                            has_yAxis="yAxis" in chart_data,
                            series_count=len(chart_data.get("series", []))
                        )
                        options.append(chart_data)
                        continue

                # ✅ 检测方式2：嵌套格式（包含 echarts_option 字段）
                if isinstance(chart_data, dict) and "echarts_option" in chart_data:
                    echarts_option = chart_data["echarts_option"]
                    if isinstance(echarts_option, dict) and "series" in echarts_option:
                        if isinstance(echarts_option["series"], list):
                            logger.info(
                                "echarts_format_detected",
                                format="nested",
                                chart_type=echarts_option.get("series", [{}])[0].get("type", "unknown") if len(echarts_option.get("series", [])) > 0 else "unknown",
                                has_xAxis="xAxis" in echarts_option,
                                has_yAxis="yAxis" in echarts_option,
                                series_count=len(echarts_option.get("series", [])),
                                original_fields=list(chart_data.keys())
                            )
                            # ✅ 直接返回 echarts_option（标准 ECharts 格式）
                            options.append(echarts_option)
                            continue

                # ✅ 检测方式3：嵌套格式（包含 data 字段，data 内有 series）
                if isinstance(chart_data, dict) and "data" in chart_data:
                    inner_data = chart_data["data"]
                    if isinstance(inner_data, dict) and "series" in inner_data:
                        if isinstance(inner_data["series"], list):
                            logger.info(
                                "echarts_format_detected",
                                format="data_nested",
                                chart_type=inner_data.get("series", [{}])[0].get("type", "unknown") if len(inner_data.get("series", [])) > 0 else "unknown",
                                has_xAxis="xAxis" in inner_data,
                                has_yAxis="yAxis" in inner_data,
                                series_count=len(inner_data.get("series", [])),
                                original_fields=list(chart_data.keys())
                            )
                            # ✅ 直接返回 inner_data（标准 ECharts 格式）
                            options.append(inner_data)
                            continue

            except (json.JSONDecodeError, ValueError):
                # 不是有效的 JSON，继续下一行
                continue

        return options

    def _build_echarts_visuals(
        self,
        echarts_options: List[Dict[str, Any]],
        *,
        generator: str,
    ) -> List[Dict[str, Any]]:
        """Convert parsed ECharts options into frontend visuals."""
        visuals: List[Dict[str, Any]] = []
        for index, echarts_data in enumerate(echarts_options):
            try:
                chart_type = self._detect_echarts_chart_type(echarts_data)
                display_title = self._detect_echarts_title(echarts_data, chart_type)
                visuals.append({
                    "id": f"echarts_{time.time_ns()}_{index}",
                    "type": chart_type,
                    "title": display_title,
                    "data": echarts_data,
                    "meta": {
                        "generator": generator,
                        "schema_version": "echarts_standard"
                    }
                })
                logger.info(
                    "echarts_format_added",
                    chart_type=chart_type,
                    display_title=display_title,
                    has_title_text="title" in echarts_data,
                    data_format=list(echarts_data.keys()) if isinstance(echarts_data, dict) else type(echarts_data).__name__
                )
            except Exception as e:
                logger.warning(
                    "echarts_processing_failed",
                    error=str(e),
                    echarts_data=echarts_data
                )
        return visuals

    def _detect_echarts_chart_type(self, echarts_data: Dict[str, Any]) -> str:
        series_list = echarts_data.get("series")
        if isinstance(series_list, list) and len(series_list) > 0:
            first_series = series_list[0]
            if isinstance(first_series, dict):
                series_type = first_series.get("type", "chart").lower()
                if first_series.get("coordinateSystem") == "polar":
                    return f"polar_{series_type}"
                return series_type
        return "chart"

    def _detect_echarts_title(self, echarts_data: Dict[str, Any], chart_type: str) -> str:
        echarts_title = echarts_data.get("title", {})
        if isinstance(echarts_title, dict):
            return echarts_title.get("text", f"{chart_type.upper()}图表")
        return echarts_title or f"{chart_type.upper()}图表"

    def _inject_data_context(self, code: str, context) -> str:
        """
        注入基于会话文件路径的数据访问上下文。

        注入内容：
        - load_data(file_path): 获取原始数据（字典列表格式）
        - save_data(data, schema="python_result", metadata=None): 保存中间结果并返回 file_path

        ⚠️ 重要：
        - 每次 execute_python 都是独立执行环境，不保留上次调用的变量
        - 跨工具调用复用的数据必须先 save_data，再在后续调用中 load_data(file_path)
        """
        # 检查 context 是否存在
        if not context:
            logger.debug("data_context_injection_skipped", reason="no_context")
            return code

        logger.info(
            "data_context_injection_started",
            has_data_manager=context.data_manager is not None
        )

        allowed_data_files_payload = self._build_allowed_data_files_payload(code, context)
        session_data_dir = json.dumps(str(context.data_manager.memory.session.data_dir), ensure_ascii=False)
        session_prefix = json.dumps(
            str(context.data_manager.memory.session.data_dir), ensure_ascii=False
        )

        # 构建注入的代码
        context_injection_code = '''# ===== 数据访问上下文（自动注入） =====
# 重要：每次 execute_python 都是独立执行环境，不保留上次调用的变量。
# 使用 load_data(file_path) 读取工具返回的数据文件。

import json as __data_context_json
from pathlib import Path as __DataContextPath

__ALLOWED_DATA_FILES__ = set(__ALLOWED_DATA_FILES_PAYLOAD__)

# 获取原始数据（字典列表格式）
def load_data(file_path: str):
    """读取当前会话中的数据文件。"""
    resolved_path = str(__DataContextPath(file_path).resolve())
    if resolved_path not in __ALLOWED_DATA_FILES__:
        raise RuntimeError(f"未找到会话数据文件: {file_path}")
    with open(resolved_path, 'r', encoding='utf-8') as data_file:
        return __data_context_json.load(data_file)

get_raw_data = load_data

def save_data(data, schema: str = 'python_result', metadata=None, version: str = 'v1'):
    """保存 Python 中间结果到会话数据文件并返回 file_path。

    适用于跨多次 execute_python 调用复用的变量、DataFrame 转换结果、核验表等。
    data 可以是 list[dict]、dict、pandas DataFrame 或其他可 JSON 序列化对象。
    """
    import json
    from pathlib import Path
    import re
    import uuid

    if metadata is None:
        metadata = {}
    metadata = dict(metadata)
    metadata.setdefault('generator', 'execute_python')
    metadata.setdefault('created_by', 'save_data_helper')

    try:
        import pandas as pd
        if isinstance(data, pd.DataFrame):
            payload = data.to_dict(orient='records')
        else:
            payload = data
    except Exception:
        payload = data

    def _json_default(value):
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        if hasattr(value, 'item'):
            return value.item()
        return str(value)

    payload = json.loads(json.dumps(payload, ensure_ascii=False, default=_json_default))

    safe_schema = re.sub(r'[^A-Za-z0-9._-]+', '_', str(schema)).strip('._') or 'python_result'
    filename = f"{safe_schema}--{uuid.uuid4().hex}.json"
    absolute_path = Path(__SESSION_DATA_DIR__) / filename
    absolute_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    file_path = __SESSION_PREFIX__ + "/" + filename
    print(f"PYTHON_DATA_FILE_SAVED:{file_path}")
    return file_path

# ===== 数据访问上下文注入完成 =====

'''
        context_injection_code = context_injection_code.replace(
            "__ALLOWED_DATA_FILES_PAYLOAD__",
            allowed_data_files_payload,
        )
        context_injection_code = context_injection_code.replace("__SESSION_DATA_DIR__", session_data_dir)
        context_injection_code = context_injection_code.replace("__SESSION_PREFIX__", session_prefix)

        # 在代码开头插入上下文代码
        injected_code = context_injection_code + code

        logger.info(
            "data_context_injection_completed",
            original_length=len(code),
            injected_length=len(injected_code)
        )

        return injected_code

    def _convert_unicode_subscript_to_latex(self, code: str) -> str:
        """
        自动转换Python代码字符串中的Unicode下标/上标字符为LaTeX格式

        策略：
        1. 使用正则表达式匹配所有字符串字面量
        2. 在字符串中检测Unicode下标/上标字符
        3. 转换为简写格式（避免LaTeX与中文混合问题）
        4. 智能处理多位数字（如2.5）

        ⚠️ 重要变更：
        由于matplotlib的mathtext引擎不支持中文字符，会导致混合LaTeX和中文的
        字符串显示异常。因此，我们采用简写格式替代LaTeX格式：
        - PM₂.₅ → PM2.5（而非 PM$_{2.5}$）
        - μg/m³ → ug/m3（而非 $\\mu$g/m$^3$）

        转换示例：
        - 'PM₂.₅浓度' → 'PM2.5浓度'
        - 'μg/m³' → 'ug/m3'
        - 'NO₂' → 'NO2'
        """
        import re

        def convert_string_content(content: str) -> tuple:
            """
            转换字符串内容

            Returns:
                (converted_content, needs_r_prefix)
            """
            # 定义常见模式的替换规则（使用简写格式，避免LaTeX与中文混合问题）
            replacements = [
                # μ符号
                (r'μ', r'u'),

                # PM2.5（特殊处理：小数形式）
                (r'PM₂\.?₅', r'PM2.5'),
                (r'pm₂\.?₅', r'pm2.5'),

                # 化学式下标（单个数字）- 简写格式
                (r'O₃', r'O3'),
                (r'NO₂', r'NO2'),
                (r'SO₂', r'SO2'),
                (r'CO₂', r'CO2'),
                (r'CH₄', r'CH4'),
                (r'N₂O', r'N2O'),
                (r'VOCs', r'VOCs'),  # 保持原样

                # 单位上标 - 简写格式
                (r'm³', r'm3'),
                (r'm²', r'm2'),
                (r'km²', r'km2'),
                (r'km³', r'km3'),
                (r'μg', r'ug'),

                # 其他上标数字（需要小数点前面的m等）
                (r'/m³', r'/m3'),
                (r'/m²', r'/m2'),
            ]

            new_content = content
            needs_r = False

            # 应用替换规则
            for old, new in replacements:
                if old in new_content:
                    new_content = new_content.replace(old, new)
                    needs_r = True

            # 处理剩余的Unicode下标/上标（兜底）
            subscript_digits = {
                '₀': '0', '₁': '1', '₂': '2', '₃': '3', '₄': '4',
                '₅': '5', '₆': '6', '₇': '7', '₈': '8', '₉': '9',
            }
            superscript_digits = {
                '⁰': '0', '¹': '1', '²': '2', '³': '3', '⁴': '4',
                '⁵': '5', '⁶': '6', '⁷': '7', '⁸': '8', '⁹': '9',
            }

            # 检查是否还有未处理的下标
            has_remaining = any(c in new_content for c in subscript_digits.keys())
            has_remaining = has_remaining or any(c in new_content for c in superscript_digits.keys())

            if has_remaining:
                # 对于剩余的Unicode字符，使用简写格式替换
                for unicode_char, normal_char in subscript_digits.items():
                    if unicode_char in new_content:
                        new_content = new_content.replace(unicode_char, normal_char)
                        needs_r = True

                for unicode_char, normal_char in superscript_digits.items():
                    if unicode_char in new_content:
                        new_content = new_content.replace(unicode_char, normal_char)
                        needs_r = True

            return new_content, needs_r

        def process_match(match):
            """处理正则匹配的字符串"""
            full_match = match.group(0)

            # 提取引号类型和内容
            if full_match.startswith("'''") or full_match.startswith('"""'):
                # 三引号字符串，暂不处理
                return full_match
            elif full_match.startswith("'"):
                quote = "'"
                content = full_match[1:-1]
            elif full_match.startswith('"'):
                quote = '"'
                content = full_match[1:-1]
            else:
                return full_match

            # 转换内容（简写格式，不需要r前缀）
            converted_content, needs_r = convert_string_content(content)

            # 直接返回转换后的内容（不需要r前缀）
            return f'{quote}{converted_content}{quote}'

        # 匹配Python字符串字面量（排除已经有r前缀的）
        # 使用负向后查找排除r前缀
        string_pattern = r"""
            (?<![rR])                      # 前面不能有r或R
            (?:
                '''(?:[^\\]|\\.)*?'''     # 三单引号
                |\"\"\"(?:[^\\]|\\.)*?\"\"\"  # 三双引号
                |'(?:[^'\\]|\\.)*'        # 单引号
                |"(?:[^"\\]|\\.)*"        # 双引号
            )
        """
        string_pattern = re.compile(string_pattern, re.VERBOSE | re.DOTALL)

        # 执行转换
        converted_code = string_pattern.sub(process_match, code)

        # 统计转换次数
        original_matches = list(string_pattern.finditer(code))
        conversion_count = 0
        for m in original_matches:
            content = m.group(0)
            if any(c in content for c in ['₂', '₃', '⁵', '⁴', '⁶', '⁷', '⁸', '⁹', '⁰',
                                          '²', '³', '¹', '⁴', '⁵', '⁶', '⁷', '⁸', '⁹', '⁰', 'μ']):
                conversion_count += 1

        if conversion_count > 0:
            logger.info(
                "unicode_subscript_converted",
                strings_converted=conversion_count
            )

        return converted_code

    def _inject_matplotlib_save_support(self, code: str) -> str:
        """Inject only matplotlib image save detection and save_chart()."""
        has_matplotlib_import = "import matplotlib" in code or "from matplotlib" in code
        if not has_matplotlib_import:
            logger.debug("matplotlib_save_injection_skipped", reason="no matplotlib import")
            return code

        images_dir_literal = repr(str(get_images_dir()))
        save_support_code = """# ===== Matplotlib 图片保存捕获（自动注入） =====
import os
from matplotlib.figure import Figure
from matplotlib.text import Text

_SUYUAN_CHINESE_FONT_PROP = None
_SUYUAN_FONT_WARNING_EMITTED = False
_SUYUAN_LAST_CONFIGURED_FONT = None

def _suyuan_configure_matplotlib_chinese_font(force_default=False):
    '''默认方正小标宋简；显式字体支持中文时优先显式字体。'''
    try:
        global _SUYUAN_CHINESE_FONT_PROP, _SUYUAN_FONT_WARNING_EMITTED, _SUYUAN_LAST_CONFIGURED_FONT
        import matplotlib.pyplot as _suyuan_plt
        from matplotlib import font_manager as _suyuan_font_manager
        from matplotlib.ft2font import FT2Font
        current_fonts = list(_suyuan_plt.rcParams.get('font.sans-serif', []))
        default_font_path = '/home/xckj/.local/share/fonts/方正小标宋简.TTF'
        if not os.path.exists(default_font_path):
            return
        _suyuan_font_manager.fontManager.addfont(default_font_path)
        default_font_prop = _suyuan_font_manager.FontProperties(fname=default_font_path)
        default_font_name = default_font_prop.get_name()

        def _font_supports_chinese(font_name):
            try:
                prop = _suyuan_font_manager.FontProperties(family=[font_name])
                path = _suyuan_font_manager.findfont(prop, fallback_to_default=False)
                charmap = FT2Font(path).get_charmap()
                return ord('中') in charmap, _suyuan_font_manager.FontProperties(fname=path)
            except Exception:
                return False, None

        selected_name = default_font_name
        selected_prop = default_font_prop
        requested_font = current_fonts[0] if current_fonts else None
        user_requested_font = (
            not force_default
            and requested_font
            and requested_font != default_font_name
            and requested_font != _SUYUAN_LAST_CONFIGURED_FONT
        )
        if user_requested_font:
            supports_chinese, font_prop = _font_supports_chinese(requested_font) if requested_font else (False, None)
            if supports_chinese:
                selected_name = font_prop.get_name()
                selected_prop = font_prop
            elif requested_font and not _SUYUAN_FONT_WARNING_EMITTED:
                print(f'字体提示：指定字体不支持中文，已回退为方正小标宋简。font={requested_font}')
                _SUYUAN_FONT_WARNING_EMITTED = True

        if not selected_name:
            return
        merged_fonts = [selected_name] + [f for f in current_fonts if f != selected_name]
        _suyuan_plt.rcParams['font.family'] = 'sans-serif'
        _suyuan_plt.rcParams['font.sans-serif'] = merged_fonts
        _suyuan_plt.rcParams['axes.unicode_minus'] = False
        _SUYUAN_CHINESE_FONT_PROP = selected_prop
        _SUYUAN_LAST_CONFIGURED_FONT = selected_name
    except Exception:
        pass

_suyuan_configure_matplotlib_chinese_font(force_default=True)

_SUYUAN_CHART_PATHS_EMITTED = set()
_suyuan_original_figure_savefig = getattr(Figure.savefig, '_suyuan_original_savefig', Figure.savefig)

def _suyuan_normalize_matplotlib_label_text(value):
    '''将常见污染物 Unicode 下标/上标转为 Matplotlib mathtext，避免字体缺字显示黑框。'''
    if not isinstance(value, str) or not value:
        return value
    replacements = [
        ('PM₂.₅', 'PM$_{2.5}$'),
        ('PM₂₅', 'PM$_{2.5}$'),
        ('PM₁₀', 'PM$_{10}$'),
        ('O₃', 'O$_3$'),
        ('NO₂', 'NO$_2$'),
        ('SO₂', 'SO$_2$'),
        ('CO₂', 'CO$_2$'),
        ('CH₄', 'CH$_4$'),
        ('N₂O', 'N$_2$O'),
        ('μg/m³', 'μg/m$^3$'),
        ('ug/m³', 'ug/m$^3$'),
        ('/m³', '/m$^3$'),
        ('m³', 'm$^3$'),
        ('km²', 'km$^2$'),
        ('m²', 'm$^2$'),
    ]
    normalized = value
    for old, new in replacements:
        normalized = normalized.replace(old, new)
    return normalized

def _suyuan_normalize_matplotlib_figure_text(fig):
    try:
        for text in fig.findobj(match=Text):
            original = text.get_text()
            normalized = _suyuan_normalize_matplotlib_label_text(original)
            if normalized != original:
                text.set_text(normalized)
    except Exception:
        pass

def _suyuan_text_contains_cjk(value):
    return isinstance(value, str) and any('\\u4e00' <= ch <= '\\u9fff' for ch in value)

def _suyuan_apply_chinese_font_to_figure(fig):
    try:
        if _SUYUAN_CHINESE_FONT_PROP is None:
            return
        for text in fig.findobj(match=Text):
            if not _suyuan_text_contains_cjk(text.get_text()):
                continue
            current_size = text.get_fontsize()
            text.set_fontproperties(_SUYUAN_CHINESE_FONT_PROP)
            text.set_fontsize(current_size)
    except Exception:
        pass

def _suyuan_emit_chart_saved(path):
    '''输出标准图片保存标记，供 execute_python 后处理登记统一资源。'''
    try:
        if path is None:
            return
        if isinstance(path, (str, os.PathLike)):
            chart_path = os.path.abspath(os.fspath(path))
        else:
            return
        if chart_path not in _SUYUAN_CHART_PATHS_EMITTED:
            _SUYUAN_CHART_PATHS_EMITTED.add(chart_path)
            print(f'CHART_SAVED:{chart_path}')
    except Exception:
        pass

def _suyuan_patched_figure_savefig(self, fname, *args, **kwargs):
    _suyuan_configure_matplotlib_chinese_font()
    _suyuan_normalize_matplotlib_figure_text(self)
    _suyuan_apply_chinese_font_to_figure(self)
    result = _suyuan_original_figure_savefig(self, fname, *args, **kwargs)
    _suyuan_emit_chart_saved(fname)
    return result

_suyuan_patched_figure_savefig._suyuan_original_savefig = _suyuan_original_figure_savefig
Figure.savefig = _suyuan_patched_figure_savefig

def save_chart(fig, filename, dpi=150, bbox_inches='tight', facecolor='white'):
    '''
    保存 matplotlib 图表并输出 CHART_SAVED 标记，便于工具登记统一资源。
    本函数不修改字体、字号、画布、布局或其他视觉设计。
    '''
    charts_dir = __SUYUAN_IMAGES_DIR__
    try:
        os.makedirs(charts_dir, exist_ok=True)
    except PermissionError:
        import tempfile
        charts_dir = tempfile.gettempdir()
        os.makedirs(charts_dir, exist_ok=True)

    filepath = os.path.join(charts_dir, filename)
    _suyuan_configure_matplotlib_chinese_font()
    _suyuan_normalize_matplotlib_figure_text(fig)
    _suyuan_apply_chinese_font_to_figure(fig)
    _suyuan_original_figure_savefig(
        fig,
        filepath,
        dpi=dpi,
        bbox_inches=bbox_inches,
        facecolor=facecolor,
    )
    _suyuan_emit_chart_saved(filepath)
    return filepath

""".replace("__SUYUAN_IMAGES_DIR__", images_dir_literal)

        injected_code = save_support_code + "\n" + code
        logger.info(
            "matplotlib_save_injection_completed",
            original_length=len(code),
            injected_length=len(injected_code),
        )
        return injected_code

    def _inject_excel_helpers(self, code: str) -> str:
        """
        条件性注入 Excel 辅助函数（仅检测到需要时才注入）

        策略：
        - 检测代码中是否涉及 Excel 操作
        - 检查 openpyxl 是否可用
        - 只在需要时注入辅助函数
        - 自动添加预览触发（即使LLM直接用openpyxl/pandas）
        """
        # 检测是否需要 Excel 操作
        excel_keywords = [
            'edit_excel_data',      # 用户调用辅助函数
            'openpyxl',             # 用户直接用 openpyxl
            '.xlsx',                # 操作 xlsx 文件
            '.xls',                 # 操作 xls 文件
            'to_excel',             # pandas to_excel
            'read_excel',           # pandas read_excel
        ]

        need_excel = any(keyword in code for keyword in excel_keywords)

        if not need_excel:
            logger.debug("excel_injection_skipped", reason="no_excel_keywords")
            return code

        # 检查 openpyxl 是否安装
        try:
            import openpyxl
            openpyxl_available = True
        except ImportError:
            openpyxl_available = False
            logger.warning("excel_injection_skipped", reason="openpyxl_not_installed")

        if not openpyxl_available:
            # openpyxl 未安装，不注入辅助函数
            # 用户代码会直接报错，错误信息更清晰
            return code

        logger.info("excel_injection_started", has_excel_keywords=need_excel)

        # 注入辅助函数
        helper_code = '''# ===== Excel 辅助函数（自动注入，保留图表和格式） =====
def edit_excel_data(file_path, updates, sheet_name=None):
    """
    修改 Excel 数据，保留图表和格式

    ⚠️ 重要：此函数使用 openpyxl 直接修改单元格，不会丢失图表和格式

    Args:
        file_path: Excel 文件路径（.xlsx 格式）
        updates: 更新数据，格式：
            - 单个单元格: {"A1": "新值"}
            - 多个单元格: {"A1": "值1", "B2": "值2", "C3": 123}
        sheet_name: 工作表名（可选，默认使用活动工作表）

    Returns:
        dict: {"success": True, "updated_count": N, "message": "...", "file_path": "..."}

    Example:
        # 修改单个单元格
        edit_excel_data("data.xlsx", {"A1": "北京"})

        # 批量修改
        edit_excel_data("data.xlsx", {
            "A2": "上海",
            "B2": 85,
            "C2": 45
        })
    """
    import openpyxl
    from pathlib import Path
    import os

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        count = 0
        for cell, value in updates.items():
            ws[cell] = value
            count += 1

        wb.save(file_path)
        wb.close()

        # 规范化文件路径
        file_path = os.path.abspath(file_path)

        # ✅ 打印特殊标记，触发前端预览生成
        print(f"EXCEL_SAVED:{file_path}")

        return {
            "success": True,
            "updated_count": count,
            "message": f"成功更新 {count} 个单元格，图表和格式已保留",
            "file_path": file_path
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": f"更新失败: {str(e)}"
        }


def read_excel_with_preview(file_path, sheet_name=None, head_rows=20):
    """
    读取 Excel 文件并生成前端预览

    ⚠️ 重要：读取文件也会生成前端预览，方便查看表格内容和格式

    Args:
        file_path: Excel 文件路径（.xlsx 格式）
        sheet_name: 工作表名（可选，默认使用活动工作表）
        head_rows: 读取前几行数据（默认20行）

    Returns:
        dict: {
            "success": True,
            "data": [...],  # 数据列表
            "columns": [...],  # 列名
            "total_rows": N,  # 总行数
            "total_columns": N,  # 总列数
            "file_path": "..."  # 文件路径（触发预览）
        }

    Example:
        # 读取并预览
        result = read_excel_with_preview("data.xlsx")
        print(f"总行数: {result['total_rows']}")
        print(f"数据: {result['data']}")
    """
    import openpyxl
    import pandas as pd
    import os

    try:
        # 规范化文件路径
        file_path = os.path.abspath(file_path)

        # 使用 pandas 读取数据
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=head_rows)

        # 使用 openpyxl 获取总行数和列数
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        total_rows = ws.max_row
        total_columns = ws.max_column
        wb.close()

        # ✅ 打印特殊标记，触发前端预览生成
        print(f"EXCEL_SAVED:{file_path}")

        return {
            "success": True,
            "data": df.to_dict("records"),
            "columns": df.columns.tolist(),
            "total_rows": total_rows,
            "total_columns": total_columns,
            "preview_rows": len(df),
            "file_path": file_path,
            "message": f"成功读取 Excel 文件，共 {total_rows} 行 x {total_columns} 列"
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": f"读取失败: {str(e)}"
        }


def merge_excel_with_charts(file_paths, output_path):
    """
    合并多个Excel文件到一个工作簿（保留图表、格式、数据）

    方法：
    1. 复制第一个文件作为基础
    2. 逐个复制其他文件的sheet（手动复制：内容+样式+图表）

    ⚠️ 重要：此方法会保留所有图表、格式和数据

    Args:
        file_paths: Excel文件路径列表
        output_path: 输出文件路径

    Returns:
        dict: {"success": True, "merged_count": N, "file_path": "...", "message": "..."}

    Example:
        files = ['/tmp/file1.xlsx', '/tmp/file2.xlsx']
        result = merge_excel_with_charts(files, '/tmp/merged.xlsx')
        # 输出文件包含所有图表和格式
    """
    import openpyxl
    from openpyxl import load_workbook
    import shutil
    import os

    if not file_paths:
        return {
            "success": False,
            "error": "文件列表为空",
            "message": "请提供至少一个Excel文件"
        }

    try:
        # 复制第一个文件作为基础（保留原文件的图表）
        shutil.copy(file_paths[0], output_path)

        # 加载输出文件
        wb_output = load_workbook(output_path)

        merged_count = 0

        # 从第二个文件开始合并
        for file_path in file_paths[1:]:
            if not os.path.exists(file_path):
                continue

            # 加载源文件
            wb_source = load_workbook(file_path)

            for sheet_name in wb_source.sheetnames:
                ws_source = wb_source[sheet_name]

                # 处理sheet名称冲突
                new_sheet_name = sheet_name
                counter = 1
                while new_sheet_name in wb_output.sheetnames:
                    new_sheet_name = f"{sheet_name}_{counter}"
                    counter += 1

                # 创建新sheet
                ws_new = wb_output.create_sheet(title=new_sheet_name)

                # 手动复制单元格内容和样式
                for row in ws_source.iter_rows():
                    for cell in row:
                        new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)

                        # 复制样式
                        if cell.has_style:
                            new_cell.font = cell.font.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.number_format = cell.number_format
                            new_cell.protection = cell.protection.copy()
                            new_cell.alignment = cell.alignment.copy()

                # 手动复制图表（关键步骤！）
                for chart in ws_source._charts:
                    try:
                        anchor = chart.anchor
                        ws_new.add_chart(chart, anchor)
                    except Exception as e:
                        # 图表复制失败时继续，不影响其他内容
                        pass

                merged_count += 1

            wb_source.close()

        # 保存输出文件
        output_path = os.path.abspath(output_path)
        wb_output.save(output_path)
        wb_output.close()

        # 打印预览触发
        print(f"EXCEL_SAVED:{output_path}")

        return {
            "success": True,
            "merged_count": merged_count + 1,  # 第一个文件也算
            "file_path": output_path,
            "message": f"成功合并 {len(file_paths)} 个文件（图表和格式已保留）"
        }

    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": f"合并失败: {str(e)}"
        }

# ===== Excel 辅助函数注入完成 =====

'''

        injected_code = helper_code + code

        # ✅ 新增：检测LLM是否直接使用openpyxl/pandas读取Excel，自动添加预览触发
        import re
        auto_preview_code = self._auto_add_excel_preview_trigger(code)
        if auto_preview_code:
            injected_code = helper_code + auto_preview_code
            logger.info(
                "excel_auto_preview_added",
                original_length=len(code),
                injected_length=len(injected_code),
                auto_preview_length=len(auto_preview_code) - len(code)
            )

        logger.info(
            "excel_injection_completed",
            original_length=len(code),
            injected_length=len(injected_code),
            injection_type="excel_helpers"
        )

        return injected_code

    def _auto_add_excel_preview_trigger(self, code: str) -> str:
        """
        自动检测Excel文件路径并添加预览触发

        检测模式：
        - file_path = 'xxx.xlsx'
        - file_path = "xxx.xlsx"
        - load_workbook('xxx.xlsx')
        - read_excel('xxx.xlsx')
        """
        import re
        import os

        # 正则模式：提取Excel文件路径
        patterns = [
            # file_path = 'xxx.xlsx' 或 file_path = "xxx.xlsx"
            r"file_path\s*=\s*['\"]([^'\"]+\.(xlsx|xls))['\"]",
            # load_workbook('xxx.xlsx') 或 load_workbook("xxx.xlsx")
            r"load_workbook\(['\"]([^'\"]+\.(xlsx|xls))['\"]",
            # read_excel('xxx.xlsx') 或 read_excel("xxx.xlsx")
            r"read_excel\(['\"]([^'\"]+\.(xlsx|xls))['\"]",
        ]

        extracted_paths = set()
        for pattern in patterns:
            matches = re.findall(pattern, code)
            for match in matches:
                extracted_paths.add(match[0])

        if not extracted_paths:
            logger.debug("auto_preview_skipped", reason="no_excel_file_path_found")
            return code

        # 在代码末尾添加预览触发
        preview_trigger_lines = [
            "\n# ===== 自动添加：Excel预览触发 =====",
            "import os",
        ]

        for file_path in extracted_paths:
            # 规范化路径
            preview_trigger_lines.append(f"print('EXCEL_SAVED:' + os.path.abspath('{file_path}'))")

        preview_trigger_lines.append("# ===== 预览触发完成 =====\n")

        modified_code = code + "\n" + "\n".join(preview_trigger_lines)

        logger.info(
            "auto_preview_trigger_added",
            excel_files=list(extracted_paths),
            lines_added=len(preview_trigger_lines)
        )

        return modified_code

    def _extract_file_paths_from_output(self, output: str) -> List[str]:
        """
        从 Python 代码输出中提取文件路径

        检测常见的文件保存输出格式：
        - "报告已生成：/path/to/file.docx"
        - "文件已保存：/path/to/file.xlsx"
        - "File saved: /path/to/file.pdf"
        - "/path/to/file.docx"

        Args:
            output: Python 代码输出

        Returns:
            文件路径列表
        """
        import re
        file_paths = []

        if not output:
            return file_paths

        # 按行扫描且限制候选行长度。旧实现在整段 stdout 上使用
        # `(.+?\.ext)...` 宽泛匹配，遇到不含文件后缀的大型 ECharts JSON
        # 会退化为 O(n²)，并占满 Web worker 直到 Uvicorn 健康检查将其终止。
        extensions = r'(?:docx|xlsx|pptx|pdf|doc|xls|ppt)'
        marker_patterns = [
            re.compile(
                rf'(?:WORD_SAVED|WORD_REPORT_SAVED|DOCX_SAVED|PPT_SAVED|PPTX_SAVED|PDF_SAVED|EXCEL_SAVED)'
                rf'[:：]\s*(.+?\.{extensions})(?:\s|$)',
                re.IGNORECASE,
            ),
            re.compile(
                rf'(?:报告已生成|文件已保存|已生成|保存成功|File saved|saved)'
                rf'[:：]\s*(.+?\.{extensions})(?:\s|$)',
                re.IGNORECASE,
            ),
        ]
        standalone_path = re.compile(
            rf'^\s*(.+?\.{extensions})\s*(?:已保存|生成完成)?\s*$',
            re.IGNORECASE,
        )

        extension_hint = re.compile(rf'\.{extensions}(?:\s|$)', re.IGNORECASE)
        for raw_line in output.splitlines():
            # 生成文件的日志行不应是巨型 payload；跳过长 JSON 也避免
            # 不受信 stdout 对主进程造成过量 CPU 消耗。
            if len(raw_line) > 8192 or not extension_hint.search(raw_line):
                continue
            line = raw_line.strip()
            for pattern in marker_patterns:
                match = pattern.search(line)
                if match:
                    file_paths.append(match.group(1))
                    break
            else:
                match = standalone_path.fullmatch(line)
                if match:
                    file_paths.append(match.group(1))

        # 去重并验证文件存在
        unique_paths = []
        seen = set()
        for path in file_paths:
            # 规范化路径
            path = os.path.abspath(path) if not os.path.isabs(path) else path
            if path not in seen and os.path.exists(path):
                unique_paths.append(path)
                seen.add(path)

        return unique_paths

    def _format_python_error(self, error: Exception, code: str) -> Dict[str, str]:
        """
        格式化 Python 错误信息，提供详细的错误上下文和修复建议

        Args:
            error: Python 异常对象
            code: 执行的代码

        Returns:
            包含错误详情的字典
        """
        import traceback
        import re

        error_type = type(error).__name__
        error_msg = str(error)
        error_lines = traceback.format_exception(type(error), error, error.__traceback__)

        # 提取行号
        line_number = None
        for line in error_lines:
            match = re.search(r'File "<ipython-input-\d+>", line (\d+)', line)
            if match:
                line_number = int(match.group(1))
                break

        # 获取错误行的代码上下文
        code_context = ""
        if line_number:
            code_lines = code.split('\n')
            if 0 < line_number <= len(code_lines):
                error_line = code_lines[line_number - 1].strip()
                code_context = f"错误行代码（第{line_number}行）: {error_line}"

        # 常见错误的修复建议
        suggestions = self._get_error_suggestions(error_type, error_msg, code_context)

        # 构建详细的错误信息
        detailed_error = f"❌ Python 执行错误\n\n"
        detailed_error += f"**错误类型**: {error_type}\n"
        detailed_error += f"**错误信息**: {error_msg}\n"
        if line_number:
            detailed_error += f"**错误行号**: {line_number}\n"
        if code_context:
            detailed_error += f"{code_context}\n"
        if suggestions:
            detailed_error += f"\n**💡 修复建议**:\n{suggestions}\n"

        logger.warning(
            "python_execution_error",
            error_type=error_type,
            error_msg=error_msg,
            line_number=line_number,
            has_suggestions=bool(suggestions)
        )

        return {
            "error_type": error_type,
            "error_message": detailed_error,
            "summary": f"❌ 执行失败: {error_type} - {error_msg}",
            "line_number": line_number,
            "code_context": code_context,
            "suggestions": suggestions
        }

    def _get_error_suggestions(self, error_type: str, error_msg: str, code_context: str) -> str:
        """
        根据错误类型提供修复建议

        Args:
            error_type: 错误类型
            error_msg: 错误消息
            code_context: 错误行的代码上下文

        Returns:
            修复建议字符串
        """
        suggestions = []

        if error_type == "TypeError":
            if "can only concatenate str" in error_msg:
                suggestions.append("• **类型不匹配**: 尝试将字符串和数字相加")
                suggestions.append("• **解决方案**: 使用 `float()` 或 `int()` 转换变量类型")
                if "wind_dir" in code_context or "wind_direction" in code_context:
                    suggestions.append("• **示例修复**: `float(wind_dir) + 11.25` 而不是 `wind_dir + 11.25`")
                suggestions.append("• **JSON 数据注意**: 从 JSON 读取的数字可能是字符串类型")

            elif "unsupported operand type" in error_msg:
                suggestions.append("• **运算符不支持**: 操作数类型不匹配")
                suggestions.append("• **解决方案**: 检查变量类型，使用 `type()` 查看类型，使用 `int()`/`float()`/`str()` 转换")

            elif "not subscriptable" in error_msg:
                suggestions.append("• **不可下标访问**: 尝试对非列表/字典类型使用索引")
                suggestions.append("• **解决方案**: 检查变量是否为列表或字典，使用 `list()` 或 `dict()` 转换")

        elif error_type == "KeyError":
            suggestions.append("• **键不存在**: 字典中没有指定的键")
            suggestions.append("• **解决方案1**: 使用 `.get(key, default)` 方法提供默认值")
            suggestions.append("• **解决方案2**: 检查键名是否正确（区分大小写）")
            suggestions.append("• **示例修复**: `value = data.get('PM2_5', 0)` 而不是 `value = data['PM2_5']`")

        elif error_type == "NameError":
            suggestions.append("• **变量未定义**: 使用了未声明的变量")
            suggestions.append("• **解决方案**: 检查变量名拼写，确保变量已定义")

        elif error_type == "ValueError":
            if "could not convert string to float" in error_msg:
                suggestions.append("• **字符串转换失败**: 无法将字符串转换为数字")
                suggestions.append("• **解决方案**: 检查数据是否包含非数字字符，使用 `try-except` 处理异常")
                suggestions.append("• **示例修复**: `try: val = float(s) except: val = 0`")

            elif "I/O operation on closed file" in error_msg:
                suggestions.append("• **文件已关闭**: 尝试操作已关闭的文件对象")
                suggestions.append("• **解决方案**: 确保文件在 `with` 块内操作，或重新打开文件")

        elif error_type == "AttributeError":
            if "'NoneType' object has no attribute" in error_msg:
                suggestions.append("• **空对象属性**: 尝试访问 None 对象的属性")
                suggestions.append("• **解决方案**: 检查对象是否为 None，添加空值检查")
                suggestions.append("• **示例修复**: `if obj is not None: obj.method()`")

            else:
                suggestions.append("• **属性不存在**: 对象没有该属性或方法")
                suggestions.append("• **解决方案**: 检查对象类型，使用 `dir()` 查看可用属性")

        elif error_type == "IndexError":
            suggestions.append("• **索引越界**: 列表索引超出范围")
            suggestions.append("• **解决方案**: 检查列表长度，使用 `len()` 确保索引有效")
            suggestions.append("• **示例修复**: `if i < len(lst): value = lst[i]`")

        elif error_type == "FileNotFoundError":
            suggestions.append("• **文件不存在**: 找不到指定的文件")
            suggestions.append("• **解决方案**: 检查文件路径是否正确，使用绝对路径或相对于当前目录的路径")

        elif error_type == "ZeroDivisionError":
            suggestions.append("• **除零错误**: 尝试除以零")
            suggestions.append("• **解决方案**: 检查除数是否为零，添加条件判断")
            suggestions.append("• **示例修复**: `if divisor != 0: result = a / divisor`")

        elif error_type == "SyntaxError":
            suggestions.append("• **语法错误**: 代码语法不正确")
            suggestions.append("• **常见原因**: 括号不匹配、冒号缺失、缩进错误")
            suggestions.append("• **解决方案**: 检查括号、引号是否配对，检查缩进是否正确")

        # 如果没有特定建议，提供通用建议
        if not suggestions:
            suggestions.append("• **检查代码**: 仔细阅读错误信息，定位问题代码")
            suggestions.append("• **打印调试**: 使用 `print()` 输出变量值和类型")
            suggestions.append("• **异常处理**: 使用 `try-except` 捕获异常")

        return "\n".join(suggestions)

    def _parse_subprocess_error(self, stderr: str, code: str) -> Dict[str, str]:
        """
        解析 subprocess 模式的错误信息

        Args:
            stderr: 标准错误输出
            code: 执行的代码

        Returns:
            包含错误详情的字典
        """
        import re

        # 尝试提取错误类型和错误消息
        error_type = "UnknownError"
        error_msg = stderr.strip() if stderr else "未知错误"

        # 常见 Python 错误模式
        patterns = [
            r"(NameError|TypeError|ValueError|KeyError|AttributeError|IndexError|FileNotFoundError|ZeroDivisionError|SyntaxError): (.+)",
            r"Traceback \(most recent call last\):\s+.*\s+(NameError|TypeError|ValueError|KeyError|AttributeError|IndexError|FileNotFoundError|ZeroDivisionError|SyntaxError): (.+)",
        ]

        for pattern in patterns:
            match = re.search(pattern, stderr)
            if match:
                error_type = match.group(1)
                error_msg = match.group(2).strip()
                break

        # 提取行号
        line_number = None
        line_match = re.search(r'File "<string>", line (\d+)', stderr)
        if line_match:
            line_number = int(line_match.group(1))

        # 获取错误行的代码上下文
        code_context = ""
        if line_number:
            code_lines = code.split('\n')
            if 0 < line_number <= len(code_lines):
                error_line = code_lines[line_number - 1].strip()
                code_context = f"错误行代码（第{line_number}行）: {error_line}"

        # 获取修复建议
        suggestions = self._get_error_suggestions(error_type, error_msg, code_context)

        # 构建详细的错误信息
        detailed_error = f"❌ Python 执行错误\n\n"
        detailed_error += f"**错误类型**: {error_type}\n"
        detailed_error += f"**错误信息**: {error_msg}\n"
        if line_number:
            detailed_error += f"**错误行号**: {line_number}\n"
        if code_context:
            detailed_error += f"{code_context}\n"
        if suggestions:
            detailed_error += f"\n**💡 修复建议**:\n{suggestions}\n"

        return {
            "error_type": error_type,
            "error_message": detailed_error,
            "summary": f"❌ 执行失败: {error_type} - {error_msg}",
            "line_number": line_number,
            "code_context": code_context,
            "suggestions": suggestions
        }

    def get_function_schema(self) -> Dict[str, Any]:
        """获取 Function Calling Schema"""
        return {
            "name": "execute_python",
            "description": (
                "通用 Python 代码执行工具，不限制于数据分析、Excel或可视化。"
                "适合需要复杂逻辑、结构化数据处理、数值计算、调用 Python 库、文件读写或文件生成的任务。"
                "代码在无网络 Bubblewrap 沙箱中执行；外部数据必须先通过查询或浏览工具获取。"
                "对于会话数据文件，承担全量扫描、复杂聚合、分组排名、关联、窗口计算、清洗和派生字段；"
                "通过 load_data(file_path) 加载，较大计算结果通过 save_data(...) 保存为新的 file_path，避免直接输出大量明细。"
                "如果任务只是查看文件、搜索文本、检查进程或调用现成 CLI，优先使用 bash；"
                "需要循环、条件分支、解析转换、程序化处理或可靠地产出文件时，优先使用 execute_python。"
                "复杂用法先阅读 backend/app/tools/utility/execute_python_manual.md。"
                "每次调用是独立环境；用 load_data(file_path) 读取会话数据文件，"
                "跨调用复用请用 save_data(...) 保存并获取 file_path。"
                "正式报告静态图表优先使用 create_report_chart；流程/架构图使用 call_sub_agent(target_mode='board') 调用画板Agent。"
                "生成文件会自动归档并发布到会话资源目录；必须复用返回的 file_path，"
                "不要自行拼接 sessions/... 路径，也不要再调用 publish_session_file；默认超时30秒。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": (
                            "要执行的 Python 代码。matplotlib 图片可用 save_chart(fig, filename) 或 fig.savefig(path) 保存；"
                            "matplotlib 中文字体由系统自动设置，不要显式设置 SimHei、DejaVu Sans 等不支持中文的字体；"
                            "工具只捕获保存路径，不接管图表字号、画布或布局。"
                        )
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "超时时间（秒），默认30"
                    }
                },
                "required": ["code"]
            }
        }


class ExecuteEChartsPythonTool(ExecutePythonTool):
    """ECharts-only Python execution wrapper with a strict visuals contract."""

    def __init__(self):
        super().__init__()
        self.name = "execute_echarts_python"
        self.category = ToolCategory.VISUALIZATION
        self.enable_echarts_visuals = True
        self.description = (
            "执行 Python 代码并将 stdout 中的一行一个纯 JSON ECharts option 转换为前端交互式图表。"
            "用于生成前端交互式 ECharts 图表（柱状图/折线图/散点图/饼图/3D图/地图等）。"
            "⚠️ **图表选择策略**："
            "① 正式报告Word/QMD静态图表 → 优先使用 create_report_chart；"
            "② 前端交互式图表/复杂数据可视化 → 使用 execute_echarts_python；"
            "③ 复杂Python绘图（3D/科研图/多子图） → 使用 execute_python + matplotlib/seaborn/plotly。"
            "通用计算和文件生成仍使用 execute_python。"
        )

    async def execute(
        self,
        context=None,
        code: str = None,
        timeout: Optional[int] = None,
        expected_charts: Optional[int] = None,
        **kwargs
    ) -> Dict[str, Any]:
        result = await super().execute(context=context, code=code, timeout=timeout, **kwargs)
        from app.tools.resource_declarations import resources_for_visuals

        echarts_visuals = [
            visual for visual in result.get("visuals", [])
            if visual.get("meta", {}).get("schema_version") == "echarts_standard"
        ]

        for visual in echarts_visuals:
            visual.setdefault("meta", {})
            visual["meta"]["generator"] = "execute_echarts_python"

        if not result.get("success"):
            result["visuals"] = echarts_visuals
            return result

        if not echarts_visuals:
            result["status"] = "failed"
            result["success"] = False
            result["visuals"] = []
            result["summary"] = (
                "❌ 未解析到有效的 ECharts option。"
                "请确保 Python stdout 每行只输出一个纯 JSON 对象，且顶层包含 series 数组。"
            )
            result.setdefault("metadata", {})
            result["metadata"]["tool_name"] = "execute_echarts_python"
            result["metadata"]["error_type"] = "NO_ECHARTS_OPTIONS"
            return result

        if expected_charts is not None and len(echarts_visuals) != expected_charts:
            result["status"] = "failed"
            result["success"] = False
            result["visuals"] = echarts_visuals
            result["summary"] = (
                f"❌ ECharts 图表数量不匹配：期望 {expected_charts} 个，"
                f"实际解析到 {len(echarts_visuals)} 个。"
            )
            result.setdefault("metadata", {})
            result["metadata"]["tool_name"] = "execute_echarts_python"
            result["metadata"]["error_type"] = "ECHARTS_COUNT_MISMATCH"
            result["metadata"]["expected_charts"] = expected_charts
            result["metadata"]["actual_charts"] = len(echarts_visuals)
            return result

        result["visuals"] = echarts_visuals
        result.setdefault("metadata", {})
        result["metadata"]["tool_name"] = "execute_echarts_python"
        result["metadata"]["visuals_count"] = len(echarts_visuals)
        result.setdefault("resources", []).extend(
            resources_for_visuals(echarts_visuals, tool_name=self.name)
        )
        result["summary"] = f"✅ ECharts 图表生成完成：{len(echarts_visuals)} 个"
        return result

    def get_function_schema(self) -> Dict[str, Any]:
        """获取 ECharts 专用 Function Calling Schema"""
        return {
            "name": "execute_echarts_python",
            "description": (
                "执行 Python 代码生成 ECharts 图表配置，并返回标准 visuals 给前端渲染。"
                "首次使用前必须先调用 read_file 阅读 "
                "backend/app/tools/utility/execute_echarts_python_manual.md。"
                "使用工具返回的 file_path，代码中通过系统注入的 load_data(file_path) 获取数据。"
                "仅用于图表模式的 ECharts 输出：Python 必须使用 print(json.dumps(option, ensure_ascii=False))，"
                "每行输出一个完整、纯 JSON 的 ECharts option，顶层必须包含 series 数组。"
                "图表通过统一会话资源目录发布和预览，不返回独立图片 URL。"
                "多图时输出多行纯 JSON。禁止输出 CHART_1: 前缀、Markdown 代码块、解释文字包裹 JSON。"
                "数据分析、清洗、中间计算和文件生成请使用 execute_python。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": (
                            "要执行的 Python 代码。使用 load_data(file_path) 读取会话数据文件。"
                            "必须在 stdout 中逐行 print 纯 JSON ECharts option；"
                            "不要打印 CHART_1:、Markdown、自然语言说明或本地路径作为图表协议。"
                        )
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "超时时间（秒），默认30"
                    },
                    "expected_charts": {
                        "type": "integer",
                        "description": "可选。期望生成的 ECharts 图表数量；不匹配时工具返回失败。"
                    }
                },
                "required": ["code"]
            }
        }
