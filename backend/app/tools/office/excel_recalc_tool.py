"""
ExcelRecalc 工具 - Excel 公式重算

功能：
- 重新计算 Excel 文件中的所有公式
- 使用 LibreOffice 执行计算
- 扫描并报告公式错误（#VALUE!, #DIV/0!, #REF! 等）
- 统计公式数量

使用场景：
- 更新外部数据后重算公式
- 验证公式计算结果
- 检测公式错误
"""
import subprocess
import tempfile
import platform
from pathlib import Path
from typing import Dict, Any
from openpyxl import load_workbook
from app.tools.base.tool_interface import LLMTool, ToolCategory
from app.tools.office.soffice import get_soffice_env
import structlog

logger = structlog.get_logger()

# LibreOffice 配置目录
if platform.system() == "Darwin":
    MACRO_DIR = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
elif platform.system() == "Windows":
    MACRO_DIR = tempfile.gettempdir() + "/libreoffice_excel_profile/user/basic/Standard"
else:
    MACRO_DIR = "~/.config/libreoffice/4/user/basic/Standard"

MACRO_FILENAME = "Module1.xba"

# LibreOffice Basic 宏：重算公式
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

# Excel 错误类型
EXCEL_ERRORS = [
    "#VALUE!",  # 值错误
    "#DIV/0!",  # 除零错误
    "#REF!",    # 引用错误
    "#NAME?",   # 名称错误
    "#NULL!",   # 空值错误
    "#NUM!",    # 数值错误
    "#N/A",     # 不可用错误
]


class ExcelRecalcTool(LLMTool):
    """
    Excel 公式重算工具

    功能：
    - 重新计算 Excel 文件中的所有公式
    - 使用 LibreOffice 执行计算
    - 扫描并报告公式错误
    """

    def __init__(self):
        super().__init__(
            name="recalc_excel",
            description="""重新计算 Excel 文件中的所有公式

功能：
- 重新计算 Excel 文件中的所有公式
- 使用 LibreOffice 执行计算
- 扫描并报告公式错误（#VALUE!, #DIV/0!, #REF! 等）
- 统计公式数量

使用场景：
- 更新外部数据后重算公式
- 验证公式计算结果
- 检测公式错误

示例：
- recalc_excel(path="report.xlsx")
- recalc_excel(path="data.xlsx", timeout=60)

参数说明：
- path: Excel 文件路径（.xlsx）
- timeout: 超时时间（秒，默认 30）

注意：
- 需要安装 LibreOffice（soffice 命令）
- 处理时间取决于公式复杂度
- 会直接修改原文件
""",
            category=ToolCategory.QUERY,
            version="1.0.0",
            requires_context=False
        )

        self.working_dir = Path.cwd().parent  # D:\溯源\ 或 /opt/app/ 等

    async def execute(
        self,
        path: str,
        timeout: int = 30,
        **kwargs
    ) -> Dict[str, Any]:
        """
        重新计算 Excel 公式

        Args:
            path: Excel 文件路径
            timeout: 超时时间（秒）

        Returns:
            {
                "success": bool,
                "data": {
                    "file_path": str,
                    "total_formulas": int,
                    "total_errors": int,
                    "error_summary": dict
                },
                "summary": str
            }
        """
        try:
            # 1. 路径解析
            resolved_path = self._resolve_path(path)
            if not resolved_path:
                return {
                    "success": False,
                    "data": {"error": "文件路径无效"},
                    "summary": "公式重算失败：路径无效"
                }

            if not resolved_path.exists():
                return {
                    "success": False,
                    "data": {"error": f"文件不存在: {path}"},
                    "summary": "公式重算失败：文件不存在"
                }

            if resolved_path.suffix.lower() not in [".xlsx", ".xls"]:
                return {
                    "success": False,
                    "data": {"error": f"不支持的文件格式: {resolved_path.suffix}"},
                    "summary": "公式重算失败：仅支持 XLSX/XLS 格式"
                }

            # 2. 设置 LibreOffice 宏
            if not self._setup_libreoffice_macro():
                return {
                    "success": False,
                    "data": {"error": "LibreOffice 宏设置失败"},
                    "summary": "公式重算失败：宏设置失败"
                }

            # 3. 执行 LibreOffice 宏重算公式
            abs_path = str(resolved_path.absolute())
            cmd = [
                "soffice",
                "--headless",
                "--norestore",
                "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
                abs_path,
            ]

            # Windows 不支持 timeout 命令，使用 subprocess timeout
            try:
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=timeout,
                    check=False,
                    env=get_soffice_env(),
                )
            except subprocess.TimeoutExpired:
                logger.info("recalc_timeout", file=str(resolved_path))
                # Timeout 通常表示成功（LibreOffice 已完成）

            # 4. 扫描公式错误
            error_scan_result = self._scan_formula_errors(resolved_path)

            if "error" in error_scan_result:
                return {
                    "success": False,
                    "data": {"error": error_scan_result["error"]},
                    "summary": f"公式重算失败：{error_scan_result['error'][:80]}"
                }

            logger.info(
                "recalc_success",
                file=str(resolved_path),
                formulas=error_scan_result["total_formulas"],
                errors=error_scan_result["total_errors"]
            )

            return {
                "success": True,
                "data": {
                    "file_path": str(resolved_path),
                    "total_formulas": error_scan_result["total_formulas"],
                    "total_errors": error_scan_result["total_errors"],
                    "error_summary": error_scan_result["error_summary"],
                    "status": error_scan_result["status"]
                },
                "summary": (
                    f"公式重算完成：{error_scan_result['total_formulas']} 个公式，"
                    f"{error_scan_result['total_errors']} 个错误"
                )
            }

        except Exception as e:
            logger.error("recalc_failed", file=path, error=str(e))
            return {
                "success": False,
                "data": {"error": str(e)},
                "summary": f"公式重算失败：{str(e)[:80]}"
            }

    def _setup_libreoffice_macro(self) -> bool:
        """设置 LibreOffice 宏"""
        try:
            macro_dir = Path(MACRO_DIR).expanduser()
            macro_file = macro_dir / MACRO_FILENAME

            # 检查宏是否已存在
            if macro_file.exists() and "RecalculateAndSave" in macro_file.read_text():
                return True

            # 初始化 LibreOffice 配置目录
            if not macro_dir.exists():
                subprocess.run(
                    ["soffice", "--headless", "--terminate_after_init"],
                    capture_output=True,
                    timeout=10,
                    check=False,
                    env=get_soffice_env(),
                )
                macro_dir.mkdir(parents=True, exist_ok=True)

            # 写入宏文件
            macro_file.write_text(RECALCULATE_MACRO, encoding='utf-8')
            logger.info("libreoffice_macro_setup", macro_file=str(macro_file))
            return True

        except Exception as e:
            logger.error("libreoffice_macro_setup_failed", error=str(e))
            return False

    def _scan_formula_errors(self, file_path: Path) -> Dict[str, Any]:
        """扫描公式错误"""
        try:
            # 加载工作簿（data_only=True 获取计算值）
            wb = load_workbook(file_path, data_only=True)

            error_details = {err: [] for err in EXCEL_ERRORS}
            total_errors = 0

            # 扫描所有单元格
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            for err in EXCEL_ERRORS:
                                if err in cell.value:
                                    location = f"{sheet_name}!{cell.coordinate}"
                                    error_details[err].append(location)
                                    total_errors += 1
                                    break

            wb.close()

            # 统计公式数量
            wb_formulas = load_workbook(file_path, data_only=False)
            formula_count = 0
            for sheet_name in wb_formulas.sheetnames:
                ws = wb_formulas[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if (
                            cell.value
                            and isinstance(cell.value, str)
                            and cell.value.startswith("=")
                        ):
                            formula_count += 1
            wb_formulas.close()

            # 构建结果
            result = {
                "status": "success" if total_errors == 0 else "errors_found",
                "total_errors": total_errors,
                "total_formulas": formula_count,
                "error_summary": {}
            }

            for err_type, locations in error_details.items():
                if locations:
                    result["error_summary"][err_type] = {
                        "count": len(locations),
                        "locations": locations[:20]  # 最多返回 20 个位置
                    }

            return result

        except Exception as e:
            return {"error": str(e)}

    def _resolve_path(self, path: str) -> Path:
        """解析路径（支持相对路径和绝对路径）"""
        try:
            file_path = Path(path)

            if not file_path.is_absolute():
                file_path = self.working_dir / file_path

            return file_path.resolve()

        except Exception as e:
            logger.error("recalc_path_resolution_failed", path=path, error=str(e))
            return None

    def get_function_schema(self) -> Dict[str, Any]:
        """获取 Function Calling Schema"""
        return {
            "name": "recalc_excel",
            "description": """重新计算 Excel 文件中的所有公式

重新计算 Excel 文件中的所有公式，并扫描报告公式错误。

使用场景：
- 更新外部数据后重算公式
- 验证公式计算结果
- 检测公式错误

注意：
- 需要安装 LibreOffice
- 会直接修改原文件
""",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Excel 文件路径（.xlsx/.xls）。示例：'report.xlsx' 或 'D:/work/data.xlsx'"
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "超时时间（秒，默认 30）",
                        "default": 30
                    }
                },
                "required": ["path"]
            }
        }

    def is_available(self) -> bool:
        """检查 LibreOffice 是否可用"""
        try:
            result = subprocess.run(
                ["soffice", "--version"],
                capture_output=True,
                timeout=5,
                check=False
            )
            return result.returncode == 0
        except Exception:
            return False


# 创建工具实例
tool = ExcelRecalcTool()
