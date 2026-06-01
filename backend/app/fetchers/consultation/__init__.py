# -*- coding: utf-8 -*-
"""
会商文件批量更新 Fetcher

每天早上7点自动生成"当月累积（截至昨日）"的会商Excel文件

功能：
- 每月自动创建子文件夹（如 /tmp/会商文件/2026年1月/）
- 每天早上7点更新数据（覆盖历史文件）
- 使用用户提供的Excel模板，脚本仅填充原始数据，保留模板图表和公式
- 数据范围：本月1号 → 昨天
- 自动数据验证

调度周期：每天早上7点 (Cron: 0 7 * * *)
数据来源：全国/全省空气质量API
输出目录：/tmp/会商文件/{年月}/
模板目录：/tmp/会商文件/模板/

author: Claude
date: 2026-05-08
"""

import shutil
import subprocess
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
import structlog

from app.fetchers.base.fetcher_interface import DataFetcher

logger = structlog.get_logger()


# 污染物配置
POLLUTANTS_CONFIG = {
    "PM2.5": {"unit": "μg/m³", "normal_range": (5, 150)},
    "PM10": {"unit": "μg/m³", "normal_range": (10, 300)},
    "NO2": {"unit": "μg/m³", "normal_range": (5, 100)},
    "O3": {"unit": "μg/m³", "normal_range": (10, 160)},
    "AQI": {"unit": "", "normal_range": (80, 100), "is_rate": True}
}

GUANGDONG_CITIES = [
    "广州", "深圳", "珠海", "佛山", "惠州", "东莞", "中山", "江门", "肇庆",
    "汕头", "汕尾", "潮州", "揭阳",
    "湛江", "茂名", "阳江",
    "韶关", "河源", "梅州", "清远", "云浮"
]

CITY_STANDARD_FIELD_ALIASES = {
    "PM2.5": ["pM2_5_Decimal", "PM2_5_Decimal", "pM2_5", "PM2_5"],
    "PM10": ["pM10_Decimal", "PM10_Decimal", "pM10", "PM10"],
    "NO2": ["nO2", "NO2"],
    "O3": ["o3_8h", "O3_8h"],
    "AQI": ["fineRate", "FineRate", "AQIStandardRate"],
    "SO2": ["sO2", "SO2"],
    "CO": ["co", "cO", "CO"],
}


# Sheet 填充配置：定义每个sheet的数据区域、列映射和表头更新规则
SHEET_CONFIG = {
    # ========== 全国污染物 sheet ==========
    "全国PM2.5": {
        "scope": "national",
        "pollutant": "PM2.5",
        "data_rows": (2, 32),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # N列计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}{month:02d}",
            "D1": "{last_year}{month:02d}",
        }
    },
    "全国PM10": {
        "scope": "national",
        "pollutant": "PM10",
        "data_rows": (2, 32),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # N列计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}{month:02d}",
            "D1": "{last_year}{month:02d}",
        }
    },
    "全国NO2": {
        "scope": "national",
        "pollutant": "NO2",
        "data_rows": (2, 32),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # N列计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}{month:02d}",
            "D1": "{last_year}{month:02d}",
        }
    },
    "全国O3": {
        "scope": "national",
        "pollutant": "O3",
        "data_rows": (2, 32),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # N列计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}{month:02d}",
            "D1": "{last_year}{month:02d}",
        }
    },
    "全国AQI": {
        "scope": "national",
        "pollutant": "AQI",
        "data_rows": (2, 32),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "I",
                "target_value_col": "K",
                "extra_targets": [
                    {"col": "J", "data_source": "last_year"},
                    {"col": "L", "data_source": "diff_pct"},
                ],
                "sort_ascending": False,  # AQI达标率越高越好，降序
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "过渡期达标率",
            "J1": "{last_year}年过渡期达标率",
            "K1": "{year}年过渡期达标率",
        }
    },
    # ========== 全省污染物 sheet ==========
    "全省PM2.5": {
        "scope": "provincial",
        "pollutant": "PM2.5",
        "data_rows": (2, 22),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # 计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}浓度",
        }
    },
    "全省PM10": {
        "scope": "provincial",
        "pollutant": "PM10",
        "data_rows": (2, 22),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # N列计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}年{month}月达标率",
        }
    },
    "全省NO2": {
        "scope": "provincial",
        "pollutant": "NO2",
        "data_rows": (2, 22),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # 计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}浓度",
        }
    },
    "全省O3": {
        "scope": "provincial",
        "pollutant": "O3",
        "data_rows": (2, 22),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "M",
                "target_value_col": "N",
                "sort_ascending": True,
                "calculate_diff": True,  # 计算同比差值
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}浓度",
        }
    },
    "全省AQI": {
        "scope": "provincial",
        "pollutant": "AQI",
        "data_rows": (2, 22),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "B",
                "target_name_col": "G",
                "target_value_col": None,
                "extra_targets": [
                    {"col": "H", "data_source": "diff_pct"},
                ],
                "sort_ascending": False,  # 达标率越高越好
            },
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": True,  # 按D列（去年数据）升序
            }
        ],
        "headers": {
            "B1": "{year}年{month}月达标率",
        }
    },
}

# 额外sheet配置
EXTRA_SHEET_CONFIG = {
    "X月全国排名": {
        "data_rows": (3, 33),  # 从第3行开始填充数据（第1行标题，第2行表头），31个省份需要到第33行
        "columns": [
            {"pollutant": "PM2.5", "name_col": "A", "value_col": "B", "rank_col": "C", "sort_ascending": True},
            {"pollutant": "PM10", "name_col": "D", "value_col": "E", "rank_col": "F", "sort_ascending": True},
            {"pollutant": "NO2", "name_col": "G", "value_col": "H", "rank_col": "I", "sort_ascending": True},
            {"pollutant": "O3", "name_col": "J", "value_col": "K", "rank_col": "L", "sort_ascending": True},
            {"pollutant": "AQI", "name_col": "M", "value_col": "N", "rank_col": "O", "sort_ascending": False},
        ]
    },
    "全省同比": {
        "data_rows": (3, 9),
        "mapping": {
            3: "PM2.5",
            4: "PM10",
            5: "NO2",
            6: "O3",
            7: "AQI",
            8: "AQI",
            9: "AQI",
        },
        "last_year_col": "B",
        "current_col": "C",
        "transition_col": "D",  # 仅第7行
        "headers": {
            "B2": "{last_year}年{month}月",
            "C2": "{year}年{month}月",
        }
    },
    "历年当月浓度": {
        "start_year": 2014,
        "start_row": 2,
        "columns": [
            {"pollutant": "AQI", "col": "B", "field": "compliance_rate"},
            {"pollutant": "PM2.5", "col": "C", "field": "PM2_5"},
            {"pollutant": "PM10", "col": "D", "field": "PM10"},
            {"pollutant": "NO2", "col": "E", "field": "NO2"},
            {"pollutant": "O3", "col": "F", "field": "O3_8h"},  # O3日均值
            {"pollutant": "SO2", "col": "G", "field": "SO2"},
            {"pollutant": "CO", "col": "H", "field": "CO_P95"},
        ],
        "year_col": "A",  # 年份列
        "standard_col": "I",
        "standard_header": "标准类型",
    },
}


class ConsultationFileFetcher(DataFetcher):
    """
    会商文件批量更新数据获取器

    功能：
    - 每天早上7点自动更新会商Excel文件
    - 使用用户提供的Excel模板
    - 脚本仅填充原始数据（地区名、去年数据、今年数据）
    - 保留模板中的图表、公式和格式
    - 数据范围：本月1号 → 昨天
    """

    def __init__(self):
        super().__init__(
            name="consultation_file_fetcher",
            description="会商文件批量更新 - 每天7点生成当月累积数据（截至昨日）",
            schedule="0 7 * * *",
            version="2.0.0"
        )

        # 会商文件根目录
        self.consultation_root = Path("/tmp/A会商文件")
        self.consultation_root.mkdir(parents=True, exist_ok=True)

        # 模板目录
        self.template_dir = self.consultation_root / "模板"
        self.template_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _city_record_name(record: Dict[str, Any]) -> str:
        """从城市统计报表记录中提取城市名。"""
        for key in (
            "cityName", "CityName", "districtName", "DistrictName",
            "areaName", "AreaName", "name", "Name"
        ):
            value = record.get(key)
            if value:
                return str(value).strip()
        return ""

    @staticmethod
    def _to_float(value: Any, default: float = 0.0) -> float:
        if value is None or value == "":
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _get_city_standard_value(
        self,
        record: Dict[str, Any],
        pollutant_or_field: str,
        *,
        suffix: str = ""
    ) -> float:
        """
        读取城市统计报表的唯一目标字段。

        pollutant_or_field 可传污染物名（PM2.5/AQI）或旧字段名（PM2_5/compliance_rate）。
        suffix 用于同比接口字段，例如 _Compare。

        字段口径：
        - PM2.5/PM10：阶段均值字段 pM2_5_Decimal / pM10_Decimal
        - AQI达标率：fineRate
        - SO2/NO2/O3/CO：修约均值字段 sO2 / nO2 / o3_8h / co
        """
        aliases = CITY_STANDARD_FIELD_ALIASES.get(pollutant_or_field)
        if aliases is None:
            reverse_alias = {
                "PM2_5": "PM2.5",
                "pM2_5_Decimal": "PM2.5",
                "pM10_Decimal": "PM10",
                "O3_8h": "O3",
                "o3_8h_Decimal": "O3",
                "compliance_rate": "AQI",
                "AQIStandardRate": "AQI",
                "CO_P95": "CO",
                "cO_Decimal": "CO",
                "cO": "CO",
            }
            aliases = CITY_STANDARD_FIELD_ALIASES.get(reverse_alias.get(pollutant_or_field, ""), [pollutant_or_field])

        expected_keys = [f"{alias}{suffix}" if suffix else alias for alias in aliases]
        for key in expected_keys:
            if key not in record:
                continue
            value = record.get(key)
            if value is not None and value != "":
                if "_Decimal" in key:
                    logger.debug("city_standard_field_found_decimal", field=key, pollutant=pollutant_or_field)
                return self._to_float(value)

        logger.warning(
            "city_standard_field_missing",
            pollutant_or_field=pollutant_or_field,
            suffix=suffix,
            expected_fields=expected_keys,
            available_keys=list(record.keys())[:20],
        )
        return 0.0

    def _records_by_city_name(self, records: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        result = {}
        for record in records:
            city = self._city_record_name(record)
            if city:
                result[city] = record
        return result

    @staticmethod
    def _historical_standard_label(year: int) -> str:
        """Return the standard label used by historical concentration queries."""
        from app.tools.query.query_city_standard_report.tool import DEFAULT_NEW_STANDARD_START

        return "新标准" if year >= DEFAULT_NEW_STANDARD_START.year else "旧标准"

    @staticmethod
    def _extract_report_records(
        query_result: Dict[str, Any],
        *,
        preferred_view: str = "raw",
        result_name: str = "query_city_standard_report",
    ) -> List[Dict[str, Any]]:
        """Return original API records from report tools.

        City standard report tools now expose reporting-view rows in data and
        store original API fields in report_data_id views. Consultation Excel
        filling still needs original fields such as cityName/pM2_5_Decimal.
        """
        records = query_result.get("result")
        if isinstance(records, list):
            return [record for record in records if isinstance(record, dict)]

        report_data_id = query_result.get("report_data_id") or (query_result.get("metadata") or {}).get("report_data_id")
        if report_data_id:
            try:
                from app.services.data_registry import data_registry

                package = data_registry.load_dataset(report_data_id)
                views = package.get("views") if isinstance(package, dict) else None
                if isinstance(views, dict):
                    for view_name in (preferred_view, "result", "raw"):
                        view_records = views.get(view_name)
                        if isinstance(view_records, list):
                            return [record for record in view_records if isinstance(record, dict)]
                    raise ValueError(
                        f"{result_name} report_data_id={report_data_id} has no usable raw/result view"
                    )
            except Exception as exc:
                logger.warning(
                    "consultation_report_view_load_failed",
                    report_data_id=report_data_id,
                    preferred_view=preferred_view,
                    error=str(exc),
                )
                raise ValueError(
                    f"{result_name} raw records could not be loaded from report_data_id={report_data_id}"
                ) from exc

        records = query_result.get("data") or []
        if isinstance(records, list):
            return [record for record in records if isinstance(record, dict)]

        raise ValueError(f"{result_name} result is not a list")

    async def _query_city_standard_records(
        self,
        start_date: str,
        end_date: str,
        *,
        cities: List[str] = None,
        pollutant_codes: List[str] = None,
        ns_type: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """调用统一的城市统计报表工具。"""
        from app.tools.query.query_city_standard_report.tool import execute_query_city_standard_report

        query_result = await execute_query_city_standard_report(
            cities=cities or GUANGDONG_CITIES,
            start_time=start_date,
            end_time=end_date,
            ns_type=ns_type,
            pollutant_codes=pollutant_codes,
            data_source=1,
            sand_type=1,
            context=None,
        )
        if not query_result or not query_result.get("success"):
            raise ValueError((query_result or {}).get("summary") or "query_city_standard_report returned empty result")
        return self._extract_report_records(
            query_result,
            preferred_view="raw",
            result_name="query_city_standard_report",
        )

    async def _query_city_standard_yoy_records(
        self,
        current_start: str,
        current_end: str,
        last_year_start: str,
        last_year_end: str,
        *,
        cities: List[str] = None,
        pollutant_codes: List[str] = None,
        ns_type: int = 2,
    ) -> List[Dict[str, Any]]:
        """
        获取城市标准同比数据。

        使用新版 query_city_standard_yoy_report；完整接口字段从 report_data_id 的 raw
        视图读取，避免使用默认 reporting 视图时丢失 cityName/pM2_5_Decimal 等填表字段。
        """
        from app.tools.query.query_city_standard_report.tool import execute_query_city_standard_yoy_report

        query_result = await execute_query_city_standard_yoy_report(
            cities=cities or GUANGDONG_CITIES,
            time_point=[current_start, current_end],
            contrast_time=[last_year_start, last_year_end],
            ns_type=ns_type,
            pollutant_codes=pollutant_codes,
            data_source=1,
            sand_type=1,
            context=None,
        )

        if not query_result or not query_result.get("success"):
            raise ValueError((query_result or {}).get("summary") or "query_city_standard_yoy_report returned empty result")

        return self._extract_report_records(
            query_result,
            preferred_view="raw",
            result_name="query_city_standard_yoy_report",
        )

    def _aggregate_city_standard_values(
        self,
        records: List[Dict[str, Any]],
        pollutant_or_field: str,
        *,
        suffix: str = ""
    ) -> float:
        values = [
            self._get_city_standard_value(record, pollutant_or_field, suffix=suffix)
            for record in records
        ]
        values = [value for value in values if value is not None]
        if not values:
            return 0.0
        return sum(values) / len(values)

    def _aggregate_all_city_standard_values(
        self,
        records: List[Dict[str, Any]],
        *,
        suffix: str = ""
    ) -> Dict[str, float]:
        return {
            pollutant: self._aggregate_city_standard_values(records, pollutant, suffix=suffix)
            for pollutant in ("PM2.5", "PM10", "NO2", "O3", "AQI")
        }

    @staticmethod
    def _competition_rankings(sorted_pairs: List[Tuple[str, float]]) -> Dict[str, int]:
        """按已排序列表计算并列排名：1, 2, 2, 4。"""
        rankings: Dict[str, int] = {}
        previous_value = None
        previous_rank = 0

        for index, (name, value) in enumerate(sorted_pairs, start=1):
            if previous_value is None or value != previous_value:
                previous_rank = index
                previous_value = value
            rankings[name] = previous_rank

        return rankings

    def _period_cache_key(
        self,
        time_range: Dict[str, str],
        last_year_start: str,
        last_year_end: str,
    ) -> Tuple[str, str, str, str]:
        return (
            time_range["start_date"],
            time_range["end_date"],
            last_year_start,
            last_year_end,
        )

    async def _replace_guangdong_in_national_cache(
        self,
        current_data: Dict[str, Any],
        last_year_data: Dict[str, Any],
        time_range: Dict[str, str],
        last_year_start: str,
        last_year_end: str,
    ) -> None:
        """用广东审核后的全省数据替换全国数据中的广东。"""
        if "广东" not in current_data.get("area_names", []):
            return

        try:
            logger.info("guangdong_data_replace_start_national")
            guangdong_data = await self._get_guangdong_province_data(
                pollutant="PM2.5",
                current_start=time_range["start_date"],
                current_end=time_range["end_date"],
                last_year_start=last_year_start,
                last_year_end=last_year_end,
                return_all_data=True,
            )

            if not guangdong_data:
                logger.warning("guangdong_national_data_empty")
                return

            area_names = current_data.get("area_names", [])
            guangdong_index = area_names.index("广东")
            all_current = current_data.get("all_data", {})
            all_last_year = last_year_data.get("all_data", {})

            if "广东" not in all_current or "广东" not in all_last_year:
                logger.warning("guangdong_not_in_national_data")
                return

            logger.info(
                "guangdong_national_data_replacing",
                index=guangdong_index,
                old_current=all_current["广东"],
                new_current=guangdong_data["current"],
            )

            all_current["广东"] = guangdong_data["current"]
            all_last_year["广东"] = guangdong_data["last_year"]
        except Exception as e:
            logger.error("guangdong_national_replace_failed", error=str(e), exc_info=True)

    async def fetch_and_store(self, full_month: bool = False):
        """
        获取并存储会商数据

        Args:
            full_month: 是否查询完整月份数据（1号-月末），False表示截至昨天

        流程：
        1. 计算时间范围（本月1号 → 昨天 或 完整月份）
        2. 创建当月子目录
        3. 复制模板到输出目录
        4. 查询全国/全省空气质量数据
        5. 填充模板各sheet（保留图表和公式）
        6. 保存文件
        """
        try:
            logger.info("consultation_file_fetch_start", full_month=full_month)

            # 计算时间范围
            time_range = self._calculate_month_to_yesterday(full_month=full_month)
            logger.info(
                "consultation_file_time_range",
                start_date=time_range["start_date"],
                end_date=time_range["end_date"],
                period_description=time_range["period_description"]
            )

            # 创建当月子目录
            month_dir = self._get_month_dir()
            month_dir.mkdir(parents=True, exist_ok=True)
            logger.info("consultation_file_month_dir", month_dir=str(month_dir))

            # 复制模板到输出目录
            template_path = self._get_template_path(time_range)
            output_path = self._get_output_path(time_range, month_dir)

            if not template_path.exists():
                logger.error("template_not_found", template_path=str(template_path))
                raise FileNotFoundError(f"模板文件不存在: {template_path}")

            shutil.copy2(str(template_path), str(output_path))
            logger.info("template_copied", template=str(template_path), output=str(output_path))

            # 打开工作簿并填充数据
            import openpyxl
            wb = openpyxl.load_workbook(str(output_path))
            self._consultation_period_cache = {}

            # 填充10个污染物sheet
            await self._fill_pollutant_sheets(wb, time_range, full_month=full_month)

            # 填充额外sheet
            await self._fill_extra_sheets(wb, time_range, full_month=full_month)

            # 保存
            wb.save(str(output_path))
            wb.close()

            # 使用LibreOffice重新保存以正确保留图表渲染
            self._resave_with_libreoffice(output_path)

            logger.info(
                "consultation_file_fetch_complete",
                output_path=str(output_path),
                sheets_updated=len(SHEET_CONFIG) + len(EXTRA_SHEET_CONFIG)
            )

        except Exception as e:
            logger.error("consultation_file_fetch_failed", error=str(e), exc_info=True)
            raise

    def _get_template_path(self, time_range: Dict[str, str]) -> Path:
        """
        获取模板文件路径

        模板命名规则：
        - 单月模板：月度会商模板（某月）.xlsx
        - 累计模板：月度会商模板（1-某月）.xlsx
        """
        year = time_range["year"]
        month = int(time_range["month"])

        # 尝试单月模板
        single_template = self.template_dir / f"月度会商模板（{year}年{month}月）.xlsx"
        if single_template.exists():
            return single_template

        # 尝试累计模板
        cumulative_template = self.template_dir / f"月度会商模板（1-{month}月）.xlsx"
        if cumulative_template.exists():
            return cumulative_template

        # 回退：查找任意匹配的模板
        for f in self.template_dir.glob("月度会商模板*.xlsx"):
            return f

        return single_template

    def _get_output_path(self, time_range: Dict[str, str], month_dir: Path) -> Path:
        """获取输出文件路径"""
        year = time_range["year"]
        month = int(time_range["month"])
        today_str = datetime.now().strftime("%Y%m%d")
        return month_dir / f"月度会商模板（{year}年{month}月）{today_str}.xlsx"

    def _calculate_month_to_yesterday(self, full_month: bool = False) -> Dict[str, str]:
        """
        计算时间范围

        Args:
            full_month: 是否查询完整月份数据（1号-月末），False表示截至昨天

        Returns:
            如果 full_month=False（实时更新模式）:
            {
                "start_date": "2026-05-01",
                "end_date": "2026-05-15",  # 昨天
                "period_description": "2026年5月份累计（截至5月15日）",
                "year": "2026",
                "month": "05",
                "last_year": "2025"
            }

            如果 full_month=True（完整月份模式）:
            - 今天是4号：查询上个月完整数据
              {
                "start_date": "2026-04-01",
                "end_date": "2026-04-30",
                "period_description": "2026年4月份",
                "year": "2026",
                "month": "04",
                "last_year": "2025"
              }
            - 其他日期：查询本月完整数据
              {
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "period_description": "2026年5月份",
                "year": "2026",
                "month": "05",
                "last_year": "2025"
              }
        """
        today = datetime.now()

        if full_month:
            # 完整月份模式
            year = today.year
            month = today.month

            # 如果今天是4号，查询上个月完整数据（确保数据审核完成）
            if today.day == 4:
                last_month = today.replace(day=1) - timedelta(days=1)
                year = last_month.year
                month = last_month.month

            # 获取该月最后一天
            from calendar import monthrange
            last_day = monthrange(year, month)[1]

            first_day_of_month = datetime(year, month, 1)
            last_day_of_month = datetime(year, month, last_day)

            start_date = first_day_of_month.strftime("%Y-%m-%d")
            end_date = last_day_of_month.strftime("%Y-%m-%d")
            period_description = f"{year}年{int(month):02d}月份"
        else:
            # 实时更新模式：查询1号到昨天
            yesterday = today - timedelta(days=1)
            first_day_of_month = today.replace(day=1)

            # 如果今天是1号，则昨天是上个月最后一天
            if today.day == 1:
                last_month = today.replace(day=1) - timedelta(days=1)
                first_day_of_month = last_month.replace(day=1)
                yesterday = last_month
                year = str(last_month.year)
                month = str(last_month.month)
            else:
                year = str(today.year)
                month = str(today.month)

            start_date = first_day_of_month.strftime("%Y-%m-%d")
            end_date = yesterday.strftime("%Y-%m-%d")

            # 生成时间段描述
            if first_day_of_month.month == yesterday.month:
                if yesterday.day == first_day_of_month.day:
                    period_description = f"{year}年{month}月份累计（截至{month}月1日）"
                else:
                    period_description = f"{year}年{month}月份累计（截至{month}月{yesterday.day}日）"
            else:
                period_description = f"{year}年{month}月份"

        return {
            "start_date": start_date,
            "end_date": end_date,
            "period_description": period_description,
            "year": str(year),
            "month": f"{int(month):02d}",
            "last_year": str(int(year) - 1),
        }

    def _get_month_dir(self) -> Path:
        """获取当月子目录路径"""
        today = datetime.now()
        if today.day == 1:
            last_month = today.replace(day=1) - timedelta(days=1)
            year = last_month.year
            month = last_month.month
        else:
            year = today.year
            month = today.month
        month_dir_name = f"{year}年{month}月"
        return self.consultation_root / month_dir_name

    def _get_last_day_of_month(self, year: int, month: int) -> str:
        """获取指定月份的最后一天"""
        from calendar import monthrange
        last_day = monthrange(year, month)[1]
        return f"{year}-{month:02d}-{last_day:02d}"

    def _get_last_year_start_date(self, time_range: Dict[str, str]) -> str:
        """按当前查询起点计算去年同期起点。"""
        current_start = datetime.strptime(time_range["start_date"], "%Y-%m-%d")
        last_year = int(time_range["last_year"])
        try:
            last_year_start = current_start.replace(year=last_year)
        except ValueError:
            last_year_start = current_start.replace(year=last_year, day=28)
        return last_year_start.strftime("%Y-%m-%d")

    def _format_period_sheet_prefix(self, time_range: Dict[str, str]) -> str:
        """生成额外 sheet 的月份前缀，例如 4月 或 1-4月。"""
        start = datetime.strptime(time_range["start_date"], "%Y-%m-%d")
        end = datetime.strptime(time_range["end_date"], "%Y-%m-%d")
        if start.month == end.month:
            return f"{end.month}月"
        return f"{start.month}-{end.month}月"

    def _get_last_year_same_day(self, time_range: Dict[str, str], full_month: bool = False) -> str:
        """
        计算去年对应的日期

        Args:
            time_range: 时间范围字典，包含 year、month 等字段
            full_month: 是否为完整月份模式

        Returns:
            日期字符串（格式：YYYY-MM-DD）

        Examples:
            full_month=False (默认):
                今天是2026-05-13 → 返回 "2025-05-12"（去年同日）
                今天是2026-03-01 → 返回 "2025-02-28"（上月最后一天）

            full_month=True:
                查询2026年4月完整数据 → 返回 "2025-04-30"（去年同月最后一天）
        """
        if full_month:
            # 完整月份模式：返回去年同月的最后一天
            year = int(time_range["last_year"])
            month = int(time_range["month"])

            # 获取去年该月的最后一天
            from calendar import monthrange
            last_day = monthrange(year, month)[1]

            return f"{year}-{month:02d}-{last_day:02d}"
        else:
            # 实时更新模式：返回去年同日
            current_date = datetime.now()

            if current_date.day == 1:
                # 如果今天是1号，则昨天是上个月最后一天
                yesterday = current_date.replace(day=1) - timedelta(days=1)
                last_year_date = datetime(
                    int(time_range["last_year"]),
                    yesterday.month,
                    yesterday.day
                )
            else:
                # 正常情况：去年同月同日（昨天对应去年的日期）
                last_year_date = datetime(
                    int(time_range["last_year"]),
                    int(time_range["month"]),
                    current_date.day - 1  # 昨天对应去年的日期
                )

            # 处理闰年2月29日的情况（如果去年不是闰年，则取2月28日）
            if last_year_date.month == 2 and last_year_date.day == 29:
                if not self._is_leap_year(last_year_date.year):
                    last_year_date = last_year_date.replace(day=28)

            return last_year_date.strftime("%Y-%m-%d")

    def _resave_with_libreoffice(self, file_path: Path) -> bool:
        """
        使用LibreOffice重新保存Excel文件以正确保留图表渲染

        openpyxl只能保留图表元数据，但无法正确保存图表渲染信息。
        使用LibreOffice重新保存可以完整保留所有图表和格式。

        Args:
            file_path: Excel文件路径

        Returns:
            bool: 是否成功
        """
        try:
            # 检查LibreOffice是否可用
            soffice_result = subprocess.run(
                ["which", "soffice"],
                capture_output=True,
                text=True
            )

            if soffice_result.returncode != 0:
                logger.warning("libreoffice_not_found", file=str(file_path))
                return False

            # 创建临时目录用于LibreOffice输出
            import tempfile
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = Path(temp_dir)

                # 使用LibreOffice转换为xlsx（会重新保存文件）
                env = {
                    "SAL_USE_VCLPLUGIN": "svp",
                    **dict(subprocess.os.environ)
                }

                result = subprocess.run(
                    [
                        "soffice",
                        "--headless",
                        "--convert-to", "xlsx",
                        "--outdir", str(temp_path),
                        str(file_path)
                    ],
                    env=env,
                    capture_output=True,
                    text=True,
                    timeout=60
                )

                if result.returncode == 0:
                    # 查找生成的文件
                    converted_files = list(temp_path.glob("*.xlsx"))
                    if converted_files:
                        # 用LibreOffice重新保存的文件替换原文件
                        shutil.move(str(converted_files[0]), str(file_path))
                        logger.info("libreoffice_resave_success", file=str(file_path))
                        return True
                    else:
                        logger.warning("libreoffice_resave_no_output", file=str(file_path))
                        return False
                else:
                    logger.warning(
                        "libreoffice_resave_failed",
                        file=str(file_path),
                        stderr=result.stderr
                    )
                    return False

        except subprocess.TimeoutExpired:
            logger.warning("libreoffice_resave_timeout", file=str(file_path))
            return False
        except Exception as e:
            logger.warning("libreoffice_resave_error", file=str(file_path), error=str(e))
            return False

    async def _fill_pollutant_sheets(self, wb, time_range: Dict[str, str], full_month: bool = False):
        """填充10个污染物sheet（优化：全省和全国数据都只查询一次）

        Args:
            wb: 工作簿对象
            time_range: 时间范围字典
            full_month: 是否查询完整月份数据
        """

        # 第一步：识别全省和全国sheet并预查询数据（避免重复查询）
        provincial_sheets = []
        national_sheets = []
        provincial_cache = {"current": None, "last_year": None}
        national_cache = {"current": None, "last_year": None}

        for sheet_name, config in SHEET_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                logger.warning("sheet_not_found", sheet=sheet_name)
                continue

            scope = config.get("scope")
            if scope == "provincial":
                provincial_sheets.append((sheet_name, config))
            elif scope == "national":
                national_sheets.append((sheet_name, config))

        # 第二步：如果有全省sheet，预查询一次全省数据
        if provincial_sheets:
            logger.info("provincial_sheets_pre_query", count=len(provincial_sheets))
            try:
                # 查询今年全省数据（一次查询获取所有污染物）
                _, current_data = await self._query_with_date_range(
                    scope="provincial",
                    pollutant="PM2.5",  # 污染物参数不重要，会返回所有污染物
                    start_date=time_range["start_date"],
                    end_date=time_range["end_date"],
                    return_all_data=True  # 返回完整数据（所有污染物）
                )
                provincial_cache["current"] = current_data

                # 查询去年全省数据
                last_year_start = self._get_last_year_start_date(time_range)
                last_year_end = self._get_last_year_same_day(time_range, full_month=full_month)
                _, last_year_data = await self._query_with_date_range(
                    scope="provincial",
                    pollutant="PM2.5",
                    start_date=last_year_start,
                    end_date=last_year_end,
                    return_all_data=True
                )
                provincial_cache["last_year"] = last_year_data

                logger.info("provincial_data_cached",
                           current_cities=len(current_data.get("area_names", [])),
                           last_year_cities=len(last_year_data.get("area_names", [])))
            except Exception as e:
                logger.error("provincial_pre_query_failed", error=str(e), exc_info=True)
                provincial_cache = {"current": None, "last_year": None}

        # 第三步：如果有全国sheet，预查询一次全国数据
        if national_sheets:
            logger.info("national_sheets_pre_query", count=len(national_sheets))
            try:
                # 查询今年全国数据（一次查询获取所有污染物）
                _, current_data = await self._query_with_date_range(
                    scope="national",
                    pollutant="PM2.5",  # 污染物参数不重要，会返回所有污染物
                    start_date=time_range["start_date"],
                    end_date=time_range["end_date"],
                    return_all_data=True  # 返回完整数据（所有污染物）
                )
                national_cache["current"] = current_data

                # 查询去年全国数据
                last_year_start = self._get_last_year_start_date(time_range)
                last_year_end = self._get_last_year_same_day(time_range, full_month=full_month)
                _, last_year_data = await self._query_with_date_range(
                    scope="national",
                    pollutant="PM2.5",
                    start_date=last_year_start,
                    end_date=last_year_end,
                    return_all_data=True
                )
                national_cache["last_year"] = last_year_data

                await self._replace_guangdong_in_national_cache(
                    current_data,
                    last_year_data,
                    time_range,
                    last_year_start,
                    last_year_end,
                )

                self._consultation_period_cache["national"] = {
                    "key": self._period_cache_key(time_range, last_year_start, last_year_end),
                    "current": current_data,
                    "last_year": last_year_data,
                }

                logger.info("national_data_cached",
                           current_provinces=len(current_data.get("area_names", [])),
                           last_year_provinces=len(last_year_data.get("area_names", [])))
            except Exception as e:
                logger.error("national_pre_query_failed", error=str(e), exc_info=True)
                national_cache = {"current": None, "last_year": None}

        # 第四步：填充所有sheet
        for sheet_name, config in SHEET_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                logger.warning("sheet_not_found", sheet=sheet_name)
                continue

            try:
                scope = config.get("scope")
                # 全省sheet使用缓存数据
                if scope == "provincial" and provincial_cache["current"]:
                    await self._fill_single_sheet_with_cache(
                        wb, sheet_name, config, time_range,
                        provincial_cache=provincial_cache
                    )
                # 全国sheet使用缓存数据
                elif scope == "national" and national_cache["current"]:
                    await self._fill_single_sheet_with_cache(
                        wb, sheet_name, config, time_range,
                        national_cache=national_cache
                    )
                else:
                    # 其他情况正常查询
                    await self._fill_single_sheet(
                        wb,
                        sheet_name,
                        config,
                        time_range,
                        full_month=full_month,
                    )

                logger.info("sheet_filled", sheet=sheet_name)
            except Exception as e:
                logger.error("sheet_fill_failed", sheet=sheet_name, error=str(e))


    async def _fill_single_sheet(
        self,
        wb,
        sheet_name: str,
        config: Dict[str, Any],
        time_range: Dict[str, str],
        full_month: bool = False
    ):
        """填充单个sheet

        Args:
            wb: 工作簿对象
            sheet_name: sheet名称
            config: sheet配置
            time_range: 时间范围字典
            full_month: 是否查询完整月份数据
        """
        ws = wb[sheet_name]
        scope = config["scope"]
        pollutant = config["pollutant"]
        start_row, end_row = config["data_rows"]

        # 查询今年数据
        area_names, current_data = await self._query_with_date_range(
            scope=scope,
            pollutant=pollutant,
            start_date=time_range["start_date"],
            end_date=time_range["end_date"]
        )

        # 查询去年数据（去年同期累积：去年同月1号 → 去年同日/月末）
        last_year_start = self._get_last_year_start_date(time_range)
        last_year_end = self._get_last_year_same_day(time_range, full_month=full_month)
        _, last_year_data = await self._query_with_date_range(
            scope=scope,
            pollutant=pollutant,
            start_date=last_year_start,
            end_date=last_year_end
        )

        # 确保数据长度匹配
        data_len = min(len(area_names), len(current_data), len(last_year_data))
        area_names = area_names[:data_len]
        current_data = current_data[:data_len]
        last_year_data = last_year_data[:data_len]

        # 替换广东省数据（全国sheet需要用审核后的全省数据替换）
        if scope == "national" and "广东" in area_names:
            try:
                logger.info("guangdong_data_replace_start", pollutant=pollutant)

                # 获取广东省的审核后数据
                guangdong_data = await self._get_guangdong_province_data(
                    pollutant=pollutant,
                    current_start=time_range["start_date"],
                    current_end=time_range["end_date"],
                    last_year_start=last_year_start,
                    last_year_end=last_year_end
                )

                if guangdong_data:
                    # 找到广东在列表中的索引
                    guangdong_index = area_names.index("广东")

                    # 记录原始数据
                    original_current = current_data[guangdong_index]
                    original_last_year = last_year_data[guangdong_index]

                    # 替换广东数据
                    current_data[guangdong_index] = guangdong_data["current"]
                    last_year_data[guangdong_index] = guangdong_data["last_year"]

                    logger.info(
                        "guangdong_data_replaced",
                        pollutant=pollutant,
                        original_current=original_current,
                        new_current=guangdong_data["current"],
                        original_last_year=original_last_year,
                        new_last_year=guangdong_data["last_year"]
                    )
                else:
                    logger.warning("guangdong_data_is_none", pollutant=pollutant)

            except Exception as e:
                logger.error("guangdong_data_replace_failed", pollutant=pollutant, error=str(e), exc_info=True)

        # 主数据区按B列（今年数据）升序排序
        paired_data = list(zip(area_names, current_data, last_year_data))
        paired_data.sort(key=lambda x: x[1])  # 按今年数据升序
        area_names = [item[0] for item in paired_data]
        current_data = [item[1] for item in paired_data]
        last_year_data = [item[2] for item in paired_data]

        # 填充主数据区
        name_col = config.get("name_col")
        current_col = config.get("current_col")
        last_year_col = config.get("last_year_col")

        for i in range(data_len):
            row = start_row + i
            if name_col:
                ws[f"{name_col}{row}"] = area_names[i]
            if current_col:
                ws[f"{current_col}{row}"] = current_data[i]
            if last_year_col:
                ws[f"{last_year_col}{row}"] = last_year_data[i]

        # 构建名称到数据的映射（用于排序副本的额外列填充）
        name_to_current = dict(zip(area_names, current_data))
        name_to_last_year = dict(zip(area_names, last_year_data))

        # 填充排序副本
        for copy_config in config.get("sort_copies", []):
            source_name_col = copy_config["source_name_col"]
            source_value_col = copy_config["source_value_col"]
            target_name_col = copy_config["target_name_col"]
            target_value_col = copy_config.get("target_value_col")
            sort_ascending = copy_config.get("sort_ascending", True)
            calculate_diff = copy_config.get("calculate_diff", False)

            # 读取源数据
            names = []
            values = []
            for i in range(data_len):
                row = start_row + i
                n = ws[f"{source_name_col}{row}"].value
                v = ws[f"{source_value_col}{row}"].value
                if n is not None and v is not None:
                    names.append(n)
                    try:
                        if calculate_diff:
                            # 需要计算同比差值
                            current_val = name_to_current.get(n, 0)
                            last_year_val = name_to_last_year.get(n, 0)
                            diff_val = current_val - last_year_val
                            values.append(diff_val)
                        else:
                            # 直接使用源列的值
                            values.append(float(v))
                    except (ValueError, TypeError):
                        values.append(0.0)

            # 排序
            paired = list(zip(names, values))
            paired.sort(key=lambda x: x[1], reverse=not sort_ascending)

            # 填充目标区域
            for i, (name, value) in enumerate(paired):
                row = start_row + i
                ws[f"{target_name_col}{row}"] = name
                if target_value_col:
                    ws[f"{target_value_col}{row}"] = value
                # 填充额外列
                for extra in copy_config.get("extra_targets", []):
                    extra_col = extra["col"]
                    data_source = extra["data_source"]
                    if data_source == "diff_pct":
                        if name in name_to_current and name in name_to_last_year:
                            diff = name_to_current[name] - name_to_last_year[name]
                            ws[f"{extra_col}{row}"] = diff
                    else:
                        source_map = name_to_current if data_source == "current" else name_to_last_year
                        if name in source_map:
                            ws[f"{extra_col}{row}"] = source_map[name]

        # 更新表头
        for cell_ref, template in config.get("headers", {}).items():
            header_value = template.format(
                year=time_range["year"],
                month=int(time_range["month"]),
                last_year=time_range["last_year"],
            )
            ws[cell_ref] = header_value


    async def _fill_single_sheet_with_cache(
        self,
        wb,
        sheet_name: str,
        config: Dict[str, Any],
        time_range: Dict[str, str],
        provincial_cache: Dict[str, Any] = None,
        national_cache: Dict[str, Any] = None
    ):
        """使用缓存数据填充单个全省或全国sheet（优化：避免重复查询）"""
        ws = wb[sheet_name]
        pollutant = config["pollutant"]
        scope = config.get("scope", "provincial")  # 默认为全省
        start_row, end_row = config["data_rows"]

        # 根据scope选择对应的缓存
        if scope == "national":
            current_cache = national_cache["current"] if national_cache else None
            last_year_cache = national_cache["last_year"] if national_cache else None
            cache_type = "national"
        else:
            current_cache = provincial_cache["current"] if provincial_cache else None
            last_year_cache = provincial_cache["last_year"] if provincial_cache else None
            cache_type = "provincial"

        if not current_cache or not last_year_cache:
            logger.error(f"{cache_type}_cache_empty", sheet=sheet_name)
            raise ValueError(f"{cache_type.capitalize()} cache is empty for {sheet_name}")

        # 提取污染物数据
        area_names = current_cache.get("area_names", [])
        all_current_data = current_cache.get("all_data", {})
        all_last_year_data = last_year_cache.get("all_data", {})
        
        # 提取当前污染物的数据
        current_data = []
        last_year_data = []
        
        for city in area_names:
            if city in all_current_data and city in all_last_year_data:
                current_data.append(all_current_data[city].get(pollutant, 0))
                last_year_data.append(all_last_year_data[city].get(pollutant, 0))
            else:
                logger.warning("city_data_missing_in_cache", city=city, pollutant=pollutant)
                current_data.append(0)
                last_year_data.append(0)
        
        logger.info(
            f"{cache_type}_sheet_filled_from_cache",
            sheet=sheet_name,
            pollutant=pollutant,
            area_count=len(area_names),
            scope=scope
        )
        
        # 确保数据长度匹配
        data_len = min(len(area_names), len(current_data), len(last_year_data))
        area_names = area_names[:data_len]
        current_data = current_data[:data_len]
        last_year_data = last_year_data[:data_len]
        
        # 主数据区按B列（今年数据）升序排序
        paired_data = list(zip(area_names, current_data, last_year_data))
        paired_data.sort(key=lambda x: x[1])  # 按今年数据升序
        area_names = [item[0] for item in paired_data]
        current_data = [item[1] for item in paired_data]
        last_year_data = [item[2] for item in paired_data]
        
        # 填充主数据区
        name_col = config.get("name_col")
        current_col = config.get("current_col")
        last_year_col = config.get("last_year_col")
        
        for i in range(data_len):
            row = start_row + i
            if name_col:
                ws[f"{name_col}{row}"] = area_names[i]
            if current_col:
                ws[f"{current_col}{row}"] = current_data[i]
            if last_year_col:
                ws[f"{last_year_col}{row}"] = last_year_data[i]
        
        # 构建名称到数据的映射（用于排序副本的额外列填充）
        name_to_current = dict(zip(area_names, current_data))
        name_to_last_year = dict(zip(area_names, last_year_data))
        
        # 填充排序副本
        for copy_config in config.get("sort_copies", []):
            source_name_col = copy_config["source_name_col"]
            source_value_col = copy_config["source_value_col"]
            target_name_col = copy_config["target_name_col"]
            target_value_col = copy_config.get("target_value_col")
            sort_ascending = copy_config.get("sort_ascending", True)
            calculate_diff = copy_config.get("calculate_diff", False)
            
            # 读取源数据
            names = []
            values = []
            for i in range(data_len):
                row = start_row + i
                n = ws[f"{source_name_col}{row}"].value
                v = ws[f"{source_value_col}{row}"].value
                if n is not None and v is not None:
                    names.append(n)
                    try:
                        if calculate_diff:
                            # 需要计算同比差值
                            current_val = name_to_current.get(n, 0)
                            last_year_val = name_to_last_year.get(n, 0)
                            diff_val = current_val - last_year_val
                            values.append(diff_val)
                        else:
                            # 直接使用源列的值
                            values.append(float(v))
                    except (ValueError, TypeError):
                        values.append(0.0)
            
            # 排序
            paired = list(zip(names, values))
            paired.sort(key=lambda x: x[1], reverse=not sort_ascending)
            
            # 填充目标区域
            for i, (name, value) in enumerate(paired):
                row = start_row + i
                ws[f"{target_name_col}{row}"] = name
                if target_value_col:
                    ws[f"{target_value_col}{row}"] = value
                # 填充额外列
                for extra in copy_config.get("extra_targets", []):
                    extra_col = extra["col"]
                    data_source = extra["data_source"]
                    if data_source == "diff_pct":
                        if name in name_to_current and name in name_to_last_year:
                            diff = name_to_current[name] - name_to_last_year[name]
                            ws[f"{extra_col}{row}"] = diff
                    else:
                        source_map = name_to_current if data_source == "current" else name_to_last_year
                        if name in source_map:
                            ws[f"{extra_col}{row}"] = source_map[name]
        
        # 更新表头
        for cell_ref, template in config.get("headers", {}).items():
            header_value = template.format(
                year=time_range["year"],
                month=int(time_range["month"]),
                last_year=time_range["last_year"],
            )
            ws[cell_ref] = header_value

    async def _fill_extra_sheets(self, wb, time_range: Dict[str, str], full_month: bool = False):
        """填充额外sheet（X月全国排名、全省同比、历年当月浓度）

        Args:
            wb: 工作簿对象
            time_range: 时间范围字典
            full_month: 是否查询完整月份数据
        """
        period_prefix = self._format_period_sheet_prefix(time_range)

        # 查找并填充全国排名sheet（处理可能的名称变体：末尾空格、月份前缀等）
        ranking_sheet = None
        for sheet_name in wb.sheetnames:
            if "全国排名" in sheet_name:
                ranking_sheet = wb[sheet_name]
                # 重命名为标准格式
                new_name = f"{period_prefix}全国排名"
                if sheet_name != new_name:
                    ranking_sheet.title = new_name
                break

        if ranking_sheet:
            try:
                await self._fill_national_ranking_sheet(wb, time_range, ranking_sheet.title, full_month=full_month)
                logger.info("sheet_filled", sheet=ranking_sheet.title)
            except Exception as e:
                logger.error("sheet_fill_failed", sheet=ranking_sheet.title, error=str(e))

        # 填充全省同比
        if "全省同比" in wb.sheetnames and "全省同比" in EXTRA_SHEET_CONFIG:
            try:
                await self._fill_provincial_summary_sheet(wb, time_range, full_month=full_month)
                logger.info("sheet_filled", sheet="全省同比")
            except Exception as e:
                logger.error("sheet_fill_failed", sheet="全省同比", error=str(e))

        # 查找并填充历年对比sheet（处理可能的名称变体）
        historical_sheet = None
        for sheet_name in wb.sheetnames:
            if "历年" in sheet_name and "浓度" in sheet_name:
                historical_sheet = wb[sheet_name]
                new_name = f"{period_prefix}浓度"
                if sheet_name != new_name:
                    historical_sheet.title = new_name
                break

        if historical_sheet:
            try:
                await self._fill_historical_comparison_sheet(wb, time_range, historical_sheet.title, full_month=full_month)
                logger.info("sheet_filled", sheet=historical_sheet.title)
            except Exception as e:
                logger.error("sheet_fill_failed", sheet=historical_sheet.title, error=str(e))

    async def _fill_national_ranking_sheet(self, wb, time_range: Dict[str, str], sheet_name: str, full_month: bool = False):
        """
        填充X月全国排名sheet（优化：2次查询 + 广东数据替换）

        功能：
        1. 查询今年和去年的全国数据（一次查询获取所有污染物）
        2. 替换广东数据为审核数据（使用 query_city_standard_yoy_report）
        3. 分别排序获得排名
        4. 使用字典方式填充，避免重复填充问题
        5. 对广东添加排名变化标记（↑X/↓X/-）

        优化：
        - 原实现：5个污染物 × 2个时间段 = 10次查询
        - 优化后：2次全国查询 + 2次广东替换查询 = 4次查询
        """
        ws = wb[sheet_name]
        config = EXTRA_SHEET_CONFIG["X月全国排名"]
        start_row, end_row = config["data_rows"]

        # 计算去年同月的时间范围
        last_year_start = self._get_last_year_start_date(time_range)

        last_year_end = self._get_last_year_same_day(time_range, full_month=full_month)

        logger.info(
            "national_ranking_pre_query",
            current_period=f"{time_range['start_date']} → {time_range['end_date']}",
            last_year_period=f"{last_year_start} → {last_year_end}"
        )

        cache_key = self._period_cache_key(time_range, last_year_start, last_year_end)
        cached_national = getattr(self, "_consultation_period_cache", {}).get("national")
        if cached_national and cached_national.get("key") == cache_key:
            current_data = cached_national["current"]
            last_year_data = cached_national["last_year"]
            logger.info("national_ranking_using_cached_data", sheet=sheet_name)
        else:
            # 第一步：预查询全国数据（一次查询获取所有污染物）
            _, current_data = await self._query_with_date_range(
                scope="national",
                pollutant="PM2.5",  # 污染物参数不重要
                start_date=time_range["start_date"],
                end_date=time_range["end_date"],
                return_all_data=True
            )

            _, last_year_data = await self._query_with_date_range(
                scope="national",
                pollutant="PM2.5",
                start_date=last_year_start,
                end_date=last_year_end,
                return_all_data=True
            )

            await self._replace_guangdong_in_national_cache(
                current_data,
                last_year_data,
                time_range,
                last_year_start,
                last_year_end,
            )
            self._consultation_period_cache["national"] = {
                "key": cache_key,
                "current": current_data,
                "last_year": last_year_data,
            }

        # 第三步：填充各污染物列
        for col_config in config["columns"]:
            pollutant = col_config["pollutant"]
            name_col = col_config["name_col"]
            value_col = col_config["value_col"]
            rank_col = col_config["rank_col"]
            sort_ascending = col_config["sort_ascending"]

            # 从缓存中提取污染物数据
            area_names = current_data.get("area_names", [])
            all_current = current_data.get("all_data", {})
            all_last_year = last_year_data.get("all_data", {})

            # 提取当前污染物的数据
            current_values = [all_current.get(area, {}).get(pollutant, 0) for area in area_names]
            last_year_values = [all_last_year.get(area, {}).get(pollutant, 0) for area in area_names]

            # 今年排序获得排名
            current_paired = list(zip(area_names, current_values))
            current_paired.sort(key=lambda x: x[1], reverse=not sort_ascending)
            current_ranking = self._competition_rankings(current_paired)

            # 去年排序获得排名
            last_paired = list(zip(area_names, last_year_values))
            last_paired.sort(key=lambda x: x[1], reverse=not sort_ascending)
            last_ranking = self._competition_rankings(last_paired)

            # 先清空数据区域（避免重复填充）
            for row in range(start_row, end_row + 1):
                ws[f"{name_col}{row}"] = None
                ws[f"{value_col}{row}"] = None
                ws[f"{rank_col}{row}"] = None

            # 使用字典方式填充数据（避免重复）
            province_row_map = {}
            data_len = min(len(current_paired), end_row - start_row + 1)
            for i in range(data_len):
                row = start_row + i
                province_name = current_paired[i][0]
                province_value = current_paired[i][1]
                current_rank = i + 1

                # 建立省份到行号的映射
                province_row_map[province_name] = row

                # 填充省份
                ws[f"{name_col}{row}"] = province_name

                # 填充指标值
                display_value = round(province_value, 2)
                ws[f"{value_col}{row}"] = display_value

                # 填充排名（如果是广东，添加变化标记）
                rank_display = str(current_ranking.get(province_name, current_rank))

                if province_name == "广东" and province_name in last_ranking:
                    current_rank = current_ranking.get(province_name, current_rank)
                    last_rank = last_ranking[province_name]
                    rank_change = last_rank - current_rank

                    if rank_change > 0:
                        rank_display = f"{current_rank}（↑{rank_change}）"
                    elif rank_change < 0:
                        rank_display = f"{current_rank}（↓{abs(rank_change)}）"
                    else:
                        rank_display = f"{current_rank}（-）"

                ws[f"{rank_col}{row}"] = rank_display

        logger.info(
            "national_ranking_filled",
            sheet=sheet_name,
            pollutants_count=len(config["columns"]),
            data_range=f"{start_row}-{end_row}"
        )

    async def _fill_provincial_summary_sheet(self, wb, time_range: Dict[str, str], full_month: bool = False):
        """
        填充全省同比sheet（优化：一次查询获取所有污染物）

        功能：
        1. 查询今年和去年同期的全省数据（如今年1-5月 vs 去年1-5月）
        2. 使用 query_city_standard_yoy_report
        3. 从城市统计报表字段聚合全省均值

        数据源：query_city_standard_yoy_report（审核数据、扣沙）

        优化：
        - 原实现：5个污染物 × 2个时间段 = 10次查询
        - 现实现：1次同比报表查询，返回当前期和对比期字段
        """
        ws = wb["全省同比"]
        config = EXTRA_SHEET_CONFIG["全省同比"]
        start_row, end_row = config["data_rows"]
        mapping = config["mapping"]

        # 时间范围计算
        # 今年数据：本月1号 → 昨天/月末
        current_start = time_range["start_date"]
        current_end = time_range["end_date"]

        # 去年数据：去年同期时段
        last_year_start = self._get_last_year_start_date(time_range)

        # 计算去年结束日期
        if full_month:
            # 完整月份模式：返回去年同月的最后一天
            from calendar import monthrange
            year = int(time_range["last_year"])
            month = int(time_range["month"])
            last_day = monthrange(year, month)[1]
            last_year_end = f"{year}-{month:02d}-{last_day:02d}"
        else:
            # 实时更新模式：返回去年同日
            current_date = datetime.now()
            if current_date.day == 1:
                # 如果今天是1号，则昨天是上个月最后一天
                yesterday = current_date.replace(day=1) - timedelta(days=1)
                last_year_date = datetime(
                    int(time_range["last_year"]),
                    yesterday.month,
                    yesterday.day
                )
            else:
                # 正常情况：去年同月同日
                last_year_date = datetime(
                    int(time_range["last_year"]),
                    int(time_range["month"]),
                    current_date.day - 1  # 昨天对应去年的日期
                )

            # 处理闰年2月29日的情况（如果去年不是闰年，则取2月28日）
            if last_year_date.month == 2 and last_year_date.day == 29:
                if not self._is_leap_year(last_year_date.year):
                    last_year_date = last_year_date.replace(day=28)

            last_year_end = last_year_date.strftime("%Y-%m-%d")

        logger.info(
            "provincial_summary_pre_query",
            current_period=f"{current_start} → {current_end}",
            last_year_period=f"{last_year_start} → {last_year_end}"
        )

        records = await self._query_city_standard_yoy_records(
            current_start,
            current_end,
            last_year_start,
            last_year_end,
            ns_type=2,
        )

        logger.info(
            "provincial_summary_data_extracted",
            records=len(records),
            ns_type=2,
        )

        old_standard_records = await self._query_city_standard_yoy_records(
            current_start,
            current_end,
            last_year_start,
            last_year_end,
            ns_type=1,
        )

        logger.info(
            "provincial_summary_old_standard_data_extracted",
            records=len(old_standard_records),
            ns_type=1,
        )

        # 筛选出"全省"记录
        province_records = [r for r in records if r.get("cityName") == "全省"]
        old_standard_province_records = [
            r for r in old_standard_records if r.get("cityName") == "全省"
        ]

        if not province_records:
            logger.error(
                "province_record_not_found_in_summary",
                message="全省同比sheet：API返回数据中未找到'全省'记录",
                total_records=len(records),
                city_names=[r.get("cityName") for r in records[:10]]
            )
            raise ValueError(
                f"全省同比sheet：API返回数据中未找到'全省'记录。"
                f"返回记录数: {len(records)}，"
                f"城市名称: {[r.get('cityName') for r in records[:10]]}"
            )
        if not old_standard_province_records:
            logger.error(
                "province_old_standard_record_not_found_in_summary",
                message="全省同比sheet：旧标准API返回数据中未找到'全省'记录",
                total_records=len(old_standard_records),
                city_names=[r.get("cityName") for r in old_standard_records[:10]]
            )
            raise ValueError(
                f"全省同比sheet：旧标准API返回数据中未找到'全省'记录。"
                f"返回记录数: {len(old_standard_records)}，"
                f"城市名称: {[r.get('cityName') for r in old_standard_records[:10]]}"
            )

        province_record = province_records[0]
        old_standard_province_record = old_standard_province_records[0]
        logger.info(
            "province_record_found_in_summary",
            city_name=province_record.get("cityName"),
            time_point=province_record.get("timePoint"),
            ns_type=2,
        )
        logger.info(
            "province_old_standard_record_found_in_summary",
            city_name=old_standard_province_record.get("cityName"),
            time_point=old_standard_province_record.get("timePoint"),
            ns_type=1,
        )

        # 第三步：填充各污染物数据。A7为老标准AQI达标率，其余使用新标准/过渡期数据。
        for row, pollutant in mapping.items():
            if row > end_row:
                continue

            try:
                source_record = old_standard_province_record if row == 7 and pollutant == "AQI" else province_record
                source_ns_type = 1 if source_record is old_standard_province_record else 2

                current_value = self._get_city_standard_value(source_record, pollutant)
                last_year_value = self._get_city_standard_value(source_record, pollutant, suffix="_Compare")

                ws[f"{config['current_col']}{row}"] = current_value
                ws[f"{config['last_year_col']}{row}"] = last_year_value

                logger.info(
                    "provincial_comparison_filled",
                    pollutant=pollutant,
                    row=row,
                    ns_type=source_ns_type,
                    current_value=current_value,
                    last_year_value=last_year_value
                )

            except Exception as e:
                logger.error("provincial_comparison_failed", pollutant=pollutant, row=row, error=str(e), exc_info=True)
                # 填充0值
                ws[f"{config['current_col']}{row}"] = 0
                ws[f"{config['last_year_col']}{row}"] = 0

        # 更新表头
        for cell_ref, template in config.get("headers", {}).items():
            header_value = template.format(
                year=time_range["year"],
                month=int(time_range["month"]),
                last_year=time_range["last_year"],
            )
            ws[cell_ref] = header_value

    async def _query_with_date_range(
        self,
        scope: str,
        pollutant: str,
        start_date: str,
        end_date: str,
        return_all_data: bool = False
    ) -> Tuple[List[str], List[float]]:
        """
        使用自定义日期范围查询数据

        数据源：
        - scope="national": 使用 NationalAirQualityQueryTool 查询全国各省数据
        - scope="provincial": 使用 query_city_standard_report 查询广东21个地级市数据

        Args:
            scope: "national" 或 "provincial"
            pollutant: 污染物名称
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            return_all_data: 是否返回所有污染物数据（仅限provincial）

        Returns:
            Tuple[List[str], List[float]]: (地区名称列表, 污染物数值列表)
            如果return_all_data=True，返回字典：{"area_names": [...], "all_data": {...}}
        """
        area_names = []
        result = []

        if scope == "national":
            # 全国数据：使用 NationalAirQualityQueryTool
            from app.tools.query.query_national_air_quality.tool import (
                NationalAirQualityQueryTool
            )

            query_tool = NationalAirQualityQueryTool()

            field_map = {
                "PM2.5": "PM2_5",
                "PM10": "PM10",
                "NO2": "NO2",
                "O3": "O3_8h",
                "AQI": "AQIStandardRate"
            }
            field = field_map.get(pollutant)
            if not field:
                raise ValueError(f"Unknown pollutant: {pollutant}")

            data = query_tool.query_province_data(
                start_date=start_date,
                end_date=end_date,
                ns_type="NS"
            )

            if return_all_data:
                # 返回所有污染物的完整数据
                all_pollutants_data = {}
                field_map_all = {
                    "PM2.5": "PM2_5",
                    "PM10": "PM10",
                    "NO2": "NO2",
                    "O3": "O3_8h",
                    "AQI": "AQIStandardRate"
                }

                for item in data:
                    area_name = item.get("AreaName", "")
                    area_names.append(area_name)

                    # 提取所有污染物数据
                    all_pollutants_data[area_name] = {}
                    for pollutant_name, field_name in field_map_all.items():
                        value = item.get(field_name, 0)
                        if value is None:
                            value = 0
                        try:
                            all_pollutants_data[area_name][pollutant_name] = float(value)
                        except (ValueError, TypeError):
                            logger.warning("invalid_value", pollutant=pollutant_name, field=field_name, value=value)
                            all_pollutants_data[area_name][pollutant_name] = 0.0

                logger.info(
                    "query_with_date_range_success",
                    scope=scope,
                    pollutant="all",
                    start_date=start_date,
                    end_date=end_date,
                    area_count=len(area_names),
                    return_all_data=True
                )

                # 返回特殊格式：所有污染物数据
                return area_names, {
                    "area_names": area_names,
                    "all_data": all_pollutants_data
                }
            else:
                # 原有逻辑：返回单个污染物数据
                for item in data:
                    area_name = item.get("AreaName", "")
                    area_names.append(area_name)

                    value = item.get(field, 0)
                    if value is None:
                        value = 0
                    try:
                        result.append(float(value))
                    except (ValueError, TypeError):
                        logger.warning("invalid_value", pollutant=pollutant, field=field, value=value)
                        result.append(0.0)

        elif scope == "provincial":
            if pollutant not in CITY_STANDARD_FIELD_ALIASES:
                raise ValueError(f"Unknown pollutant: {pollutant}")

            records = await self._query_city_standard_records(start_date, end_date)
            city_stats = self._records_by_city_name(records)

            if return_all_data:
                all_pollutants_data = {}
                for city in GUANGDONG_CITIES:
                    area_names.append(city)
                    city_data = city_stats.get(city)
                    if not city_data:
                        logger.warning("city_data_missing", city=city)
                        all_pollutants_data[city] = {name: 0.0 for name in ("PM2.5", "PM10", "NO2", "O3", "AQI")}
                        continue

                    all_pollutants_data[city] = {
                        pollutant_name: self._get_city_standard_value(city_data, pollutant_name)
                        for pollutant_name in ("PM2.5", "PM10", "NO2", "O3", "AQI")
                    }

                logger.info(
                    "query_with_date_range_success",
                    scope=scope,
                    pollutant="all",
                    start_date=start_date,
                    end_date=end_date,
                    area_count=len(area_names),
                    return_all_data=True
                )

                return area_names, {
                    "area_names": area_names,
                    "all_data": all_pollutants_data
                }
            else:
                for city in GUANGDONG_CITIES:
                    area_names.append(city)
                    city_data = city_stats.get(city)
                    if not city_data:
                        logger.warning("city_data_missing", city=city)
                        result.append(0.0)
                        continue
                    result.append(self._get_city_standard_value(city_data, pollutant))
        else:
            raise ValueError(f"Unknown scope: {scope}")

        logger.info(
            "query_with_date_range_success",
            scope=scope,
            pollutant=pollutant,
            start_date=start_date,
            end_date=end_date,
            area_count=len(area_names),
        )

        return area_names, result

    async def _get_guangdong_province_data(
        self,
        pollutant: str,
        current_start: str,
        current_end: str,
        last_year_start: str,
        last_year_end: str,
        return_all_data: bool = False
    ) -> Dict[str, float]:
        """
        获取广东省的全省数据（用于替换全国sheet中的广东数据）

        使用 query_city_standard_yoy_report 查询广东省审核数据，
        直接使用API返回的"全省"记录（不进行聚合计算）。

        Args:
            pollutant: 污染物名称（PM2.5、PM10、NO2、O3、AQI）
            current_start: 今年开始日期
            current_end: 今年结束日期
            last_year_start: 去年开始日期
            last_year_end: 去年结束日期
            return_all_data: 是否返回所有污染物数据

        Returns:
            如果 return_all_data=False: {"current": 今年数值, "last_year": 去年数值}
            如果 return_all_data=True: {"current": {所有污染物今年值}, "last_year": {所有污染物去年值}}
        """
        try:
            logger.info(
                "query_guangdong_yoy_start",
                pollutant=pollutant,
                current_start=current_start,
                current_end=current_end,
                last_year_start=last_year_start,
                last_year_end=last_year_end,
            )
            records = await self._query_city_standard_yoy_records(
                current_start,
                current_end,
                last_year_start,
                last_year_end,
            )

            # 筛选出"全省"记录
            province_records = [r for r in records if r.get("cityName") == "全省"]

            if not province_records:
                logger.error(
                    "province_record_not_found",
                    message="API返回数据中未找到'全省'记录",
                    total_records=len(records),
                    city_names=[r.get("cityName") for r in records[:10]]
                )
                raise ValueError(
                    f"API返回数据中未找到'全省'记录。"
                    f"返回记录数: {len(records)}，"
                    f"城市名称: {[r.get('cityName') for r in records[:10]]}"
                )

            province_record = province_records[0]
            logger.info(
                "province_record_found",
                city_name=province_record.get("cityName"),
                time_point=province_record.get("timePoint")
            )

            if return_all_data:
                # 直接使用"全省"记录的所有污染物数据
                all_current = {
                    "PM2.5": self._get_city_standard_value(province_record, "PM2.5"),
                    "PM10": self._get_city_standard_value(province_record, "PM10"),
                    "NO2": self._get_city_standard_value(province_record, "NO2"),
                    "O3": self._get_city_standard_value(province_record, "O3"),
                    "AQI": self._get_city_standard_value(province_record, "AQI"),
                }
                all_last_year = {
                    "PM2.5": self._get_city_standard_value(province_record, "PM2.5", suffix="_Compare"),
                    "PM10": self._get_city_standard_value(province_record, "PM10", suffix="_Compare"),
                    "NO2": self._get_city_standard_value(province_record, "NO2", suffix="_Compare"),
                    "O3": self._get_city_standard_value(province_record, "O3", suffix="_Compare"),
                    "AQI": self._get_city_standard_value(province_record, "AQI", suffix="_Compare"),
                }

                logger.info(
                    "guangdong_all_pollutants_retrieved",
                    source="province_record",
                    current_city_count=1,
                    pollutants_count=len(all_current)
                )

                return {
                    "current": all_current,
                    "last_year": all_last_year
                }
            else:
                if pollutant not in CITY_STANDARD_FIELD_ALIASES:
                    logger.warning("unknown_pollutant_for_guangdong", pollutant=pollutant)
                    return None

                current_value = self._get_city_standard_value(province_record, pollutant)
                last_year_value = self._get_city_standard_value(province_record, pollutant, suffix="_Compare")

                logger.info(
                    "pollutant_value_from_province_record",
                    pollutant=pollutant,
                    current_value=current_value,
                    last_year_value=last_year_value
                )

                return {
                    "current": float(current_value),
                    "last_year": float(last_year_value)
                }

        except Exception as e:
            logger.error("guangdong_data_query_failed", pollutant=pollutant, error=str(e), exc_info=True)
            return None

    async def _fill_historical_comparison_sheet(self, wb, time_range: Dict[str, str], sheet_name: str, full_month: bool = False):
        """
        填充历年当月浓度sheet（2014年至当前年，每年目标月份一行）

        功能：
        1. 从2014年开始，逐年查询目标月份的全省数据
        2. 填充7个指标：AQI达标率、PM2.5、PM10、NO2、O3、SO2、CO
        3. 数据来源：广东省城市统计报表接口（query_city_standard_report）

        示例：
        - 当前为2026年4月完整月报，则查询：
          - 2014年4月1日-4月30日
          - ...
          - 2026年4月1日-4月30日

        Args:
            wb: 工作簿对象
            time_range: 时间范围字典
            sheet_name: sheet名称（动态生成，如"5月浓度"）
        """
        ws = wb[sheet_name]
        config = EXTRA_SHEET_CONFIG["历年当月浓度"]
        start_year = config["start_year"]
        start_row = config["start_row"]
        year_col = config["year_col"]
        standard_col = config.get("standard_col")
        standard_header = config.get("standard_header", "标准类型")

        # 当前年份和月份
        current_year = int(time_range["year"])
        start_month = int(str(time_range["start_date"]).split("-")[1])
        start_day = int(str(time_range["start_date"]).split("-")[2])
        current_month = int(time_range["month"])
        target_end_day = int(str(time_range["end_date"]).split("-")[-1])

        historical_years = list(range(start_year, current_year + 1))
        required_rows = start_row + len(historical_years) - 1
        if ws.max_row < required_rows:
            for _ in range(required_rows - ws.max_row):
                ws.append([None] * ws.max_column)

        if standard_col:
            from openpyxl.styles import Alignment, Font, PatternFill

            header_row = start_row - 1
            standard_header_cell = ws[f"{standard_col}{header_row}"]
            source_header_cell = ws[f"{config['columns'][-1]['col']}{header_row}"]
            standard_header_cell.value = standard_header
            if source_header_cell.has_style:
                standard_header_cell._style = copy(source_header_cell._style)
            standard_header_cell.font = copy(source_header_cell.font) if source_header_cell.has_style else Font(bold=True)
            standard_header_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[standard_col].width = max(ws.column_dimensions[standard_col].width or 0, 12)
            old_standard_fill = PatternFill("solid", fgColor="FFF2CC")
            new_standard_fill = PatternFill("solid", fgColor="D9EAD3")
        else:
            old_standard_fill = None
            new_standard_fill = None

        logger.info(
            "historical_comparison_pre_query",
            start_year=start_year,
            current_year=current_year,
            month=current_month,
            years_count=len(historical_years),
            full_month=full_month,
        )

        from calendar import monthrange

        def build_year_range(year: int) -> Tuple[str, str]:
            """构造某一年目标月份的查询区间。"""
            if year == current_year:
                return time_range["start_date"], time_range["end_date"]

            month_last_day = monthrange(year, current_month)[1]
            if full_month:
                end_day = month_last_day
            else:
                end_day = min(target_end_day, month_last_day)

            return (
                f"{year}-{start_month:02d}-{start_day:02d}",
                f"{year}-{current_month:02d}-{end_day:02d}",
            )

        # 清理并重写年份区域，避免模板旧数据残留或目标年份变化造成错位。
        for row in range(start_row, required_rows + 1):
            ws[f"{year_col}{row}"] = None
            for col_config in config["columns"]:
                ws[f"{col_config['col']}{row}"] = None
            if standard_col:
                ws[f"{standard_col}{row}"] = None

        for year in historical_years:
            row = start_row + (year - start_year)
            query_start, query_end = build_year_range(year)
            ws[f"{year_col}{row}"] = year
            standard_label = self._historical_standard_label(year)
            if standard_col:
                standard_cell = ws[f"{standard_col}{row}"]
                standard_cell.value = standard_label
                standard_cell.alignment = Alignment(horizontal="center", vertical="center")
                standard_cell.fill = new_standard_fill if standard_label == "新标准" else old_standard_fill

            try:
                records = await self._query_city_standard_records(
                    query_start,
                    query_end,
                )

                province_records = [r for r in records if r.get("cityName") == "全省"]
                if not province_records:
                    logger.error(
                        "province_record_not_found_in_historical",
                        message="历年当月浓度sheet：API返回数据中未找到'全省'记录",
                        year=year,
                        start_date=query_start,
                        end_date=query_end,
                        total_records=len(records),
                        city_names=[r.get("cityName") for r in records[:10]],
                    )
                    raise ValueError(
                        f"历年当月浓度sheet：{year}年API返回数据中未找到'全省'记录。"
                        f"返回记录数: {len(records)}，"
                        f"城市名称: {[r.get('cityName') for r in records[:10]]}"
                    )

                province_record = province_records[0]
                for col_config in config["columns"]:
                    col = col_config["col"]
                    field = col_config["field"]
                    ws[f"{col}{row}"] = self._get_city_standard_value(province_record, field)

                logger.info(
                    "historical_year_filled",
                    year=year,
                    row=row,
                    start_date=query_start,
                    end_date=query_end,
                )

            except Exception as e:
                logger.error(
                    "historical_year_fill_failed",
                    year=year,
                    row=row,
                    start_date=query_start,
                    end_date=query_end,
                    error=str(e),
                    exc_info=True,
                )
                for col_config in config["columns"]:
                    ws[f"{col_config['col']}{row}"] = 0

        logger.info(
            "historical_comparison_filled",
            start_year=start_year,
            current_year=current_year,
            month=current_month,
            years_count=len(historical_years),
            period=f"{current_month}月",
        )

    def _is_leap_year(self, year: int) -> bool:
        """
        判断是否为闰年

        Args:
            year: 年份

        Returns:
            True表示闰年，False表示平年
        """
        if year % 4 != 0:
            return False
        elif year % 100 != 0:
            return True
        else:
            return year % 400 == 0


# 导出
__all__ = ["ConsultationFileFetcher", "MonthlyConsultationFileFetcher"]

# 导入月度完整会商文件Fetcher
from app.fetchers.consultation.monthly import MonthlyConsultationFileFetcher
