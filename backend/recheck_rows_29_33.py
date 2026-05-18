#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
重新检查第29行和第33行的省份
"""

import openpyxl
from pathlib import Path

def recheck_rows_29_33():
    """重新检查第29行和第33行"""

    excel_file = Path("/tmp/会商文件/2026年5月/月度会商模板（2026年5月）20260514.xlsx")

    if not excel_file.exists():
        print(f"❌ 文件不存在: {excel_file}")
        return

    wb = openpyxl.load_workbook(str(excel_file))
    ws = wb["5月全国排名"]

    print("=" * 80)
    print("🔍 重新检查第29行和第33行")
    print("=" * 80)

    # 检查第29行和第33行的所有污染物
    rows_to_check = [29, 33]

    for row in rows_to_check:
        print(f"\n第{row}行数据:")
        for col_idx in range(1, 16, 2):  # A, C, E, G, I, K, M列（省份列）
            col_letter = chr(ord('A') + col_idx - 1)
            province = ws.cell(row=row, column=col_idx).value
            if province:
                value_col = chr(ord('A') + col_idx)  # 下一列是数值列
                value = ws.cell(row=row, column=col_idx + 1).value
                rank_col = chr(ord('A') + col_idx + 1)  # 下下1列是排名列
                rank = ws.cell(row=row, column=col_idx + 2).value
                print(f"  {col_letter}列 (省份): {province}")
                print(f"  {value_col}列 (数值): {value}")
                print(f"  {rank_col}列 (排名): {rank}")

    # 检查A列（PM2.5省份）第29行和第33行
    print(f"\n{'='*80}")
    print("A列（PM2.5省份）详细检查:")
    print("="*80)

    for row in [29, 33]:
        province = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        rank = ws.cell(row=row, column=3).value
        print(f"第{row}行: {province} = {value} (排名: {rank})")

    wb.close()

if __name__ == "__main__":
    recheck_rows_29_33()
