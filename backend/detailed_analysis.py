#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更详细分析全国排名sheet的问题
"""

import openpyxl
from pathlib import Path

def detailed_analysis():
    """更详细分析全国排名sheet"""

    excel_file = Path("/tmp/会商文件/2026年5月/月度会商模板（2026年5月）20260514.xlsx")

    if not excel_file.exists():
        print(f"❌ 文件不存在: {excel_file}")
        return

    wb = openpyxl.load_workbook(str(excel_file))
    ws = wb["5月全国排名"]

    print("=" * 80)
    print("🔍 详细数据分析")
    print("=" * 80)

    # 检查每个污染物的省份列
    pollutants = [
        {"name": "PM2.5", "name_col": "A", "value_col": "B", "rank_col": "C"},
        {"name": "PM10", "name_col": "D", "value_col": "E", "rank_col": "F"},
        {"name": "NO2", "name_col": "G", "value_col": "H", "rank_col": "I"},
        {"name": "O3", "name_col": "J", "value_col": "K", "rank_col": "L"},
        {"name": "AQI", "name_col": "M", "value_col": "N", "rank_col": "O"},
    ]

    for pollutant in pollutants:
        print(f"\n{'='*20} {pollutant['name']} {'='*20}")

        name_col = pollutant["name_col"]
        value_col = pollutant["value_col"]
        rank_col = pollutant["rank_col"]

        name_col_idx = ord(name_col) - ord('A') + 1
        value_col_idx = ord(value_col) - ord('A') + 1
        rank_col_idx = ord(rank_col) - ord('A') + 1

        # 收集所有行的数据
        rows_with_data = []
        for row in range(3, 40):
            province = ws.cell(row=row, column=name_col_idx).value
            if province and str(province).strip():
                value = ws.cell(row=row, column=value_col_idx).value
                rank = ws.cell(row=row, column=rank_col_idx).value
                rows_with_data.append({
                    "row": row,
                    "province": str(province).strip(),
                    "value": value,
                    "rank": rank
                })

        print(f"数据行数: {len(rows_with_data)}")

        # 检查是否有第33行及以后的数据
        overflow = [r for r in rows_with_data if r["row"] > 32]
        if overflow:
            print(f"⚠️  超出第32行的数据:")
            for r in overflow:
                print(f"  第{r['row']}行: {r['province']} = {r['value']} (排名: {r['rank']})")

        # 检查省份重复
        province_positions = {}
        for r in rows_with_data:
            prov = r["province"]
            if prov not in province_positions:
                province_positions[prov] = []
            province_positions[prov].append(r["row"])

        duplicates = {prov: rows for prov, rows in province_positions.items() if len(rows) > 1}
        if duplicates:
            print(f"❌ 重复省份:")
            for prov, rows in duplicates.items():
                print(f"  '{prov}' 出现在: {rows}")
        else:
            print(f"✅ 无重复省份")

    # 检查是否有其他sheet影响了这个sheet
    print(f"\n{'='*20} 检查模板问题 {'='*20}")
    print(f"所有sheets: {wb.sheetnames}")

    wb.close()

if __name__ == "__main__":
    detailed_analysis()
