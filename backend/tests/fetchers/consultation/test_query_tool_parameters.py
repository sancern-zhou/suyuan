import inspect

from openpyxl import Workbook
import pytest

from app.fetchers.consultation import ConsultationFileFetcher
from app.fetchers.consultation.annual_ytd import AnnualYtdConsultationFileFetcher
from app.tools.query.query_city_standard_report.tool import _default_ns_type_for_range


@pytest.mark.asyncio
async def test_city_standard_records_use_current_tool_signature(monkeypatch):
    captured = {}

    async def fake_execute_query_city_standard_report(
        *,
        cities=None,
        start_time,
        end_time,
        ns_type=None,
        pollutant_codes=None,
        data_source=1,
        sand_type=1,
        context=None,
    ):
        captured.update(
            {
                "cities": cities,
                "start_time": start_time,
                "end_time": end_time,
                "ns_type": ns_type,
                "pollutant_codes": pollutant_codes,
                "data_source": data_source,
                "sand_type": sand_type,
                "context": context,
            }
        )
        return {"success": True, "data": [{"cityName": "全省"}]}

    monkeypatch.setattr(
        "app.tools.query.query_city_standard_report.tool.execute_query_city_standard_report",
        fake_execute_query_city_standard_report,
    )

    fetcher = object.__new__(ConsultationFileFetcher)
    records = await fetcher._query_city_standard_records("2014-01-01", "2014-01-31")

    assert records == [{"cityName": "全省"}]
    assert captured["start_time"] == "2014-01-01"
    assert captured["end_time"] == "2014-01-31"
    assert captured["ns_type"] is None


@pytest.mark.asyncio
async def test_city_standard_yoy_records_use_current_tool_signature(monkeypatch):
    captured = {}

    async def fake_execute_query_city_standard_yoy_report(
        *,
        cities=None,
        time_point,
        contrast_time,
        ns_type=2,
        pollutant_codes=None,
        data_source=1,
        sand_type=1,
        context=None,
    ):
        captured.update(
            {
                "cities": cities,
                "time_point": time_point,
                "contrast_time": contrast_time,
                "ns_type": ns_type,
                "pollutant_codes": pollutant_codes,
                "data_source": data_source,
                "sand_type": sand_type,
                "context": context,
            }
        )
        return {"success": True, "data": [{"cityName": "全省"}]}

    monkeypatch.setattr(
        "app.tools.query.query_city_standard_report.tool.execute_query_city_standard_yoy_report",
        fake_execute_query_city_standard_yoy_report,
    )

    fetcher = object.__new__(ConsultationFileFetcher)
    records = await fetcher._query_city_standard_yoy_records(
        "2026-01-01",
        "2026-01-31",
        "2025-01-01",
        "2025-01-31",
    )

    assert records == [{"cityName": "全省"}]
    assert captured["time_point"] == ["2026-01-01", "2026-01-31"]
    assert captured["contrast_time"] == ["2025-01-01", "2025-01-31"]
    assert captured["ns_type"] == 2


def test_historical_comparison_uses_auto_standard_selection():
    source = inspect.getsource(ConsultationFileFetcher._fill_historical_comparison_sheet)

    assert "ns_type=2" not in source


def test_tool_auto_standard_selection_boundary_for_historical_years():
    assert _default_ns_type_for_range("2014-01-01", "2014-12-31") == 1
    assert _default_ns_type_for_range("2015-01-01", "2015-12-31") == 1
    assert _default_ns_type_for_range("2016-01-01", "2016-12-31") == 1
    assert _default_ns_type_for_range("2025-01-01", "2025-12-31") == 2


def test_historical_standard_label_matches_tool_boundary():
    assert ConsultationFileFetcher._historical_standard_label(2014) == "旧标准"
    assert ConsultationFileFetcher._historical_standard_label(2016) == "旧标准"
    assert ConsultationFileFetcher._historical_standard_label(2025) == "新标准"


def test_city_standard_value_does_not_fallback_to_alternate_fields():
    fetcher = object.__new__(ConsultationFileFetcher)
    record = {
        "cityName": "全省",
        "pM2_5": "21",
        "PM2_5_Decimal": "21.1",
        "pM10": "35",
        "PM10_Decimal": "34.5",
        "NO2": "11",
        "O3_8h": "122",
        "FineRate": "98",
    }

    assert fetcher._get_city_standard_value(record, "PM2.5") == 0.0
    assert fetcher._get_city_standard_value(record, "PM10") == 0.0
    assert fetcher._get_city_standard_value(record, "NO2") == 0.0
    assert fetcher._get_city_standard_value(record, "O3") == 0.0
    assert fetcher._get_city_standard_value(record, "AQI") == 0.0


@pytest.mark.asyncio
async def test_national_ranking_writes_aqi_rate_as_excel_percent_value():
    fetcher = object.__new__(ConsultationFileFetcher)
    time_range = {
        "start_date": "2026-05-01",
        "end_date": "2026-05-31",
        "period_description": "2026年05月份",
        "year": "2026",
        "month": "05",
        "last_year": "2025",
    }
    last_year_start = fetcher._get_last_year_start_date(time_range)
    last_year_end = fetcher._get_last_year_same_day(time_range, full_month=True)
    cache_key = fetcher._period_cache_key(time_range, last_year_start, last_year_end)
    fetcher._consultation_period_cache = {
        "national": {
            "key": cache_key,
            "current": {
                "area_names": ["广东", "浙江"],
                "all_data": {
                    "广东": {"PM2.5": 18.0, "PM10": 30.5, "NO2": 10.0, "O3": 120.0, "AQI": 100.0},
                    "浙江": {"PM2.5": 20.0, "PM10": 33.0, "NO2": 12.0, "O3": 128.0, "AQI": 98.5},
                },
            },
            "last_year": {
                "area_names": ["广东", "浙江"],
                "all_data": {
                    "广东": {"PM2.5": 19.0, "PM10": 31.0, "NO2": 11.0, "O3": 122.0, "AQI": 99.0},
                    "浙江": {"PM2.5": 21.0, "PM10": 34.0, "NO2": 13.0, "O3": 130.0, "AQI": 97.0},
                },
            },
        }
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "5月全国排名"
    ws["N3"].number_format = "0.0%"

    await fetcher._fill_national_ranking_sheet(wb, time_range, "5月全国排名", full_month=True)

    assert ws["N3"].value == 1.0
    assert ws["E3"].value == 30.5


@pytest.mark.asyncio
async def test_national_ranking_formats_pm25_and_pm10_with_one_decimal_place():
    fetcher = object.__new__(ConsultationFileFetcher)
    time_range = {
        "start_date": "2026-05-01",
        "end_date": "2026-05-31",
        "period_description": "2026年05月份",
        "year": "2026",
        "month": "05",
        "last_year": "2025",
    }
    last_year_start = fetcher._get_last_year_start_date(time_range)
    last_year_end = fetcher._get_last_year_same_day(time_range, full_month=True)
    fetcher._consultation_period_cache = {
        "national": {
            "key": fetcher._period_cache_key(time_range, last_year_start, last_year_end),
            "current": {
                "area_names": ["广东"],
                "all_data": {
                    "广东": {
                        "PM2.5": 12.0,
                        "PM10": 22.0,
                        "NO2": 10.0,
                        "O3": 120.0,
                        "AQI": 100.0,
                    },
                },
            },
            "last_year": {
                "area_names": ["广东"],
                "all_data": {
                    "广东": {
                        "PM2.5": 13.0,
                        "PM10": 23.0,
                        "NO2": 11.0,
                        "O3": 122.0,
                        "AQI": 99.0,
                    },
                },
            },
        }
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "5月全国排名"
    ws["B3"].number_format = "General"
    ws["E3"].number_format = "General"

    await fetcher._fill_national_ranking_sheet(wb, time_range, "5月全国排名", full_month=True)

    assert ws["B3"].value == 12.0
    assert ws["B3"].number_format == "0.0"
    assert ws["E3"].value == 22.0
    assert ws["E3"].number_format == "0.0"


@pytest.mark.asyncio
async def test_guangdong_all_pollutants_prefers_single_period_decimal_values():
    fetcher = object.__new__(ConsultationFileFetcher)

    async def fake_yoy_records(*args, **kwargs):
        raise AssertionError("return_all_data must not use the YoY API")

    async def fake_standard_records(start_date, end_date, **kwargs):
        if start_date == "2026-01-01":
            return [
                {
                    "cityName": "全省",
                    "pM2_5_Decimal": "21.0",
                    "pM2_5": "21",
                    "pM10_Decimal": None,
                    "pM10": "35",
                    "fineRate": "98.0",
                }
            ]
        return [
            {
                "cityName": "全省",
                "pM2_5_Decimal": "22.7",
                "pM2_5": "23",
                "pM10_Decimal": "37.6",
                "pM10": "38",
                "fineRate": "97.0",
            }
        ]

    fetcher._query_city_standard_yoy_records = fake_yoy_records
    fetcher._query_city_standard_records = fake_standard_records

    data = await fetcher._get_guangdong_province_data(
        pollutant="PM2.5",
        current_start="2026-01-01",
        current_end="2026-06-30",
        last_year_start="2025-01-01",
        last_year_end="2025-06-30",
        return_all_data=True,
    )

    assert data["current"]["PM2.5"] == 21.0
    assert data["current"]["PM10"] == 0.0
    assert data["last_year"]["PM2.5"] == 22.7
    assert data["last_year"]["PM10"] == 37.6


@pytest.mark.asyncio
async def test_pollutant_sheet_cache_orders_values_descending():
    fetcher = object.__new__(ConsultationFileFetcher)
    wb = Workbook()
    ws = wb.active
    ws.title = "全省PM10"
    config = {
        "scope": "provincial",
        "pollutant": "PM10",
        "data_rows": (2, 4),
        "name_col": "A",
        "current_col": "B",
        "last_year_col": "D",
        "sort_copies": [
            {
                "source_name_col": "A",
                "source_value_col": "D",
                "target_name_col": "W",
                "target_value_col": "X",
                "sort_ascending": False,
            }
        ],
    }
    provincial_cache = {
        "current": {
            "area_names": ["广州", "深圳", "珠海"],
            "all_data": {
                "广州": {"PM10": 40.0},
                "深圳": {"PM10": 55.0},
                "珠海": {"PM10": 35.0},
            },
        },
        "last_year": {
            "area_names": ["广州", "深圳", "珠海"],
            "all_data": {
                "广州": {"PM10": 46.0},
                "深圳": {"PM10": 42.0},
                "珠海": {"PM10": 60.0},
            },
        },
    }

    await fetcher._fill_single_sheet_with_cache(
        wb,
        "全省PM10",
        config,
        {},
        provincial_cache=provincial_cache,
    )

    assert [ws[f"A{row}"].value for row in range(2, 5)] == ["深圳", "广州", "珠海"]
    assert [ws[f"B{row}"].value for row in range(2, 5)] == [55.0, 40.0, 35.0]
    assert [ws[f"W{row}"].value for row in range(2, 5)] == ["珠海", "广州", "深圳"]
    assert [ws[f"X{row}"].value for row in range(2, 5)] == [60.0, 46.0, 42.0]


@pytest.mark.asyncio
async def test_annual_ytd_provincial_summary_headers_use_cumulative_period():
    fetcher = AnnualYtdConsultationFileFetcher()

    async def fake_yoy_records(*args, ns_type=2, **kwargs):
        return [
            {
                "cityName": "全省",
                "pM2_5_Decimal": 15.6,
                "pM2_5_Decimal_Compare": 16.9,
                "pM10_Decimal": 26,
                "pM10_Decimal_Compare": 29,
                "nO2": 12,
                "nO2_Compare": 15,
                "o3_8h": 135,
                "o3_8h_Compare": 143,
                "fineRate": 95.5,
                "fineRate_Compare": 94.5,
            }
        ]

    async def fake_standard_records(start_date, end_date, **kwargs):
        if start_date == "2026-01-01":
            return [
                {
                    "cityName": "全省",
                    "pM2_5_Decimal": 15.6,
                    "pM10_Decimal": 26.4,
                    "nO2": 12,
                    "o3_8h": 135,
                    "fineRate": 95.5,
                }
            ]
        return [
            {
                "cityName": "全省",
                "pM2_5_Decimal": 16.9,
                "pM10_Decimal": 29.3,
                "nO2": 15,
                "o3_8h": 143,
                "fineRate": 94.5,
            }
        ]

    fetcher._query_city_standard_yoy_records = fake_yoy_records
    fetcher._query_city_standard_records = fake_standard_records

    wb = Workbook()
    wb.active.title = "全省同比"
    time_range = {
        "start_date": "2026-01-01",
        "end_date": "2026-05-31",
        "period_description": "2026年1-5月累计",
        "year": "2026",
        "month": "05",
        "last_year": "2025",
    }

    await fetcher._fill_provincial_summary_sheet(wb, time_range, full_month=True)

    ws = wb["全省同比"]
    assert ws["B2"].value == "2025年1-5月"
    assert ws["C2"].value == "2026年1-5月"
    assert ws["B4"].value == 29.3
    assert ws["C4"].value == 26.4
