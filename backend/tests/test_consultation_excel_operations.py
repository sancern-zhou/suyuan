from pathlib import Path

import openpyxl

from app.tools.assistant.excel_operations import ConsultationExcelOperator


def test_update_consultation_file_by_name_keeps_guangdong_unique(tmp_path: Path):
    file_path = tmp_path / "consultation.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    provinces = ["浙江", "河南", "广东", "北京"]
    for index, province in enumerate(provinces, start=2):
        ws[f"A{index}"] = province

    wb.save(file_path)

    current_records = [
        {"name": "河南", "value": 137},
        {"name": "广东省", "value": 128},
        {"name": "浙江", "value": 120},
        {"name": "北京市", "value": 90},
    ]
    last_year_records = [
        {"name": "北京", "value": 80},
        {"name": "浙江省", "value": 118},
        {"name": "广东", "value": 130},
        {"name": "河南省", "value": 125},
    ]

    operator = ConsultationExcelOperator(str(file_path))
    operator.load_file()
    operator.update_consultation_file_by_name(
        current_records=current_records,
        last_year_records=last_year_records,
        current_period="2026年4月份",
        last_year_period="2025年4月份",
        data_start_row=2,
        data_end_row=5,
    )
    operator.save_file()

    result = openpyxl.load_workbook(file_path, data_only=True)
    ws = result.active

    assert ws["B4"].value == 128
    sorted_current_names = [ws[f"G{row}"].value for row in range(2, 6)]
    sorted_current_values = [ws[f"H{row}"].value for row in range(2, 6)]

    assert sorted_current_names.count("广东") == 1
    assert sorted_current_values[sorted_current_names.index("广东")] == 130

    diff_names = [ws[f"K{row}"].value for row in range(2, 6)]
    assert diff_names.count("广东") == 1


def test_update_consultation_file_by_name_rejects_duplicate_api_names(tmp_path: Path):
    file_path = tmp_path / "consultation.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A2"] = "广东"
    wb.save(file_path)

    operator = ConsultationExcelOperator(str(file_path))
    operator.load_file()

    duplicate_records = [
        {"name": "广东", "value": 128},
        {"name": "广东省", "value": 137},
    ]

    try:
        operator.align_records_to_names(duplicate_records, start_row=2, end_row=2)
    except ValueError as exc:
        assert "重复地区" in str(exc)
    else:
        raise AssertionError("duplicate area names should be rejected")
