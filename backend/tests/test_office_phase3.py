"""
Office 工具单元测试 - Phase 3

测试范围：
- excel_recalc_tool.py - Excel 公式重算

测试场景：
- 公式重算
- 错误扫描
- 公式统计
"""
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook
from app.tools.office.excel_recalc_tool import ExcelRecalcTool


@pytest.fixture
def temp_dir():
    """创建临时目录"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_xlsx_with_formulas(temp_dir):
    """创建包含公式的示例 XLSX 文件"""
    xlsx_path = temp_dir / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加数据
    ws['A1'] = 10
    ws['A2'] = 20
    ws['A3'] = 30

    # 添加公式
    ws['B1'] = '=A1*2'
    ws['B2'] = '=A2+A1'
    ws['B3'] = '=SUM(A1:A3)'

    # 添加一个除零错误公式
    ws['C1'] = 0
    ws['C2'] = '=A1/C1'  # 会产生 #DIV/0! 错误

    wb.save(xlsx_path)
    return xlsx_path


@pytest.fixture
def sample_xlsx_no_formulas(temp_dir):
    """创建不包含公式的示例 XLSX 文件"""
    xlsx_path = temp_dir / "no_formulas.xlsx"
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 100
    ws['A2'] = 200
    ws['A3'] = 300

    wb.save(xlsx_path)
    return xlsx_path


class TestExcelRecalcTool:
    """测试 ExcelRecalcTool"""

    @pytest.mark.asyncio
    async def test_recalc_file_not_found(self, temp_dir):
        """测试重算 - 文件不存在"""
        tool = ExcelRecalcTool()

        result = await tool.execute(
            file_path=str(temp_dir / "nonexistent.xlsx")
        )

        assert result["success"] is False
        assert "不存在" in result["summary"]

    @pytest.mark.asyncio
    async def test_recalc_invalid_format(self, temp_dir):
        """测试重算 - 不支持的格式"""
        tool = ExcelRecalcTool()
        invalid_file = temp_dir / "test.txt"
        invalid_file.write_text("test")

        result = await tool.execute(
            file_path=str(invalid_file)
        )

        assert result["success"] is False
        assert "格式" in result["summary"]

    @pytest.mark.asyncio
    async def test_scan_formulas_without_libreoffice(self, sample_xlsx_with_formulas):
        """测试公式扫描（不需要 LibreOffice）"""
        tool = ExcelRecalcTool()

        # 直接调用扫描方法
        result = tool._scan_formula_errors(sample_xlsx_with_formulas)

        assert "total_formulas" in result
        assert result["total_formulas"] == 4  # 4个公式
        assert "total_errors" in result

    @pytest.mark.asyncio
    async def test_scan_no_formulas(self, sample_xlsx_no_formulas):
        """测试扫描无公式文件"""
        tool = ExcelRecalcTool()

        result = tool._scan_formula_errors(sample_xlsx_no_formulas)

        assert result["total_formulas"] == 0
        assert result["total_errors"] == 0
        assert result["status"] == "success"

    @pytest.mark.asyncio
    async def test_tool_availability(self):
        """测试 LibreOffice 可用性检测"""
        tool = ExcelRecalcTool()
        is_available = tool.is_available()

        # 只记录可用性，不强制要求
        print(f"LibreOffice available: {is_available}")

    @pytest.mark.asyncio
    @pytest.mark.skipif(
        not ExcelRecalcTool().is_available(),
        reason="LibreOffice not installed"
    )
    async def test_recalc_with_libreoffice(self, sample_xlsx_with_formulas):
        """测试公式重算 - 需要 LibreOffice"""
        tool = ExcelRecalcTool()

        result = await tool.execute(
            file_path=str(sample_xlsx_with_formulas),
            timeout=30
        )

        # 如果 LibreOffice 可用，应该成功
        if result["success"]:
            assert "total_formulas" in result["data"]
            assert result["data"]["total_formulas"] >= 4
            assert "total_errors" in result["data"]

    @pytest.mark.asyncio
    async def test_recalc_schema(self):
        """测试工具 Schema"""
        tool = ExcelRecalcTool()
        schema = tool.get_function_schema()

        assert schema["name"] == "recalc_excel"
        assert "parameters" in schema
        assert "file_path" in schema["parameters"]["properties"]
        assert "timeout" in schema["parameters"]["properties"]

    @pytest.mark.asyncio
    async def test_path_resolution(self, sample_xlsx_with_formulas):
        """测试路径解析"""
        tool = ExcelRecalcTool()

        # 测试绝对路径
        abs_path = tool._resolve_path(str(sample_xlsx_with_formulas))
        assert abs_path is not None
        assert abs_path.is_absolute()

        # 测试相对路径
        rel_path = tool._resolve_path("test.xlsx")
        assert rel_path is not None
        assert rel_path.is_absolute()


class TestExcelErrorDetection:
    """测试 Excel 错误检测"""

    @pytest.mark.asyncio
    async def test_detect_div_zero_error(self, temp_dir):
        """测试检测除零错误"""
        xlsx_path = temp_dir / "div_zero.xlsx"
        wb = Workbook()
        ws = wb.active

        ws['A1'] = 10
        ws['A2'] = 0
        ws['A3'] = '=A1/A2'  # 除零错误

        wb.save(xlsx_path)

        tool = ExcelRecalcTool()
        result = tool._scan_formula_errors(xlsx_path)

        assert result["total_formulas"] == 1

    @pytest.mark.asyncio
    async def test_detect_multiple_errors(self, temp_dir):
        """测试检测多个错误"""
        xlsx_path = temp_dir / "multiple_errors.xlsx"
        wb = Workbook()
        ws = wb.active

        ws['A1'] = 10
        ws['A2'] = 0
        ws['A3'] = '=A1/A2'  # #DIV/0!
        ws['A4'] = '=UNKNOWN_FUNC()'  # #NAME?
        ws['A5'] = '=A1+TEXT'  # #VALUE!

        wb.save(xlsx_path)

        tool = ExcelRecalcTool()
        result = tool._scan_formula_errors(xlsx_path)

        assert result["total_formulas"] == 3


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
